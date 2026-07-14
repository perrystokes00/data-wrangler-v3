"""Build loader test fixtures from the real PPDM training data in well_picks.zip."""
import os, glob
import pandas as pd

SRC = "/home/claude/wp"
OUT = "/home/claude/fixtures"
os.makedirs(OUT, exist_ok=True)

TABLES = ["well_header", "well_picks", "well_log", "well_log_curve", "well_core",
          "well_dir_survey_hdr", "well_dir_survey_data"]
T = {n: pd.read_csv(f"{SRC}/{n}.csv", dtype=str) for n in TABLES}
for n, df in T.items():                       # drop the trailing unnamed col some files have
    T[n] = df.loc[:, ~df.columns.str.startswith("Unnamed")]

# wells with the richest cross-table coverage
FOCUS = ["42-329-10001-0000", "17-069-10005-0000", "42-003-10013-0000", "17-017-10017-0000"]
# a wider slice for the workbook so the loader sees realistic volume
WB_WELLS = list(dict.fromkeys(FOCUS + list(T["well_header"]["UWI"].head(12))))


def hdr(uwi):
    r = T["well_header"][T["well_header"].UWI == uwi]
    return r.iloc[0].to_dict() if len(r) else {}


def rows(name, uwi):
    df = T[name]
    return df[df.UWI == uwi]


# ── 1. multi-sheet Excel workbook ────────────────────────────────────────────
def build_workbook():
    from openpyxl.styles import Font, Alignment, PatternFill
    p = f"{OUT}/well_data_workbook.xlsx"
    with pd.ExcelWriter(p, engine="openpyxl") as w:
        for n in TABLES:
            df = T[n]
            df = df[df.UWI.isin(WB_WELLS)] if "UWI" in df.columns else df
            df.to_excel(w, sheet_name=n[:31], index=False)
        wb = w.book
        head_fill = PatternFill("solid", fgColor="1F3864")
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            for c in ws[1]:
                c.font = Font(name="Arial", bold=True, color="FFFFFF")
                c.fill = head_fill
                c.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                letter = col[0].column_letter
                width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
                ws.column_dimensions[letter].width = min(max(width + 2, 10), 34)
            for row in ws.iter_rows(min_row=2):
                for c in row:
                    c.font = Font(name="Arial")
    print(f"  {os.path.basename(p)}: {len(TABLES)} sheets, {len(WB_WELLS)} wells")
    return p


# ── 2. synthetic LAS files (LAS 2.0, real curves/ranges) ─────────────────────
def build_las(uwi):
    import random
    h = hdr(uwi)
    log = rows("well_log", uwi)
    curves = rows("well_log_curve", uwi)
    if not len(log) or not len(curves):
        return None
    lg = log.iloc[0]
    top, base = float(lg.TOP_DEPTH), float(lg.BASE_DEPTH)
    step = 0.5
    n = min(int((base - top) / step) + 1, 400)          # keep fixtures small
    base = top + step * (n - 1)
    api = uwi.replace("-", "")
    name = f"{OUT}/{h['WELL_NAME'].strip().replace(' ', '_')}_{lg.LOG_ID}.las"

    cv = curves.to_dict("records")
    L = []
    L.append("~Version Information")
    L.append(" VERS.                          2.0 : CWLS LOG ASCII STANDARD - VERSION 2.0")
    L.append(" WRAP.                           NO : ONE LINE PER DEPTH STEP")
    L.append("~Well Information Block")
    L.append("#MNEM.UNIT            DATA                       DESCRIPTION")
    L.append("#---------    -----------------------            -----------------------------")
    L.append(f" STRT .FT     {top:>22.4f} : START DEPTH")
    L.append(f" STOP .FT     {base:>22.4f} : STOP DEPTH")
    L.append(f" STEP .FT     {step:>22.4f} : STEP")
    L.append(f" NULL .       {-999.25:>22} : NULL VALUE")
    L.append(f" COMP .       {h.get('OPERATOR',''):>22} : COMPANY")
    L.append(f" WELL .       {h.get('WELL_NAME',''):>22} : WELL")
    L.append(f" FLD  .       {h.get('FIELD_NAME',''):>22} : FIELD")
    L.append(f" LOC  .       {h.get('COUNTY',''):>22} : LOCATION")
    L.append(f" CNTY .       {h.get('COUNTY',''):>22} : COUNTY")
    L.append(f" STAT .       {h.get('PROVINCE_STATE',''):>22} : STATE")
    L.append(f" CTRY .       {h.get('COUNTRY',''):>22} : COUNTRY")
    L.append(f" SRVC .       {'PPDM_TRAINING':>22} : SERVICE COMPANY")
    L.append(f" DATE .       {str(lg.LOG_DATE):>22} : LOG DATE")
    L.append(f" API  .       {api:>22} : API NUMBER")
    L.append(f" UWI  .       {api:>22} : UNIQUE WELL ID")
    L.append(f" EKB  .FT     {h.get('KB_ELEV',''):>22} : KB ELEVATION")
    L.append(f" EGL  .FT     {h.get('GL_ELEV',''):>22} : GL ELEVATION")
    L.append(f" LATI .DEG    {h.get('SURFACE_LATITUDE',''):>22} : LATITUDE")
    L.append(f" LONG .DEG    {h.get('SURFACE_LONGITUDE',''):>22} : LONGITUDE")
    L.append("~Curve Information Block")
    L.append("#MNEM.UNIT                 API CODE   CURVE DESCRIPTION")
    L.append("#---------               -----------  -----------------")
    L.append(" DEPT .FT                            : DEPTH")
    for c in cv:
        L.append(f" {c['CURVE_NAME']:<4} .{str(c['CURVE_UNIT']):<8}                   : {c['CURVE_NAME']} CURVE")
    L.append("~Parameter Information Block")
    L.append(f" RUN  .       {str(lg.RUN_NO):>22} : RUN NUMBER")
    L.append(f" LTYP .       {str(lg.LOG_TYPE):>22} : LOG TYPE")
    L.append("~Other")
    L.append(f" Synthetic LAS generated from PPDM training data for loader testing.")
    L.append("~ASCII Log Data")
    L.append("#" + "DEPT".rjust(9) + "".join(c["CURVE_NAME"].rjust(11) for c in cv))

    rnd = random.Random(hash(uwi) & 0xFFFF)
    state = {c["CURVE_NAME"]: (float(c["MIN_VALUE"]) + float(c["MAX_VALUE"])) / 2 for c in cv}
    for i in range(n):
        d = top + i * step
        vals = []
        for c in cv:
            lo, hi = float(c["MIN_VALUE"]), float(c["MAX_VALUE"])
            span = (hi - lo) or 1.0
            v = state[c["CURVE_NAME"]] + rnd.uniform(-span * 0.03, span * 0.03)
            v = min(max(v, lo), hi)                      # random walk inside the real range
            state[c["CURVE_NAME"]] = v
            vals.append(f"{v:11.4f}")
        L.append(f"{d:10.4f}" + "".join(vals))
    open(name, "w", newline="\n").write("\n".join(L) + "\n")
    print(f"  {os.path.basename(name)}: {len(cv)} curves, {n} depth steps")
    return name


# ── 3. PDF scout tickets (text layer, real data) ─────────────────────────────
def build_scout(uwi):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    S = getSampleStyleSheet()
    H = ParagraphStyle("sec", parent=S["Heading3"], textColor=colors.HexColor("#2b6b4f"),
                       spaceBefore=10, spaceAfter=4)
    TITLE = ParagraphStyle("t", parent=S["Title"], fontSize=17, textColor=colors.HexColor("#333333"))
    SUB = ParagraphStyle("s", parent=S["Normal"], fontSize=9, textColor=colors.HexColor("#777777"))

    def grid(data, widths, header=True):
        t = Table(data, colWidths=widths, hAlign="LEFT")
        st = [("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#bbbbbb")),
              ("FONTSIZE", (0, 0), (-1, -1), 8),
              ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
              ("LEFTPADDING", (0, 0), (-1, -1), 5)]
        if header:
            st += [("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef3f0")),
                   ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold")]
        t.setStyle(TableStyle(st))
        return t

    h = hdr(uwi)
    api = uwi.replace("-", "")
    p = f"{OUT}/scout_ticket_{h['WELL_NAME'].strip().replace(' ', '_')}.pdf"
    doc = SimpleDocTemplate(p, pagesize=letter, topMargin=0.6 * inch,
                            leftMargin=0.6 * inch, rightMargin=0.6 * inch)
    story = [Paragraph("WELL SCOUT TICKET", TITLE),
             Paragraph(f"DataView &middot; {h['OPERATOR']} &middot; {h['STATUS']}", SUB),
             Spacer(1, 8), Paragraph("Well Header", H)]
    story.append(grid([["API:", api, "Well Name:", h["WELL_NAME"]],
                       ["Well Type:", h["WELL_CLASS"], "Status:", h["STATUS"]],
                       ["Operator:", h["OPERATOR"], "Field:", h["FIELD_NAME"]],
                       ["County:", h["COUNTY"], "State:", h["PROVINCE_STATE"]],
                       ["Spud Date:", h["SPUD_DATE"], "Completion Date:", h["COMPLETION_DATE"]],
                       ["Total Depth MD:", f"{h['DRILLERS_TD']} ft", "KB Elevation:", f"{h['KB_ELEV']} ft"],
                       ["UWI:", api, "Surface Location:",
                        f"{h['SURFACE_LATITUDE']}N {h['SURFACE_LONGITUDE']}W"]],
                      [1.15 * inch, 1.9 * inch, 1.35 * inch, 2.0 * inch], header=False))

    pk = rows("well_picks", uwi)
    if len(pk):
        story.append(Paragraph("Stratigraphy &mdash; Formation Tops", H))
        data = [["Formation", "Top MD (ft)", "Base MD (ft)", "Interp Date", "Interp By"]]
        data += [[r.STRAT_UNIT_ID, r.TOP_MD, r.BASE_MD, r.INTERP_DATE, r.INTERP_BY]
                 for r in pk.itertuples()]
        story.append(grid(data, [1.7 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch, 1.0 * inch]))

    sv = rows("well_dir_survey_data", uwi).head(12)
    if len(sv):
        story.append(Paragraph(f"Directional Survey &mdash; first {len(sv)} stations", H))
        data = [["MD (ft)", "Inc", "Azi", "TVDSS (ft)"]]
        data += [[r.MD, r.INCLINATION, r.AZIMUTH, r.TVDSS] for r in sv.itertuples()]
        story.append(grid(data, [1.3 * inch] * 4))

    co = rows("well_core", uwi)
    if len(co):
        story.append(Paragraph("Core Runs", H))
        data = [["Core ID", "Type", "Top MD (ft)", "Base MD (ft)", "Recovery (%)", "Formation"]]
        data += [[r.CORE_ID, r.CORE_TYPE, r.TOP_DEPTH, r.BASE_DEPTH, r.RECOVERY_PCT, r.FORMATION]
                 for r in co.itertuples()]
        story.append(grid(data, [1.6 * inch, 1.3 * inch, 0.95 * inch, 0.95 * inch, 0.95 * inch, 1.2 * inch]))

    doc.build(story)
    print(f"  {os.path.basename(p)}: header + {len(pk)} picks + {len(sv)} stations + {len(co)} cores")
    return p


if __name__ == "__main__":
    print("Excel workbook:")
    build_workbook()
    print("LAS files:")
    for u in FOCUS[:3]:
        build_las(u)
    print("PDF scout tickets:")
    for u in FOCUS[:2]:
        build_scout(u)
    print("\nfixtures:", sorted(os.path.basename(f) for f in glob.glob(f"{OUT}/*")))
