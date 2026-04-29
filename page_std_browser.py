"""
page_las_catalog.py

Data Wrangler — LAS File Catalog page.

Four tabs:
  📊 Dashboard   — summary stats, top curves
  🔍 Search      — query by UWI, well name, curve, field, depth interval
  📁 Repositories — register and manage storage locations
  ⚙  Catalog     — scan directories into the catalog
"""

from pathlib import Path

import pandas as pd
import streamlit as st

try:
    from modules.segy_catalog import (
        parse_segy_header, catalog_segy_file, catalog_segy_directory,
        get_segy_summary,
    )
    from modules.p190_catalog import (
        parse_p190_header, catalog_p190_file, catalog_p190_directory,
        get_p190_summary,
    )
    from modules.las_catalog import (
        ensure_catalog_schema,
        list_repositories,
        add_repository,
        catalog_file,
        catalog_directory,
        search_catalog,
        get_file_curves,
        get_catalog_summary,
        update_uwi_match,
        well_exists,
        create_well_from_las,
        parse_las_header,
        export_files,
        get_distinct_values,
    )
    from modules.dlis_catalog import (
        catalog_dlis_file,
        catalog_lis_file,
        parse_dlis_header,
        parse_lis_header,
    )
    from modules.las_loader import fetch_ppdm_uwis, fuzzy_match_uwi
    from modules.wl_file_map import (
        scan_directory,
        import_manifest,
        save_map,
        load_map,
        update_uwi,
        confirm_rows,
        catalog_confirmed,
        clear_catalogued,
        skip_rows,
        ensure_map_table,
        preview_rename,
        rename_files,
        crawl_walk,
        crawl_process,
        _now_str as _wl_now_str,
    )
    _AVAILABLE = True
except ImportError as _err:
    _AVAILABLE = False
    _IMPORT_ERROR = str(_err)

_REPO_KEY = "las_catalog_repo_id"


def _get_engine():
    """Get engine from session state — same object used by the main pipeline."""
    engine = st.session_state.get("engine")
    if engine is not None:
        return engine
    try:
        from modules.db_pool import get_engine
        return get_engine()
    except ImportError:
        return None


def run():
    # Breadcrumb back to landing
    if st.button("← File Catalog", key="wl_back_to_landing"):
        st.session_state["file_catalog_domain"] = None
        st.session_state.pop("file_catalog_domain", None)
        st.session_state.app_mode = "file_catalog"
        st.rerun()
    st.title("🛢️ Petroleum File Catalog")
    st.caption("Browse and search LAS, DLIS and LIS files — matched to PPDM wells.")

    if not _AVAILABLE:
        st.error(f"Catalog dependencies missing:\n\n`{_IMPORT_ERROR}`")
        return

    engine = _get_engine()
    if engine is None:
        st.warning("No database connection. Connect via the main pipeline first.")
        return

    # Ensure schema exists on every load (idempotent)
    with st.spinner("Checking catalog schema…"):
        try:
            created = ensure_catalog_schema(engine)
            if created:
                st.success(f"Created catalog tables: {', '.join(created)}")
        except Exception as e:
            st.error(f"Could not create catalog schema: {e}")
            return

    # Ensure mapping staging table exists
    try:
        ensure_map_table(engine)
    except Exception:
        pass  # Non-fatal — mapping tab will show error if needed

    tab_search, tab_dash = st.tabs([
        "🔍 Search", "📊 Dashboard",
    ])

    with tab_search:
        _render_search(engine)

    with tab_dash:
        _render_dashboard(engine)


def run_seismic():
    """Entry point when launched from the Seismic domain card."""
    if st.button("← File Catalog", key="seis_back_to_landing"):
        st.session_state["file_catalog_domain"] = None
        st.session_state.app_mode = "file_catalog"
        st.rerun()
    st.title("🌊 Petroleum File Catalog")
    st.caption("Browse and search SEG-Y and P1/90 files — bounding boxes, survey maps, EBCDIC headers.")

    if not _AVAILABLE:
        st.error(f"Catalog dependencies missing:\n\n`{_IMPORT_ERROR}`")
        return

    engine = _get_engine()
    if engine is None:
        st.warning("No database connection. Connect via the main pipeline first.")
        return

    with st.spinner("Checking catalog schema…"):
        try:
            ensure_catalog_schema(engine)
        except Exception as e:
            st.error(f"Could not verify catalog schema: {e}")
            return

    try:
        ensure_map_table(engine)
    except Exception:
        pass

    try:
        from modules.segy_catalog import ensure_seis_catalog_columns
        added = ensure_seis_catalog_columns(engine)
        if added:
            st.info(f"Schema updated — added columns: {', '.join(added)}")
    except Exception as _e:
        st.warning(f"Schema check failed: {_e}")

    try:
        repos = list_repositories(engine)
        repo_opts = {"(none — assign later)": ""} | {
            f"{r['REPOSITORY_NAME']} ({r['BASE_PATH']})": r["REPOSITORY_ID"]
            for _, r in repos.iterrows()
        } if not repos.empty else {"(none)": ""}
    except Exception:
        repo_opts = {"(none)": ""}

    _render_seismic(engine)


def run_repos():
    """Entry point when launched from the Manage Repositories card."""
    if st.button("← File Catalog", key="repos_back_to_landing"):
        st.session_state["file_catalog_domain"] = None
        st.session_state.app_mode = "file_catalog"
        st.rerun()
    st.title("📁 Manage Repositories")
    st.caption("Register, edit and delete the file system locations where your files are stored.")

    if not _AVAILABLE:
        st.error(f"Catalog dependencies missing:\n\n`{_IMPORT_ERROR}`")
        return

    engine = _get_engine()
    if engine is None:
        st.warning("No database connection. Connect via the main pipeline first.")
        return

    with st.spinner("Checking catalog schema…"):
        try:
            ensure_catalog_schema(engine)
        except Exception as e:
            st.error(f"Could not verify catalog schema: {e}")
            return

    _render_repositories(engine)


def run_spatial():
    """Spatial domain — import Shapefile / GeoJSON into PPDM."""
    if st.button("← File Catalog", key="spatial_back"):
        st.session_state["file_catalog_domain"] = None
        st.session_state.app_mode = "file_catalog"
        st.rerun()
    st.title("🗺️ Spatial Catalog")
    st.caption(
        "Import well locations, field boundaries and seismic line geometry "
        "from Shapefile or GeoJSON into PPDM."
    )

    if not _AVAILABLE:
        st.error(f"Catalog dependencies missing: `{_IMPORT_ERROR}`")
        return

    engine = _get_engine()
    if engine is None:
        st.warning("No database connection. Connect via the main pipeline first.")
        return

    st.divider()

    tab_wl, tab_field, tab_seis = st.tabs([
        "📍 Well Locations", "🔷 Field Boundaries", "〰️ Seismic Lines"
    ])

    with tab_wl:
        _render_spatial_wells(engine)

    with tab_field:
        st.info("Field boundary import — coming soon.")

    with tab_seis:
        st.info("Seismic line geometry import — coming soon.")


def _render_spatial_wells(engine):
    """
    Import well locations from Shapefile or GeoJSON.
    Matches each feature to a PPDM WELL record by UWI or well name,
    then updates SURFACE_LATITUDE and SURFACE_LONGITUDE.
    """
    from sqlalchemy import text

    st.markdown("**Import well locations from a spatial file**")
    st.caption(
        "Supported formats: Shapefile (.shp + .dbf + .shx), GeoJSON (.geojson / .json). "
        "Each feature must have a field containing a UWI or well name to match against PPDM."
    )

    file_path = st.text_input(
        "File path",
        key="spatial_wl_path",
        placeholder=r"e.g. C:\GIS\well_locations.shp  or  C:\GIS\wells.geojson"
    )

    if not file_path:
        return

    from pathlib import Path
    p = Path(file_path)
    if not p.exists():
        st.error(f"File not found: `{file_path}`")
        return

    ext = p.suffix.lower()
    if ext not in (".shp", ".geojson", ".json"):
        st.error(f"Unsupported format: `{ext}`. Use .shp or .geojson")
        return

    # Read file
    try:
        try:
            import geopandas as gpd
            gdf = gpd.read_file(file_path)
            has_gpd = True
        except ImportError:
            has_gpd = False
            # Fallback to pyshp for shapefiles, json for geojson
            if ext == ".shp":
                try:
                    import shapefile as pyshp
                    sf  = pyshp.Reader(file_path)
                    fields = [f[0] for f in sf.fields[1:]]
                    records = sf.records()
                    shapes  = sf.shapes()
                    import pandas as pd
                    rows = []
                    for rec, shp in zip(records, shapes):
                        d = dict(zip(fields, rec))
                        if shp.points:
                            d["_lon"] = shp.points[0][0]
                            d["_lat"] = shp.points[0][1]
                        rows.append(d)
                    gdf = pd.DataFrame(rows)
                except ImportError:
                    st.error(
                        "Neither **geopandas** nor **pyshp** is installed. "
                        "Install one with: `pip install geopandas` or `pip install pyshp`"
                    )
                    return
            else:
                import json, pandas as pd
                data  = json.loads(p.read_text())
                rows  = []
                for feat in data.get("features", []):
                    d = feat.get("properties", {}) or {}
                    coords = feat.get("geometry", {}).get("coordinates", [])
                    if coords and len(coords) >= 2:
                        d["_lon"] = coords[0]
                        d["_lat"] = coords[1]
                    rows.append(d)
                gdf = pd.DataFrame(rows)

    except Exception as e:
        st.error(f"Could not read file: {e}")
        return

    # Extract coordinates
    if has_gpd:
        import pandas as pd
        gdf = gdf.to_crs("EPSG:4326") if hasattr(gdf, "crs") and gdf.crs else gdf
        gdf["_lon"] = gdf.geometry.x if hasattr(gdf, "geometry") else None
        gdf["_lat"] = gdf.geometry.y if hasattr(gdf, "geometry") else None
        gdf = pd.DataFrame(gdf.drop(columns="geometry", errors="ignore"))

    st.success(f"✅ Loaded {len(gdf):,} feature(s)")
    st.dataframe(gdf.head(5), hide_index=True, use_container_width=True)

    if "_lon" not in gdf.columns or "_lat" not in gdf.columns:
        st.error("Could not extract coordinates. Ensure the file contains point geometry.")
        return

    # Field mapping
    st.divider()
    st.markdown("**Map fields to PPDM**")
    str_cols = ["— none —"] + [c for c in gdf.columns
                                if c not in ("_lon","_lat")]

    col1, col2 = st.columns(2)
    with col1:
        uwi_field = st.selectbox(
            "UWI / API field", str_cols, key="spatial_uwi_field",
            help="Field containing the well UWI or API number"
        )
    with col2:
        name_field = st.selectbox(
            "Well name field (fallback)", str_cols, key="spatial_name_field",
            help="Used for fuzzy matching if UWI doesn't match"
        )

    crs_note = st.text_input(
        "Coordinate system note (optional)",
        key="spatial_crs",
        placeholder="e.g. WGS84 / GDA2020",
        help="Stored as a remark — coordinates are assumed to be WGS84 lat/lon"
    )

    if st.button("🗺️ Preview matches", key="spatial_preview_btn",
                 disabled=uwi_field == "— none —"):
        try:
            ppdm_uwis = {
                r[0]: r[1] for r in engine.connect().execute(text(
                    "SELECT UWI, WELL_NAME FROM dbo.WELL WHERE UWI IS NOT NULL"
                )).fetchall()
            }
        except Exception as e:
            st.error(f"Could not load PPDM wells: {e}")
            return

        matched, unmatched = [], []
        for _, row in gdf.iterrows():
            uwi = str(row.get(uwi_field, "")).strip() if uwi_field != "— none —" else ""
            lat = row.get("_lat")
            lon = row.get("_lon")
            if uwi in ppdm_uwis:
                matched.append({
                    "UWI": uwi,
                    "WELL_NAME": ppdm_uwis[uwi],
                    "LAT": lat, "LON": lon,
                    "MATCH": "✅ Exact"
                })
            else:
                unmatched.append({
                    "UWI": uwi, "LAT": lat, "LON": lon,
                    "MATCH": "❌ No match"
                })

        all_rows = matched + unmatched
        import pandas as pd
        preview_df = pd.DataFrame(all_rows)
        st.session_state["spatial_preview"] = preview_df
        st.session_state["spatial_matched"] = matched

    if "spatial_preview" in st.session_state:
        prev = st.session_state["spatial_preview"]
        n_match = len(st.session_state.get("spatial_matched", []))
        st.metric("Matched", n_match)
        st.metric("Unmatched", len(prev) - n_match)
        st.dataframe(prev, hide_index=True, use_container_width=True)

        if n_match > 0:
            if st.button(
                f"📥 Update {n_match} well location(s) in PPDM",
                type="primary", key="spatial_import_btn"
            ):
                updated = 0
                errors  = []
                for row in st.session_state["spatial_matched"]:
                    try:
                        with engine.begin() as con:
                            con.execute(text("""
                                UPDATE dbo.WELL
                                SET SURFACE_LATITUDE  = :lat,
                                    SURFACE_LONGITUDE = :lon,
                                    ROW_CHANGED_BY    = 'SPATIAL_IMPORT',
                                    ROW_CHANGED_DATE  = GETUTCDATE()
                                WHERE UWI = :uwi
                            """), {"lat": row["LAT"], "lon": row["LON"],
                                   "uwi": row["UWI"]})
                        updated += 1
                    except Exception as e:
                        errors.append(f"{row['UWI']}: {e}")
                if updated:
                    st.success(f"✅ Updated {updated} well location(s) in PPDM.")
                if errors:
                    with st.expander(f"❌ {len(errors)} error(s)"):
                        for e in errors:
                            st.text(e)
                st.session_state.pop("spatial_preview", None)
                st.session_state.pop("spatial_matched", None)


# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD TAB
# ─────────────────────────────────────────────────────────────────────────────

def _render_dashboard(engine):
    st.subheader("Catalog overview")

    try:
        stats = get_catalog_summary(engine)
    except Exception as e:
        st.error(str(e))
        return

    # Row 1 — file counts
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total files",     f"{stats['total_files']:,}")
    c2.metric("Matched wells",   f"{stats['matched_wells']:,}")
    c3.metric("Repositories",    stats["repositories"])
    c4.metric("Total size",      f"{stats['total_size_mb']:,.1f} MB")

    # Row 2 — depth and length
    c5, c6, c7, c8 = st.columns(4)
    c5.metric("Shallowest",   f"{stats['shallowest']:,.1f}"  if stats["shallowest"]  else "—")
    c6.metric("Deepest",      f"{stats['deepest']:,.1f}"     if stats["deepest"]     else "—")
    c7.metric("Avg length",   f"{stats['avg_length']:,.1f}"  if stats["avg_length"]  else "—")
    c8.metric("Max length",   f"{stats['max_length']:,.1f}"  if stats["max_length"]  else "—")

    # Row 3 — location coverage
    c9, c10, c11 = st.columns(3)
    c9.metric("Countries",  stats["countries"])
    c10.metric("States / Provinces", stats["states"])
    c11.metric("Counties",  stats["counties"])

    st.divider()

    col_curves, col_geo = st.columns(2)

    with col_curves:
        if stats["top_curves"]:
            st.markdown("**Most common curves**")
            curve_df = pd.DataFrame(stats["top_curves"])
            st.bar_chart(curve_df.set_index("curve")["count"])

    with col_geo:
        if stats["top_countries"]:
            st.markdown("**Top countries**")
            st.dataframe(
                pd.DataFrame(stats["top_countries"]), hide_index=True
            )
        if stats["top_states"]:
            st.markdown("**Top states / provinces**")
            st.dataframe(
                pd.DataFrame(stats["top_states"]), hide_index=True
            )


# ─────────────────────────────────────────────────────────────────────────────
# SEARCH TAB
# ─────────────────────────────────────────────────────────────────────────────


def _search_dlis(engine, uwi, well_name, channel, field,
                  company, depth_min_m, depth_max_m) -> pd.DataFrame:
    """Query DLIS tables and return unified result rows."""
    from sqlalchemy import text
    where = ["df.ACTIVE_IND = 'Y'"]
    params = {}
    if uwi:
        uwi_p = uwi if ("%" in uwi or "_" in uwi) else f"%{uwi}%"
        where.append("df.UWI LIKE :uwi"); params["uwi"] = uwi_p
    if well_name:
        where.append("lf.WELL_NAME LIKE :wn"); params["wn"] = f"%{well_name}%"
    if field:
        where.append("lf.FIELD_NAME LIKE :field"); params["field"] = f"%{field}%"
    if company:
        where.append("lf.COMPANY LIKE :company"); params["company"] = f"%{company}%"
    if depth_min_m is not None:
        where.append("fr.BASE_DEPTH_M >= :dmin"); params["dmin"] = depth_min_m
    if depth_max_m is not None:
        where.append("fr.TOP_DEPTH_M <= :dmax"); params["dmax"] = depth_max_m

    join_ch = ""
    if channel:
        join_ch = "JOIN [las_catalog].[DLIS_CHANNEL] ch ON ch.DLIS_FILE_ID = fr.DLIS_FILE_ID AND ch.LOGICAL_FILE_IDX = fr.LOGICAL_FILE_IDX AND ch.FRAME_NAME = fr.FRAME_NAME"
        where.append("ch.CHANNEL_NAME = :ch"); params["ch"] = channel.upper()

    sql = f"""
        SELECT DISTINCT
            'DLIS' AS FORMAT,
            df.UWI, lf.WELL_NAME, lf.FIELD_NAME AS FIELD,
            lf.COMPANY AS OPERATOR, NULL AS COUNTRY,
            NULL AS STATE_PROVINCE, NULL AS COUNTY,
            fr.TOP_DEPTH_M AS TOP_DEPTH, fr.BASE_DEPTH_M AS BASE_DEPTH,
            fr.BASE_DEPTH_M - fr.TOP_DEPTH_M AS LENGTH,
            'M' AS DEPTH_UOM,
            fr.CHANNEL_COUNT AS CURVE_COUNT,
            fr.SAMPLE_COUNT, CONVERT(NVARCHAR(50), lf.CREATION_TIME, 120) AS LOG_DATE,
            lf.PRODUCER_NAME AS SERVICE_COMPANY,
            df.FILE_SIZE_KB,
            CASE WHEN RIGHT(r.BASE_PATH,1) = '\\' THEN r.BASE_PATH ELSE r.BASE_PATH + '\\' END + df.FILE_NAME AS FULL_PATH,
            df.FILE_NAME, r.REPOSITORY_NAME,
            df.CATALOG_DATE, df.LAST_SEEN_DATE,
            fr.FRAME_NAME AS EXTRA_INFO,
            df.DLIS_FILE_ID AS FILE_ID
        FROM [las_catalog].[DLIS_FILE] df
        JOIN [las_catalog].[DLIS_LOGICAL_FILE] lf
          ON lf.DLIS_FILE_ID = df.DLIS_FILE_ID
        JOIN [las_catalog].[DLIS_FRAME] fr
          ON fr.DLIS_FILE_ID = lf.DLIS_FILE_ID
         AND fr.LOGICAL_FILE_IDX = lf.LOGICAL_FILE_IDX
        JOIN [las_catalog].[WL_REPOSITORY] r
          ON r.REPOSITORY_ID = df.REPOSITORY_ID
        {join_ch}
        WHERE {" AND ".join(where)}
        ORDER BY df.UWI, fr.TOP_DEPTH_M
    """
    with engine.connect() as con:
        rows = con.execute(text(sql), params).fetchall()
    return pd.DataFrame(rows, columns=[
        "FORMAT","UWI","WELL_NAME","FIELD","OPERATOR","COUNTRY",
        "STATE_PROVINCE","COUNTY","TOP_DEPTH","BASE_DEPTH","LENGTH",
        "DEPTH_UOM","CURVE_COUNT","SAMPLE_COUNT","LOG_DATE","SERVICE_COMPANY",
        "FILE_SIZE_KB","FULL_PATH","FILE_NAME","REPOSITORY_NAME",
        "CATALOG_DATE","LAST_SEEN_DATE","EXTRA_INFO","FILE_ID"
    ])


def _search_lis(engine, uwi, well_name, channel,
                depth_min, depth_max, depth_uom) -> pd.DataFrame:
    """Query LIS tables and return unified result rows."""
    from sqlalchemy import text
    where = ["f.ACTIVE_IND = 'Y'"]
    params = {}
    if uwi:
        uwi_p = uwi if ("%" in uwi or "_" in uwi) else f"%{uwi}%"
        where.append("f.UWI LIKE :uwi"); params["uwi"] = uwi_p
    if well_name:
        where.append("f.WELL_NAME LIKE :wn"); params["wn"] = f"%{well_name}%"
    if depth_min is not None:
        where.append("f.BASE_DEPTH >= :dmin"); params["dmin"] = depth_min
    if depth_max is not None:
        where.append("f.TOP_DEPTH <= :dmax"); params["dmax"] = depth_max
    if depth_uom:
        where.append("f.DEPTH_UOM = :duom"); params["duom"] = depth_uom.upper()

    join_ch = ""
    if channel:
        join_ch = "JOIN [las_catalog].[LIS_CHANNEL] ch ON ch.LIS_FILE_ID = f.LIS_FILE_ID"
        where.append("ch.CHANNEL_NAME = :ch"); params["ch"] = channel.upper()

    sql = f"""
        SELECT DISTINCT
            'LIS' AS FORMAT,
            f.UWI, f.WELL_NAME, f.FIELD_NAME AS FIELD,
            f.COMPANY AS OPERATOR, NULL AS COUNTRY,
            NULL AS STATE_PROVINCE, NULL AS COUNTY,
            f.TOP_DEPTH, f.BASE_DEPTH,
            CASE WHEN f.BASE_DEPTH IS NOT NULL AND f.TOP_DEPTH IS NOT NULL
                 THEN f.BASE_DEPTH - f.TOP_DEPTH ELSE NULL END AS LENGTH,
            f.DEPTH_UOM, f.CHANNEL_COUNT AS CURVE_COUNT,
            f.SAMPLE_COUNT, f.LOG_DATE, f.SERVICE_COMPANY,
            f.FILE_SIZE_KB,
            CASE WHEN RIGHT(r.BASE_PATH,1) = '\\' THEN r.BASE_PATH ELSE r.BASE_PATH + '\\' END + f.FILE_NAME AS FULL_PATH,
            f.FILE_NAME, r.REPOSITORY_NAME,
            CONVERT(NVARCHAR(30), f.CATALOG_DATE, 120) AS CATALOG_DATE,
            CONVERT(NVARCHAR(30), f.LAST_SEEN_DATE, 120) AS LAST_SEEN_DATE,
            NULL AS EXTRA_INFO,
            f.LIS_FILE_ID AS FILE_ID
        FROM [las_catalog].[LIS_FILE] f
        JOIN [las_catalog].[WL_REPOSITORY] r
          ON r.REPOSITORY_ID = f.REPOSITORY_ID
        {join_ch}
        WHERE {" AND ".join(where)}
        ORDER BY f.UWI, f.TOP_DEPTH
    """
    with engine.connect() as con:
        rows = con.execute(text(sql), params).fetchall()
    return pd.DataFrame(rows, columns=[
        "FORMAT","UWI","WELL_NAME","FIELD","OPERATOR","COUNTRY",
        "STATE_PROVINCE","COUNTY","TOP_DEPTH","BASE_DEPTH","LENGTH",
        "DEPTH_UOM","CURVE_COUNT","SAMPLE_COUNT","LOG_DATE","SERVICE_COMPANY",
        "FILE_SIZE_KB","FULL_PATH","FILE_NAME","REPOSITORY_NAME",
        "CATALOG_DATE","LAST_SEEN_DATE","EXTRA_INFO","FILE_ID"
    ])


def _do_search(engine, uwi, well_name, curve, field, operator,
               country, state, county,
               depth_min, depth_max, depth_uom,
               length_min, length_max, fmt_filter):
    """on_click callback — runs cross-format search."""
    try:
        frames = []

        if fmt_filter in ("All", "LAS"):
            las_res = search_catalog(
                engine, uwi=uwi, well_name=well_name, curve_id=curve,
                field=field, operator=operator, country=country,
                state_province=state, county=county,
                depth_min=depth_min or None, depth_max=depth_max or None,
                depth_uom=depth_uom,
                length_min=length_min or None, length_max=length_max or None,
            )
            if not las_res.empty:
                las_res.insert(0, "FORMAT", "LAS")
                las_res["EXTRA_INFO"] = None
                las_res["FILE_ID"] = las_res["LAS_FILE_ID"]
                frames.append(las_res)

        if fmt_filter in ("All", "DLIS"):
            # Convert depth to metres for DLIS search (stored as metres)
            dm = _to_metres_search(depth_min, depth_uom)
            dx = _to_metres_search(depth_max, depth_uom)
            dlis_res = _search_dlis(engine, uwi, well_name, curve,
                                     field, operator, dm, dx)
            if not dlis_res.empty:
                frames.append(dlis_res)

        if fmt_filter in ("All", "LIS"):
            lis_res = _search_lis(engine, uwi, well_name, curve,
                                   depth_min or None, depth_max or None,
                                   depth_uom)
            if not lis_res.empty:
                frames.append(lis_res)

        if frames:
            combined = pd.concat(frames, ignore_index=True)
            # Coerce any mixed-type columns (e.g. datetime + string) to str
            # so PyArrow can serialize the DataFrame for st.dataframe()
            for col in combined.columns:
                if combined[col].dtype == object:
                    try:
                        combined[col] = combined[col].where(
                            combined[col].isna(),
                            combined[col].astype(str)
                        )
                    except Exception:
                        pass
        else:
            combined = pd.DataFrame()

        st.session_state["cat_results"] = combined
    except Exception as e:
        st.session_state["cat_results"] = str(e)


def _to_metres_search(val, uom: str) -> object:
    """Convert depth filter value to metres for DLIS search."""
    if val is None:
        return None
    uom_lower = str(uom).lower().strip()
    multipliers = {"m": 1.0, "ft": 0.3048, "f": 0.3048, "feet": 0.3048}
    mult = multipliers.get(uom_lower, 1.0)
    return float(val) * mult


def _render_search(engine):
    st.subheader("Search catalog")
    st.caption(
        "UWI supports SQL wildcards: **%** = any characters, **_** = single character  "
        "— e.g. `17-031%` or `%SMITH%`  |  "
        "Results are sorted by the order selected below the results table."
    )

    # Distinct values for dropdowns — cached 5 min
    @st.cache_data(ttl=300, show_spinner=False)
    def _opts(col):
        return [""] + get_distinct_values(engine, col)

    # ── Filter panel ─────────────────────────────────────────────────
    with st.expander("🔎 Search filters", expanded=True):

        # Row 1 — identifiers
        r1a, r1b, r1c, r1d = st.columns([2, 2, 2, 2])
        with r1a:
            f_uwi      = st.text_input("UWI", key="cat_f_uwi",
                                        placeholder="e.g. 17-031% or full UWI")
        with r1b:
            f_well_name = st.selectbox("Well name", options=_opts("WELL_NAME"),
                                        key="cat_f_wn")
        with r1c:
            f_field    = st.selectbox("Field", options=_opts("FIELD"),
                                       key="cat_f_field")
        with r1d:
            f_operator = st.selectbox("Operator", options=_opts("OPERATOR"),
                                       key="cat_f_op")

        # Row 2 — location
        r2a, r2b, r2c, r2d = st.columns([2, 2, 2, 2])
        with r2a:
            f_country  = st.selectbox("Country", options=_opts("COUNTRY"),
                                       key="cat_f_country")
        with r2b:
            f_state    = st.selectbox("State / Province",
                                       options=_opts("STATE_PROVINCE"),
                                       key="cat_f_state")
        with r2c:
            f_county   = st.selectbox("County", options=_opts("COUNTY"),
                                       key="cat_f_county")
        with r2d:
            f_curve    = st.text_input("Curve mnemonic (exact)", key="cat_f_curve",
                                        placeholder="e.g. GR").upper()

        # Row 3 — depth / length / UOM
        r3a, r3b, r3c, r3d, r3e = st.columns([2, 2, 2, 2, 1])
        with r3a:
            f_depth_min  = st.number_input("Min depth", value=None,
                                            key="cat_f_dmin", placeholder="e.g. 500")
        with r3b:
            f_depth_max  = st.number_input("Max depth", value=None,
                                            key="cat_f_dmax", placeholder="e.g. 3000")
        with r3c:
            f_length_min = st.number_input("Min length", value=None,
                                            key="cat_f_lmin", placeholder="e.g. 500",
                                            help="BASE_DEPTH − TOP_DEPTH")
        with r3d:
            f_length_max = st.number_input("Max length", value=None,
                                            key="cat_f_lmax", placeholder="e.g. 2000")
        with r3e:
            f_uom = st.selectbox("UOM", options=_opts("DEPTH_UOM"), key="cat_f_uom")

        fmt_filter = st.radio(
            "Format", options=["All", "LAS", "DLIS", "LIS"],
            horizontal=True, key="cat_fmt_filter",
        )

        st.button("🔍 Search", type="primary", key="cat_search_btn",
                  on_click=_do_search, args=(
                      engine, f_uwi, f_well_name, f_curve, f_field, f_operator,
                      f_country, f_state, f_county,
                      f_depth_min, f_depth_max, f_uom,
                      f_length_min, f_length_max, fmt_filter,
                  ))

    if "cat_results" not in st.session_state:
        return

    results = st.session_state["cat_results"]

    if isinstance(results, str):
        st.error(results)
        return

    st.markdown(f"**{len(results):,} file(s) found**")

    if results.empty:
        st.info("No files match the search criteria.")
        return

    # ── Sort builder ─────────────────────────────────────────────────
    _SORTABLE = [
        "", "UWI", "WELL_NAME", "FIELD", "OPERATOR", "COUNTRY",
        "STATE_PROVINCE", "COUNTY", "TOP_DEPTH", "BASE_DEPTH",
        "LENGTH", "CURVE_COUNT", "LOG_DATE", "FILE_SIZE_KB", "REPOSITORY_NAME",
    ]
    # Only show columns that exist in results
    _available = [""] + [c for c in _SORTABLE[1:] if c in results.columns]

    st.markdown("**Sort results**")
    sc1, sd1, sc2, sd2, sc3, sd3 = st.columns([2, 1, 2, 1, 2, 1])
    with sc1: sort1 = st.selectbox("Primary",  _available, key="rs_c1", label_visibility="collapsed")
    with sd1: dir1  = st.selectbox("↑↓", ["ASC","DESC"], key="rs_d1", label_visibility="collapsed")
    with sc2: sort2 = st.selectbox("Then by",  _available, key="rs_c2", label_visibility="collapsed")
    with sd2: dir2  = st.selectbox("↑↓", ["ASC","DESC"], key="rs_d2", label_visibility="collapsed")
    with sc3: sort3 = st.selectbox("Then by",  _available, key="rs_c3", label_visibility="collapsed")
    with sd3: dir3  = st.selectbox("↑↓", ["ASC","DESC"], key="rs_d3", label_visibility="collapsed")

    # Apply sorts in Python
    sort_pairs = [(c, d) for c, d in [(sort1,dir1),(sort2,dir2),(sort3,dir3)] if c]
    display_df = results.copy()
    if sort_pairs:
        sort_by  = [p[0] for p in sort_pairs]
        sort_asc = [p[1] == "ASC" for p in sort_pairs]
        display_df = display_df.sort_values(by=sort_by, ascending=sort_asc)

    # Results table — hide internal ID
    display_cols = [c for c in display_df.columns if c != "LAS_FILE_ID"]
    st.dataframe(display_df[display_cols], hide_index=True)

    # ── Curve inspector ───────────────────────────────────────────────
    st.divider()
    st.markdown("**Inspect curves for a file**")
    file_options = results["FULL_PATH"].tolist()
    selected_path = st.selectbox("Select file", options=file_options,
                                 key="cat_file_select")
    if selected_path:
        sel_row = results[results["FULL_PATH"] == selected_path].iloc[0]

        # FILE_ID and FORMAT are present in all three format result sets
        file_id = sel_row.get("FILE_ID") or sel_row.get("LAS_FILE_ID") or ""
        fmt     = str(sel_row.get("FORMAT", "LAS")).upper()

        def _metric_val(col):
            if col not in sel_row.index:
                return None
            v = sel_row[col]
            if v is None:
                return None
            try:
                import math
                f = float(v)
                return None if math.isnan(f) else f
            except (TypeError, ValueError):
                # Non-numeric — return as string (for UOM, FORMAT etc.)
                s = str(v).strip()
                return s if s not in ("", "None", "nan") else None

        def _safe_float(col):
            v = _metric_val(col)
            if v is None:
                return None
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        top    = _safe_float("TOP_DEPTH")
        base   = _safe_float("BASE_DEPTH")
        length = _safe_float("LENGTH")
        uom    = _metric_val("DEPTH_UOM") or ""
        if isinstance(uom, float):
            uom = ""

        col_a, col_b, col_c, col_d = st.columns(4)
        col_a.metric("Format",     fmt)
        col_b.metric("Top depth",  f"{top:,.1f} {uom}"    if top    is not None else "—")
        col_c.metric("Base depth", f"{base:,.1f} {uom}"   if base   is not None else "—")
        col_d.metric("Length",     f"{length:,.1f} {uom}" if length is not None else "—")

        if fmt == "LAS" and file_id:
            curves_df = get_file_curves(engine, file_id)
            if not curves_df.empty:
                st.dataframe(curves_df, hide_index=True)
            st.divider()
            if st.checkbox("📈 Plot curves", key="cat_plot_curves", value=False):
                import os as _os
                if _os.path.exists(selected_path):
                    _render_curve_plot(selected_path, curves_df)
                else:
                    st.warning(f"File not found at catalogued path:\n\n`{selected_path}`\n\nEnter the current file path:")
                    _ov = st.text_input("File path", key="las_path_override",
                                        placeholder=r"e.g. C:\Wells\myfile.las")
                    if _ov and _os.path.exists(_ov):
                        _render_curve_plot(_ov, curves_df)
                    elif _ov:
                        st.error(f"File not found: `{_ov}`")

        elif fmt == "DLIS" and file_id:
            try:
                from sqlalchemy import text
                with engine.connect() as con:
                    rows = con.execute(text("""
                        SELECT ch.CHANNEL_NAME, ch.UNITS, ch.LONG_NAME,
                               ch.IS_INDEX, fr.FRAME_NAME,
                               fr.TOP_DEPTH_M, fr.BASE_DEPTH_M, fr.DEPTH_UOM
                        FROM [las_catalog].[DLIS_CHANNEL] ch
                        JOIN [las_catalog].[DLIS_FRAME] fr
                          ON fr.DLIS_FILE_ID     = ch.DLIS_FILE_ID
                         AND fr.LOGICAL_FILE_IDX = ch.LOGICAL_FILE_IDX
                         AND fr.FRAME_NAME       = ch.FRAME_NAME
                        WHERE ch.DLIS_FILE_ID = :id
                        ORDER BY fr.FRAME_NAME, ch.IS_INDEX DESC, ch.CHANNEL_NAME
                    """), {"id": file_id}).fetchall()
                if rows:
                    st.dataframe(
                        pd.DataFrame(rows, columns=[
                            "CHANNEL_NAME","UNITS","LONG_NAME","IS_INDEX",
                            "FRAME_NAME","TOP_DEPTH_M","BASE_DEPTH_M","DEPTH_UOM"
                        ]), hide_index=True
                    )
                else:
                    st.info("No channel details found.")
            except Exception as e:
                st.error(str(e))

            st.divider()
            if st.checkbox("📈 Plot curves", key="cat_dlis_plot_cb", value=False):
                import os as _os
                if _os.path.exists(selected_path):
                    _render_dlis_plot(selected_path)
                else:
                    st.warning(f"File not found at catalogued path:\n\n`{selected_path}`\n\nEnter the current file path:")
                    _ov = st.text_input("File path", key="dlis_path_override",
                                        placeholder=r"e.g. C:\Wells\myfile.dlis")
                    if _ov and _os.path.exists(_ov):
                        _render_dlis_plot(_ov)
                    elif _ov:
                        st.error(f"File not found: `{_ov}`")

        elif fmt == "LIS" and file_id:
            try:
                from sqlalchemy import text
                with engine.connect() as con:
                    rows = con.execute(text("""
                        SELECT CHANNEL_NAME, UNITS, DESCRIPTION, IS_INDEX
                        FROM [las_catalog].[LIS_CHANNEL]
                        WHERE LIS_FILE_ID = :id
                        ORDER BY IS_INDEX DESC, CHANNEL_NAME
                    """), {"id": file_id}).fetchall()
                if rows:
                    st.dataframe(
                        pd.DataFrame(rows, columns=[
                            "CHANNEL_NAME","UNITS","DESCRIPTION","IS_INDEX"
                        ]), hide_index=True
                    )
                else:
                    st.info("No channel details found.")
            except Exception as e:
                st.error(str(e))

            st.divider()
            if st.checkbox("📈 Plot curves", key="cat_lis_plot_cb", value=False):
                import os as _os
                if _os.path.exists(selected_path):
                    _render_lis_plot(selected_path)
                else:
                    st.warning(f"File not found at catalogued path:\n\n`{selected_path}`\n\nEnter the current file path:")
                    _ov = st.text_input("File path", key="lis_path_override",
                                        placeholder=r"e.g. C:\Wells\myfile.lis")
                    if _ov and _os.path.exists(_ov):
                        _render_lis_plot(_ov)
                    elif _ov:
                        st.error(f"File not found: `{_ov}`")
    # ── Export row ────────────────────────────────────────────────────
    st.divider()
    exp_col, csv_col, spacer = st.columns([3, 1, 1])

    # Work out what formats are in the results
    formats_present = results["FORMAT"].unique().tolist() if "FORMAT" in results.columns else ["LAS"]
    has_las_only    = formats_present == ["LAS"]
    fmt_label       = "/".join(sorted(set(formats_present))) if len(formats_present) > 1 else formats_present[0]
    file_count      = len(results)

    with exp_col:
        export_dest = st.text_input(
            "Destination folder",
            key="cat_export_dest",
            placeholder=r"e.g. C:\Projects\Well_Logs\Export",
        )
        chk1, chk2 = st.columns(2)
        with chk1:
            overwrite = st.checkbox("Overwrite existing files", value=False,
                                    key="cat_export_overwrite")
        with chk2:
            # Header update only meaningful for LAS files
            if has_las_only:
                update_headers = st.checkbox(
                    "Update ~W header from PPDM",
                    value=False,
                    key="cat_export_update_headers",
                    help=(
                        "Updates existing ~W (Well info) values from PPDM before writing. "
                        "Only updates values already in the LAS file where PPDM value is "
                        "non-empty. Writes LAS 2.0. Original files are never modified."
                    ),
                )
            else:
                update_headers = False
                if "LAS" in formats_present:
                    st.caption("~W header update available for LAS-only results.")

    btn_export, btn_csv = st.columns([1, 1])
    with btn_export:
        if st.button(
            f"📤 Export {file_count} {fmt_label} file(s)",
            type="primary",
            key="cat_export_btn",
            disabled=not export_dest,
        ):
            with st.spinner(f"Copying {file_count} file(s)…"):
                try:
                    exp_result = export_files(
                        results, export_dest, overwrite,
                        update_headers=update_headers,
                        engine=engine,
                    )
                    st.session_state["cat_export_result"] = exp_result
                except Exception as e:
                    st.error(str(e))

    with btn_csv:
        st.download_button(
            label="⬇  Export Summary CSV",
            data=results.to_csv(index=False),
            file_name="las_catalog_results.csv",
            mime="text/csv",
            key="cat_download_btn",
        )
    if "cat_export_result" in st.session_state:
        r = st.session_state["cat_export_result"]
        if r["errors"] == 0 and r["missing"] == 0:
            st.success(f"✅ {r['copied']} copied, {r['skipped']} skipped")
        else:
            st.warning(
                f"{r['copied']} copied · {r['skipped']} skipped · "
                f"{r['missing']} missing · {r['errors']} error(s)"
            )
        detail_df = pd.DataFrame(r["details"])
        # Drop "updated" column if empty (header update not used)
        if "updated" in detail_df.columns and detail_df["updated"].eq("").all():
            detail_df = detail_df.drop(columns=["updated"])
        with st.expander("Export details", expanded=r["errors"] > 0):
            st.dataframe(detail_df, hide_index=True)

def _plot_tracks(file_name: str, depths, curves: list[dict],
                 depth_uom: str, key_prefix: str):
    """
    Shared well log plot — depth on inverted Y axis, one track per curve.

    curves: list of { "name": str, "unit": str, "data": np.array }
    """
    import math
    import numpy as np
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        st.error("matplotlib is required: pip install matplotlib")
        return

    if not curves:
        st.info("No plottable curves found.")
        return

    # Curve selection
    options = [c["name"] for c in curves]
    selected = st.multiselect(
        "Select curves to plot",
        options=options,
        default=options[:min(6, len(options))],
        key=f"{key_prefix}_curve_select",
    )
    if not selected:
        return

    plot_curves = [c for c in curves if c["name"] in selected]
    n = len(plot_curves)
    colours = [
        "#1f77b4", "#d62728", "#2ca02c", "#ff7f0e",
        "#9467bd", "#8c564b", "#e377c2", "#7f7f7f",
    ]

    fig, axes = plt.subplots(1, n, figsize=(max(3 * n, 6), 10), sharey=True)
    if n == 1:
        axes = [axes]
    fig.patch.set_facecolor("#f8f9fa")

    for i, (ax, curve) in enumerate(zip(axes, plot_curves)):
        colour = colours[i % len(colours)]
        data = np.where(curve["data"] == -999.25, np.nan, curve["data"])

        ax.plot(data, depths, colour, linewidth=0.8, alpha=0.9)
        ax.set_ylim(depths.max(), depths.min())
        title = f"{curve['name']}"
        if curve.get("unit"):
            title += f"\n({curve['unit']})"
        ax.set_title(title, fontsize=8, pad=2, color=colour, fontweight="bold")
        ax.xaxis.set_label_position("top")
        ax.xaxis.tick_top()
        ax.tick_params(axis="x", labelsize=7, colors=colour)
        ax.tick_params(axis="y", labelsize=7)
        ax.grid(True, axis="y", linestyle="--", alpha=0.4, linewidth=0.5)
        ax.grid(True, axis="x", linestyle=":",  alpha=0.3, linewidth=0.4)
        ax.set_facecolor("white")
        ax.spines["top"].set_color(colour)
        ax.spines["top"].set_linewidth(2)
        if i == 0:
            ax.set_ylabel(f"Depth ({depth_uom})", fontsize=9)
        else:
            ax.tick_params(labelleft=False)

    fig.suptitle(file_name, fontsize=10, y=0.02, color="#444")
    plt.tight_layout(rect=[0, 0.03, 1, 1])
    st.pyplot(fig)
    plt.close(fig)


def _render_curve_plot(las_path: str, curves_df: pd.DataFrame):
    """Plot a LAS file using lasio."""
    import math, numpy as np
    try:
        import lasio
    except ImportError:
        st.error("lasio is required: pip install lasio")
        return

    from pathlib import Path
    if not Path(las_path).exists():
        st.warning(f"File not found: `{las_path}`")
        return

    with st.spinner("Reading LAS file…"):
        try:
            las = lasio.read(las_path, ignore_header_errors=True)
        except Exception as e:
            st.error(f"Could not read LAS file: {e}")
            return

    depth_mnemonic = las.curves[0].mnemonic if las.curves else "DEPT"
    try:
        depths = las[depth_mnemonic]
    except Exception:
        st.error("Could not read depth curve.")
        return

    curves = []
    for curve in las.curves:
        if curve.mnemonic == depth_mnemonic:
            continue
        try:
            data = las[curve.mnemonic]
            if len(data) == len(depths) and not all(
                math.isnan(float(v)) for v in data[:10]
            ):
                curves.append({
                    "name": curve.mnemonic,
                    "unit": curve.unit or "",
                    "data": data,
                })
        except Exception:
            continue

    depth_uom = las.curves[0].unit if las.curves else ""
    _plot_tracks(Path(las_path).name, depths, curves, depth_uom,
                 key_prefix="las_plot")


def _render_dlis_plot(dlis_path: str, frame_name: str = ""):
    """Plot curves from a DLIS file using dlisio."""
    import numpy as np, warnings
    try:
        from dlisio import dlis
    except ImportError:
        st.error("dlisio is required: pip install dlisio")
        return

    from pathlib import Path
    if not Path(dlis_path).exists():
        st.warning(f"File not found: `{dlis_path}`")
        return

    with st.spinner("Reading DLIS file…"):
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                with dlis.load(dlis_path) as lfs:
                    lf = lfs[0]

                    # Pick frame
                    frames = list(lf.frames)
                    if not frames:
                        st.info("No frames found in this DLIS file.")
                        return

                    frame_names = [f.name for f in frames]
                    if len(frames) > 1:
                        selected_frame = st.selectbox(
                            "Frame", options=frame_names,
                            key="dlis_plot_frame"
                        )
                        frame = next(f for f in frames if f.name == selected_frame)
                    else:
                        frame = frames[0]
                        st.caption(f"Frame: {frame.name}")

                    # Load curves
                    try:
                        data = frame.curves()
                    except Exception as e:
                        st.error(f"Could not read frame curves: {e}")
                        return

                    idx_name = frame.index
                    if not idx_name or idx_name not in data.dtype.names:
                        st.error("Could not identify depth index channel.")
                        return

                    depths = data[idx_name].astype(float)

                    # Find index channel unit
                    idx_ch = next(
                        (c for c in frame.channels if c.name == idx_name), None
                    )
                    depth_uom = idx_ch.units if idx_ch else ""

                    curves = []
                    for ch in frame.channels:
                        if ch.name == idx_name:
                            continue
                        try:
                            ch_data = data[ch.name]
                            # Only scalar channels — skip array/image curves
                            if ch_data.ndim != 1:
                                continue
                            ch_data = ch_data.astype(float)
                            curves.append({
                                "name": ch.name,
                                "unit": ch.units or "",
                                "data": ch_data,
                            })
                        except Exception:
                            continue

        except Exception as e:
            st.error(f"Could not read DLIS file: {e}")
            return

    _plot_tracks(Path(dlis_path).name, depths, curves, depth_uom,
                 key_prefix="dlis_plot")


def _render_lis_plot(lis_path: str):
    """Plot curves from a LIS file using dlisio."""
    import numpy as np, warnings
    try:
        from dlisio import lis
    except ImportError:
        st.error("dlisio is required: pip install dlisio")
        return

    from pathlib import Path
    if not Path(lis_path).exists():
        st.warning(f"File not found: `{lis_path}`")
        return

    with st.spinner("Reading LIS file…"):
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                with lis.load(lis_path) as lfs:
                    lf = lfs[0]
                    specs = lf.data_format_specs()
                    if not specs:
                        st.info("No data format specs found.")
                        return

                    # Use first spec
                    spec = specs[0]
                    try:
                        data = lis.curves(lf, spec)
                    except Exception as e:
                        st.error(f"Could not read LIS curves: {e}")
                        return

                    names = data.dtype.names or []
                    if not names:
                        st.info("No curves found.")
                        return

                    depth_col = names[0]
                    depths    = data[depth_col].astype(float)

                    # Get units from wellsite data
                    channel_units = {}
                    for rec in lf.wellsite_data():
                        comps = rec.components()
                        current = None
                        for c in comps:
                            if c.mnemonic == "MNEM":
                                current = str(c.component).strip()
                            elif c.mnemonic == "PUNI" and current:
                                channel_units[current] = str(c.component).strip()
                                current = None

                    depth_uom = channel_units.get(depth_col.strip(), "")

                    curves = []
                    for name in names[1:]:
                        try:
                            ch_data = data[name].astype(float)
                            curves.append({
                                "name": name.strip(),
                                "unit": channel_units.get(name.strip(), ""),
                                "data": ch_data,
                            })
                        except Exception:
                            continue

        except Exception as e:
            st.error(f"Could not read LIS file: {e}")
            return

    _plot_tracks(Path(lis_path).name, depths, curves, depth_uom,
                 key_prefix="lis_plot")



# ─────────────────────────────────────────────────────────────────────────────
# MAPPING TAB
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# REPOSITORIES TAB
# ─────────────────────────────────────────────────────────────────────────────

def _render_repositories(engine):
    st.subheader("Repositories")
    st.caption("Register physical storage locations. Files are stored relative to the base path.")

    try:
        repos = list_repositories(engine)
    except Exception as e:
        st.error(str(e)); return

    if not repos.empty:
        st.dataframe(
            repos[["REPOSITORY_NAME","REPOSITORY_TYPE","BASE_PATH","ACTIVE_IND","REMARK"]],
            hide_index=True, use_container_width=True
        )
    else:
        st.info("No repositories registered yet.")

    st.divider()
    sub_add, sub_edit, sub_delete = st.tabs(["➕ Add", "✏️ Edit", "🗑 Delete"])

    with sub_add:
        st.markdown("**Register a new repository**")
        r_name   = st.text_input("Name",  key="repo_add_name",
                                  placeholder="e.g. Chevron DLIS Delivery")
        r_type   = st.selectbox("Type",
                                 ["LOCAL","UNC","S3","AZURE_BLOB","SHAREPOINT"],
                                 key="repo_add_type")
        r_path   = st.text_input("Base path", key="repo_add_path",
                                  placeholder=r"e.g. C:\WellLogs\DLIS")
        r_remark = st.text_input("Remark (optional)", key="repo_add_remark")

        if st.button("Add repository", type="primary", key="repo_add_btn"):
            if not r_name.strip():
                st.error("Name is required.")
            elif not r_path.strip():
                st.error("Base path is required.")
            else:
                try:
                    rid = add_repository(engine, r_name.strip(), r_type,
                                         r_path.strip(), r_remark.strip())
                    st.success(f"✅ Added — ID: `{rid}`")
                    st.rerun()
                except Exception as e:
                    st.error(str(e))

    with sub_edit:
        if repos.empty:
            st.info("No repositories to edit."); return
        from sqlalchemy import text as _text
        sel = st.selectbox("Select", repos["REPOSITORY_NAME"].tolist(), key="repo_edit_sel")
        row = repos[repos["REPOSITORY_NAME"] == sel].iloc[0]
        e_name   = st.text_input("Name",      value=row["REPOSITORY_NAME"], key="repo_e_name")
        _types   = ["LOCAL","UNC","S3","AZURE_BLOB","SHAREPOINT"]
        e_type   = st.selectbox("Type", _types,
                                 index=_types.index(row["REPOSITORY_TYPE"])
                                       if row["REPOSITORY_TYPE"] in _types else 0,
                                 key="repo_e_type")
        e_path   = st.text_input("Base path", value=row["BASE_PATH"],    key="repo_e_path")
        e_remark = st.text_input("Remark",    value=row["REMARK"] or "", key="repo_e_remark")
        if st.button("Save changes", type="primary", key="repo_edit_btn"):
            try:
                import datetime as _dt
                with engine.begin() as con:
                    con.execute(_text("""
                        UPDATE [las_catalog].[WL_REPOSITORY]
                        SET REPOSITORY_NAME=:nm, REPOSITORY_TYPE=:rt,
                            BASE_PATH=:bp, REMARK=:rm,
                            ROW_CHANGED_BY='DATA_WRANGLER', ROW_CHANGED_DATE=:now
                        WHERE REPOSITORY_ID=:id
                    """), {"nm": e_name.strip(), "rt": e_type,
                           "bp": e_path.strip(), "rm": e_remark.strip(),
                           "now": _dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
                           "id": row["REPOSITORY_ID"]})
                st.success("✅ Saved.")
                st.rerun()
            except Exception as e:
                st.error(str(e))

    with sub_delete:
        if repos.empty:
            st.info("No repositories to delete."); return
        from sqlalchemy import text as _text
        del_sel = st.selectbox("Select", repos["REPOSITORY_NAME"].tolist(), key="repo_del_sel")
        del_row = repos[repos["REPOSITORY_NAME"] == del_sel].iloc[0]
        del_id  = del_row["REPOSITORY_ID"]
        try:
            with engine.connect() as con:
                n = sum([
                    con.execute(_text("SELECT COUNT(*) FROM [las_catalog].[LAS_FILE]  WHERE REPOSITORY_ID=:id"),{"id":del_id}).scalar() or 0,
                    con.execute(_text("SELECT COUNT(*) FROM [las_catalog].[DLIS_FILE] WHERE REPOSITORY_ID=:id"),{"id":del_id}).scalar() or 0,
                    con.execute(_text("SELECT COUNT(*) FROM [las_catalog].[LIS_FILE]  WHERE REPOSITORY_ID=:id"),{"id":del_id}).scalar() or 0,
                ])
        except Exception:
            n = -1
        if n > 0:
            st.warning(f"Repository has {n:,} file(s). Remove or reassign them first.")
        else:
            st.warning(f"This will permanently delete **{del_sel}**.")
            if st.button("🗑 Confirm delete", type="primary", key="repo_del_btn"):
                try:
                    with engine.begin() as con:
                        con.execute(_text(
                            "DELETE FROM [las_catalog].[WL_REPOSITORY] WHERE REPOSITORY_ID=:id"
                        ), {"id": del_id})
                    st.success("✅ Deleted.")
                    st.rerun()
                except Exception as e:
                    st.error(str(e))


# ─────────────────────────────────────────────────────────────────────────────
# CATALOG TAB
# ─────────────────────────────────────────────────────────────────────────────

def _goto_wl_map():
    st.session_state["app_mode"]           = "file_catalog"
    st.session_state["file_catalog_domain"] = "wells"
    st.session_state["wl_goto_mapping"]    = True  # flag to auto-open mapping tab


def _validate_file_path(file_path: str, extensions: list) -> tuple:
    from pathlib import Path
    p = Path(file_path)
    if p.is_dir():
        return False, f"That is a directory, not a file."
    if not p.exists():
        return False, f"File not found: `{file_path}`"
    if extensions and p.suffix.lower() not in [e.lower() for e in extensions]:
        return False, f"Expected {' or '.join(extensions)}, got `{p.suffix}`"
    size_kb = round(p.stat().st_size / 1024, 2)
    return True, f"✅ `{p.name}` — {size_kb:,.1f} KB"


def _render_catalog(engine):
    # Auto-jump to UWI Mapping if navigated here from DLIS/LIS tab
    # Use persistent flag (not pop) so crawl/scan reruns stay in mapping
    if st.session_state.get("wl_goto_mapping"):
        col1, col2 = st.columns([6, 1])
        col1.info("📍 **File → UWI Mapping** — crawl, match and bulk catalog DLIS/LIS files.")
        if col2.button("✕ Exit", key="wl_exit_mapping"):
            st.session_state.pop("wl_goto_mapping", None)
            st.rerun()
        _render_mapping(engine)
        return

    st.subheader("Catalog well log files")

    try:
        repos = list_repositories(engine)
    except Exception as e:
        st.error(str(e)); return

    if repos.empty:
        st.warning("No repositories registered. Add one in **📁 Repositories** first.")
        return

    col_repo, col_src = st.columns([3, 1])
    with col_repo:
        repo_opts = {
            f"{r['REPOSITORY_NAME']} ({r['BASE_PATH']})": r["REPOSITORY_ID"]
            for _, r in repos.iterrows()
        }
        repo_label    = st.selectbox("Repository", list(repo_opts.keys()), key="cat_repo_sel")
        repository_id = repo_opts[repo_label]
        base_path     = repos[repos["REPOSITORY_ID"] == repository_id]["BASE_PATH"].iloc[0]
    with col_src:
        source = st.text_input("Source", value="DATA_WRANGLER", key="cat_source")

    st.divider()
    fmt_las, fmt_dlis, fmt_lis = st.tabs(["📄 LAS", "📦 DLIS", "📋 LIS"])

    with fmt_las:
        s1, s2 = st.tabs(["Single file", "Directory"])
        with s1: _render_catalog_las_single(engine, repository_id, base_path, source)
        with s2: _render_catalog_las_directory(engine, repository_id, base_path, source)

    with fmt_dlis:
        s1, s2 = st.tabs(["Single file", "Directory"])
        with s1: _render_catalog_dlis_single(engine, repository_id, base_path, source)
        with s2: _render_catalog_dlis_directory(engine, repository_id, base_path, source)

    with fmt_lis:
        s1, s2 = st.tabs(["Single file", "Directory"])
        with s1: _render_catalog_lis_single(engine, repository_id, base_path, source)
        with s2: _render_catalog_lis_directory(engine, repository_id, base_path, source)


def _render_catalog_las_single(engine, repository_id, base_path, source):
    st.markdown("**Catalog a single LAS file**")
    file_path = st.text_input("Full file path", key="cat_las_path",
                               placeholder=r"e.g. C:\WellLogs\well_01.las")
    if not file_path: return
    valid, msg = _validate_file_path(file_path, [".las", ".LAS"])
    if not valid: st.error(msg); return
    st.caption(msg)

    if st.button("🔍 Read header", key="cat_las_preview_btn"):
        try:
            hdr = parse_las_header(file_path)
            st.session_state["cat_las_hdr"]      = hdr
            st.session_state["cat_las_hdr_path"] = file_path
        except Exception as e:
            st.error(f"Could not parse: {e}"); return

    if "cat_las_hdr" not in st.session_state: return
    if st.session_state.get("cat_las_hdr_path") != file_path:
        st.session_state.pop("cat_las_hdr", None); return

    hdr = st.session_state["cat_las_hdr"]
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Version",    hdr.get("version","—"))
    c2.metric("Curves",     hdr.get("curve_count","—"))
    c3.metric("Top depth",  hdr.get("top_depth","—"))
    c4.metric("Base depth", hdr.get("base_depth","—"))

    for k,v in {"Well": hdr.get("well_name",""), "UWI": hdr.get("uwi",""),
                "Operator": hdr.get("operator",""), "Field": hdr.get("field","")}.items():
        st.text(f"{k}: {v or '—'}")

    uwi_override = st.text_input("UWI override", value=hdr.get("uwi",""), key="cat_las_uwi")

    if st.button("📥 Catalog LAS file", type="primary", key="cat_las_btn"):
        uwi = (uwi_override.strip() or hdr.get("uwi","")).strip()
        if not uwi: st.error("UWI is required."); return
        try:
            r = catalog_file(engine, file_path, uwi, repository_id, source=source)
            if r.get("ok"):
                st.success(f"✅ {r.get('action','').capitalize()} — {uwi}")
            else:
                st.error(r.get("error","Unknown error"))
        except Exception as e:
            st.error(str(e))


def _show_catalog_btn(engine, fmt: str, table: str, pk: str,
                      name_col: str = "FILE_NAME",
                      extra_cols: list = None, key_suffix: str = ""):
    """Show a collapsible catalog summary for a given format table."""
    from sqlalchemy import text as _t
    if st.button(f"📋 Show {fmt} catalog", key=f"show_cat_{fmt}_{key_suffix}",
                 use_container_width=False):
        st.session_state[f"show_cat_open_{fmt}_{key_suffix}"] = True

    if st.session_state.get(f"show_cat_open_{fmt}_{key_suffix}"):
        try:
            cols = extra_cols or [name_col]
            col_sql = ", ".join(f"[{c}]" for c in cols) + ", [FILE_SIZE_KB], [CATALOG_DATE]"
            with engine.connect() as con:
                rows = con.execute(_t(
                    f"SELECT TOP 500 {col_sql} "
                    f"FROM [las_catalog].[{table}] "
                    f"ORDER BY CATALOG_DATE DESC"
                )).fetchall()
            if rows:
                df_cols = cols + ["FILE_SIZE_KB", "CATALOG_DATE"]
                df = pd.DataFrame(rows, columns=df_cols)
                df["FILE_SIZE_KB"] = (df["FILE_SIZE_KB"].fillna(0) / 1024).round(2)
                df = df.rename(columns={"FILE_SIZE_KB": "Size (MB)"})
                st.caption(f"{len(rows):,} {fmt} file(s) in catalog")
                st.dataframe(df, hide_index=True, use_container_width=True)
            else:
                st.info(f"No {fmt} files cataloged yet.")
        except Exception as e:
            st.error(str(e))
        if st.button("✕ Close", key=f"close_cat_{fmt}_{key_suffix}"):
            st.session_state.pop(f"show_cat_open_{fmt}_{key_suffix}", None)
            st.rerun()


def _render_catalog_las_directory(engine, repository_id, base_path, source):
    st.markdown("**Catalog a directory of LAS files**")
    folder = st.text_input("Directory path", value=base_path, key="cat_las_dir",
                            placeholder=r"e.g. C:\WellLogs\LAS")
    if not folder: return
    try:
        from pathlib import Path
        files = [f for f in Path(folder).rglob("*")
                 if f.is_file() and f.suffix.lower() == ".las"]
        st.info(f"{len(files)} LAS file(s) found.")
    except Exception: files = []

    if st.button(f"🚀 Catalog {len(files)} LAS file(s)", type="primary",
                  key="cat_las_dir_btn", disabled=len(files)==0):
        prog = st.progress(0, text="Starting…")
        def _cb(i, t, n): prog.progress(i/t, text=f"{n} ({i}/{t})…")
        try:
            results = catalog_directory(engine, folder, repository_id,
                                         source=source, progress_callback=_cb)
            prog.empty()
            ins = sum(1 for r in results if r.get("action")=="inserted")
            upd = sum(1 for r in results if r.get("action")=="updated")
            err = sum(1 for r in results if not r.get("ok"))
            st.success(f"✅ {ins} new · {upd} updated · {err} error(s)")
            if err:
                with st.expander(f"❌ {err} error(s)"):
                    st.dataframe(pd.DataFrame([
                        {"File": r.get("file_name",""), "Error": r.get("error","")}
                        for r in results if not r.get("ok")
                    ]), hide_index=True)
        except Exception as e:
            st.error(str(e))


def _render_catalog_dlis_single(engine, repository_id, base_path, source):
    from modules.dlis_catalog import catalog_dlis_file, parse_dlis_header
    st.markdown("**Catalog a single DLIS file**")
    file_path = st.text_input("Full file path", key="cat_dlis_path",
                               placeholder=r"e.g. C:\WellLogs\well_01.dlis")
    if not file_path: return
    valid, msg = _validate_file_path(file_path, [".dlis", ".DLIS"])
    if not valid: st.error(msg); return
    st.caption(msg)

    if st.button("🔍 Preview header", key="cat_dlis_preview_btn"):
        try:
            hdr = parse_dlis_header(file_path)
            st.session_state["cat_dlis_hdr"]      = hdr
            st.session_state["cat_dlis_hdr_path"] = file_path
        except Exception as e:
            st.error(f"Could not parse: {e}"); return

    if "cat_dlis_hdr" not in st.session_state: return
    if st.session_state.get("cat_dlis_hdr_path") != file_path:
        st.session_state.pop("cat_dlis_hdr", None); return

    hdr = st.session_state["cat_dlis_hdr"]
    for k,v in {"Well": hdr.get("well_name",""), "Well ID": hdr.get("well_id",""),
                "Company": hdr.get("company",""), "Field": hdr.get("field_name","")}.items():
        st.text(f"{k}: {v or '—'}")

    uwi = st.text_input("UWI", key="cat_dlis_uwi")
    if st.button("📥 Catalog DLIS file", type="primary", key="cat_dlis_btn"):
        if not uwi.strip(): st.error("UWI is required."); return
        try:
            r = catalog_dlis_file(engine, file_path, uwi.strip(), repository_id, source=source)
            if r.get("ok"):
                st.success(f"✅ {r.get('action','').capitalize()} — {uwi}")
            else:
                st.error(r.get("error","Unknown error"))
        except Exception as e:
            st.error(str(e))


def _render_catalog_dlis_directory(engine, repository_id, base_path, source):
    st.info(
        "DLIS files rarely contain a reliable UWI in their headers. "
        "Use **File → UWI Mapping** in the Well Logs section to fuzzy-match "
        "files to PPDM wells, confirm matches, and bulk catalog."
    )
    if st.button("Open File → UWI Mapping →", type="primary",
                  key="cat_dlis_dir_goto_map", on_click=_goto_wl_map):
        pass


def _render_catalog_lis_single(engine, repository_id, base_path, source):
    from modules.dlis_catalog import catalog_lis_file, parse_lis_header
    st.markdown("**Catalog a single LIS file**")
    file_path = st.text_input("Full file path", key="cat_lis_path",
                               placeholder=r"e.g. C:\WellLogs\well_01.lis")
    if not file_path: return
    valid, msg = _validate_file_path(file_path, [".lis", ".LIS"])
    if not valid: st.error(msg); return
    st.caption(msg)

    if st.button("🔍 Preview header", key="cat_lis_preview_btn"):
        try:
            hdr = parse_lis_header(file_path)
            st.session_state["cat_lis_hdr"]      = hdr
            st.session_state["cat_lis_hdr_path"] = file_path
        except Exception as e:
            st.error(f"Could not parse: {e}"); return

    if "cat_lis_hdr" not in st.session_state: return
    if st.session_state.get("cat_lis_hdr_path") != file_path:
        st.session_state.pop("cat_lis_hdr", None); return

    hdr = st.session_state["cat_lis_hdr"]
    for k,v in {"Well": hdr.get("well_name",""), "Company": hdr.get("company","")}.items():
        st.text(f"{k}: {v or '—'}")

    uwi = st.text_input("UWI", key="cat_lis_uwi")
    if st.button("📥 Catalog LIS file", type="primary", key="cat_lis_btn"):
        if not uwi.strip(): st.error("UWI is required."); return
        try:
            r = catalog_lis_file(engine, file_path, uwi.strip(), repository_id, source=source)
            if r.get("ok"):
                st.success(f"✅ {r.get('action','').capitalize()} — {uwi}")
            else:
                st.error(r.get("error","Unknown error"))
        except Exception as e:
            st.error(str(e))


def _render_catalog_lis_directory(engine, repository_id, base_path, source):
    st.info(
        "LIS files rarely contain a reliable UWI.\n\n"
        "Use **File → UWI Mapping** in the Well Logs section to scan, match, and bulk catalog LIS files."
    )
    if st.session_state.get("cat_goto_map"):
        st.session_state.pop("cat_goto_map", None)

    else:
        if st.button("Open File → UWI Mapping →", type="primary",
                      key="cat_lis_dir_goto_map", on_click=_goto_wl_map):
            pass


def _render_mapping(engine):
    st.subheader("File → UWI Mapping")
    st.caption(
        "Stage DLIS, LIS, and LAS files → assign UWIs → bulk catalog. "
        "Files are never modified."
    )
    tab_crawl, tab_scan, tab_manifest, tab_review = st.tabs([
        "🕷 Crawl", "🔍 Scan", "📋 Manifest", "✅ Review & Catalog"
    ])
    with tab_crawl:
        _render_crawl(engine)
    with tab_scan:
        _render_scan(engine)
    with tab_manifest:
        _render_manifest(engine)
    with tab_review:
        _render_review(engine)



def _build_crawl_excel(file_list: list[dict]) -> bytes:
    """
    Build an Excel workbook from crawl results.

    Sheet 1 — Summary: count and total size by parent directory and file type
    Sheet 2 — Full listing: every file found
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("pip install openpyxl")

    df_full = pd.DataFrame(file_list)
    if df_full.empty:
        raise ValueError("No files to export.")

    # ── Summary sheet ─────────────────────────────────────────────────
    summary = (
        df_full.groupby(["PARENT", "FILE_TYPE"])
        .agg(
            FILE_COUNT=("FILE_NAME", "count"),
            TOTAL_SIZE_KB=("SIZE_KB", "sum"),
        )
        .reset_index()
        .sort_values(["PARENT", "FILE_TYPE"])
    )
    # Add totals row per parent
    parent_totals = (
        summary.groupby("PARENT")
        .agg(FILE_COUNT=("FILE_COUNT","sum"),
             TOTAL_SIZE_KB=("TOTAL_SIZE_KB","sum"))
        .reset_index()
        .assign(FILE_TYPE="ALL")
    )
    grand_total = pd.DataFrame([{
        "PARENT": "GRAND TOTAL",
        "FILE_TYPE": "",
        "FILE_COUNT": summary["FILE_COUNT"].sum(),
        "TOTAL_SIZE_KB": summary["TOTAL_SIZE_KB"].sum(),
    }])

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    alt_fill    = PatternFill("solid", fgColor="DCE6F1")
    total_fill  = PatternFill("solid", fgColor="BDD7EE")
    grand_fill  = PatternFill("solid", fgColor="1F4E79")
    grand_font  = Font(bold=True, color="FFFFFF")

    s1_headers = ["Parent Directory", "File Type", "File Count", "Total Size (KB)"]
    for col, h in enumerate(s1_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for _, r in summary.iterrows():
        ws1.cell(row=row, column=1, value=r["PARENT"])
        ws1.cell(row=row, column=2, value=r["FILE_TYPE"])
        ws1.cell(row=row, column=3, value=int(r["FILE_COUNT"]))
        ws1.cell(row=row, column=4, value=round(r["TOTAL_SIZE_KB"], 2))
        if row % 2 == 0:
            for col in range(1, 5):
                ws1.cell(row=row, column=col).fill = alt_fill
        row += 1

    # Grand total row
    ws1.cell(row=row, column=1, value="GRAND TOTAL").font = grand_font
    ws1.cell(row=row, column=1).fill = grand_fill
    ws1.cell(row=row, column=2, value="").fill = grand_fill
    ws1.cell(row=row, column=3, value=int(grand_total["FILE_COUNT"].iloc[0])).font = grand_font
    ws1.cell(row=row, column=3).fill = grand_fill
    ws1.cell(row=row, column=4, value=round(grand_total["TOTAL_SIZE_KB"].iloc[0], 2)).font = grand_font
    ws1.cell(row=row, column=4).fill = grand_fill

    # Auto-width
    for col in range(1, 5):
        col_letter = get_column_letter(col)
        max_len = max(
            len(str(ws1.cell(r, col).value or ""))
            for r in range(1, row + 1)
        )
        ws1.column_dimensions[col_letter].width = min(max_len + 4, 80)

    # ── Sheet 2: Full Listing ─────────────────────────────────────────
    ws2 = wb.create_sheet("Full Listing")
    s2_headers = ["Full Path", "Parent Directory", "File Name", "File Type", "Size (KB)"]
    for col, h in enumerate(s2_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(file_list, 2):
        ws2.cell(row=row_idx, column=1, value=r["FULL_PATH"])
        ws2.cell(row=row_idx, column=2, value=r["PARENT"])
        ws2.cell(row=row_idx, column=3, value=r["FILE_NAME"])
        ws2.cell(row=row_idx, column=4, value=r["FILE_TYPE"])
        ws2.cell(row=row_idx, column=5, value=r["SIZE_KB"])
        if row_idx % 2 == 0:
            for col in range(1, 6):
                ws2.cell(row=row_idx, column=col).fill = alt_fill

    for col in range(1, 6):
        col_letter = get_column_letter(col)
        max_len = max(
            len(str(ws2.cell(r, col).value or ""))
            for r in range(1, min(len(file_list) + 2, 200))
        )
        ws2.column_dimensions[col_letter].width = min(max_len + 4, 100)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_crawl(engine):
    st.markdown("**Crawl filesystem for well log files**")
    st.caption(
        "Phase 1 walks directories quickly — no header reading. "
        "Review the results, select files to process, then Phase 2 "
        "extracts headers and matches UWIs."
    )

    col1, col2 = st.columns([3, 1])
    with col1:
        root = st.text_input(
            "Root directory", key="crawl_root",
            placeholder=r"e.g. C:\WellLogs  or  \\\\server\\data",
        )
    with col2:
        max_workers = st.select_slider(
            "Threads", options=[1, 2, 4, 6, 8, 10, 12, 16],
            value=__import__("os").cpu_count() and min(max(__import__("os").cpu_count()-2,2),12) or 4,
            key="crawl_workers"
        )

    # Format selector
    st.caption("Well log formats")
    fmt_cols = st.columns(3)
    with fmt_cols[0]: find_las  = st.checkbox("LAS",  value=True,  key="crawl_las")
    with fmt_cols[1]: find_dlis = st.checkbox("DLIS", value=True,  key="crawl_dlis")
    with fmt_cols[2]: find_lis  = st.checkbox("LIS",  value=True,  key="crawl_lis")

    st.caption("Seismic / positioning formats")
    fmt_cols2 = st.columns(3)
    with fmt_cols2[0]: find_segy = st.checkbox("SEG-Y (.segy, .sgy)", value=False, key="crawl_segy")
    with fmt_cols2[1]: find_p190 = st.checkbox("P1/90 (.p190, .p90)", value=False, key="crawl_p190")

    formats = [f for f, v in [
        ("LAS", find_las), ("DLIS", find_dlis), ("LIS", find_lis),
        ("SEGY", find_segy), ("P190", find_p190),
    ] if v]

    # Repository selector
    try:
        repos = list_repositories(engine)
        if not repos.empty:
            repo_opts = {"(none — assign later)": ""} | {
                f"{r['REPOSITORY_NAME']} ({r['BASE_PATH']})": r["REPOSITORY_ID"]
                for _, r in repos.iterrows()
            }
            selected_repo = st.selectbox(
                "Repository", options=list(repo_opts.keys()),
                key="crawl_repo"
            )
            repository_id = repo_opts[selected_repo]
        else:
            st.warning("No repositories registered.")
            repository_id = ""
    except Exception as e:
        st.error(str(e))
        return

    if not root or not formats:
        return

    # Warn if root looks like a drive root
    from pathlib import Path as _Path
    root_path = _Path(root)
    if str(root_path) in (root_path.anchor, root_path.anchor.rstrip("\\/")):
        st.warning(
            f"⚠️ `{root}` is a drive root — this may take a long time. "
            "Consider a more targeted path like `C:\\WellLogs`."
        )

    # ── Phase 1: Walk ─────────────────────────────────────────────────
    if "crawl_abort" not in st.session_state:
        st.session_state["crawl_abort"] = [False]

    col_walk, col_abort = st.columns([1, 1])
    with col_walk:
        walk_btn = st.button(
            "🔍 Phase 1 — Find files",
            type="primary", key="crawl_walk_btn",
            disabled=not root,
        )
    with col_abort:
        if st.button("⛔ Abort", key="crawl_abort_btn",
                     disabled="crawl_walking" not in st.session_state):
            st.session_state["crawl_abort"][0] = True

    if walk_btn:
        st.session_state["crawl_abort"] = [False]
        st.session_state["crawl_walking"] = True
        st.session_state.pop("crawl_results", None)

        walk_status = st.empty()
        walk_status.info("Walking directories…")

        try:
            from modules.wl_file_map import crawl_walk
            result = crawl_walk(
                root, formats, engine,
                abort_flag=st.session_state["crawl_abort"],
            )
            st.session_state["crawl_results"] = result
        except Exception as e:
            st.error(str(e))
            del st.session_state["crawl_walking"]
            return

        del st.session_state["crawl_walking"]
        walk_status.empty()

    # ── Results grid ──────────────────────────────────────────────────
    if "crawl_results" not in st.session_state:
        return

    result = st.session_state["crawl_results"]
    file_list = result.get("file_list", [])

    # Summary metrics
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Found",            f"{len(file_list):,}")
    m2.metric("Already staged",   f"{result.get('skipped_existing', 0):,}")
    m3.metric("Already catalogued", f"{result.get('skipped_catalogued', 0):,}")
    m4.metric("Aborted",          "Yes" if result.get("aborted") else "No")

    if not file_list:
        st.info("No new files found.")
        return

    # Format breakdown
    fmt_counts = pd.DataFrame(file_list).groupby("FILE_TYPE").agg(
        Files=("FILE_NAME", "count"),
        Size_MB=("SIZE_KB", lambda x: round(x.sum() / 1024, 1))
    ).reset_index()
    st.dataframe(fmt_counts, hide_index=True)

    st.divider()
    st.markdown("**Select files to process**")
    st.caption(
        "Check the rows you want to extract headers from and stage for cataloguing. "
        "Use the format filter to narrow the list."
    )

    # Format filter
    fmt_filter = st.selectbox(
        "Show format", options=["All"] + list(set(r["FILE_TYPE"] for r in file_list)),
        key="crawl_fmt_filter"
    )
    display_list = file_list if fmt_filter == "All" else [
        r for r in file_list if r["FILE_TYPE"] == fmt_filter
    ]

    # Selection state — keyed by FULL_PATH so it survives reruns
    if "crawl_selected" not in st.session_state:
        st.session_state["crawl_selected"] = set()

    # Select All / Deselect All buttons — update session state before building grid
    _col_sa, _col_da, _col_sp = st.columns([1, 1, 4])
    with _col_sa:
        if st.button("☑ Select all", key="crawl_sel_all_top"):
            st.session_state["crawl_selected"] = {
                r["FULL_PATH"] for r in display_list
            }
            st.rerun()
    with _col_da:
        if st.button("☐ Deselect all", key="crawl_desel_all_top"):
            st.session_state["crawl_selected"] = set()
            st.rerun()

    # Build display dataframe — restore SELECT from session state
    display_df = pd.DataFrame(display_list)
    display_df.insert(0, "SELECT",
        display_df["FULL_PATH"].isin(st.session_state["crawl_selected"])
    )

    edited = st.data_editor(
        display_df,
        hide_index=True,
        column_config={
            "SELECT":     st.column_config.CheckboxColumn("Select", width="small"),
            "FULL_PATH":  st.column_config.Column("Full Path",   disabled=True),
            "PARENT":     st.column_config.Column("Directory",   disabled=True),
            "FILE_NAME":  st.column_config.Column("File",        disabled=True, width="medium"),
            "FILE_TYPE":  st.column_config.Column("Type",        disabled=True, width="small"),
            "SIZE_KB":    st.column_config.NumberColumn("Size KB", disabled=True, width="small"),
        },
        key="crawl_grid",
    )

    # Persist checkbox changes back into session state
    for _, row in edited.iterrows():
        path = row["FULL_PATH"]
        if row["SELECT"]:
            st.session_state["crawl_selected"].add(path)
        else:
            st.session_state["crawl_selected"].discard(path)

    selected_rows = edited[edited["SELECT"] == True].drop(columns=["SELECT"])
    n_selected = len(selected_rows)

    col_selall, col_desel, col_process, col_abort2 = st.columns([1, 1, 2, 1])

    with col_selall:
        pass  # select all moved above the grid

    with col_process:
        process_btn = st.button(
            f"⚙ Phase 2 — Process {n_selected} selected file(s)",
            type="primary",
            key="crawl_process_btn",
            disabled=n_selected == 0,
        )

    with col_abort2:
        if st.button("⛔ Abort processing", key="crawl_abort2_btn",
                     disabled="crawl_processing" not in st.session_state):
            st.session_state["crawl_abort"][0] = True

    # ── Phase 2: Process selected ────────────────────────────────────
    if process_btn and n_selected > 0:
        st.session_state["crawl_abort"] = [False]
        st.session_state["crawl_processing"] = True

        with st.spinner("Loading PPDM wells…"):
            try:
                ppdm_uwis = fetch_ppdm_uwis(engine)
            except Exception as e:
                st.error(str(e))
                del st.session_state["crawl_processing"]
                return

        progress_bar = st.progress(0, text="Starting…")
        p_status     = st.empty()

        def _proc_progress(done, total, staged):
            pct = min(done / total, 1.0) if total > 0 else 0
            progress_bar.progress(
                pct,
                text=f"Extracting headers… {done:,} / {total:,} files  |  {staged:,} staged"
            )

        selected_dicts = selected_rows.to_dict("records")

        try:
            from modules.wl_file_map import crawl_process
            proc_result = crawl_process(
                selected_dicts, engine, ppdm_uwis,
                repository_id=repository_id,
                max_workers=max_workers,
                progress_callback=_proc_progress,
                abort_flag=st.session_state["crawl_abort"],
            )
        except Exception as e:
            st.error(str(e))
            del st.session_state["crawl_processing"]
            return

        del st.session_state["crawl_processing"]
        progress_bar.empty()
        p_status.empty()

        if proc_result.get("aborted"):
            st.warning(
                f"⛔ Processing aborted — "
                f"{proc_result['saved']:,} staged · {proc_result['errors']} error(s)"
            )
        else:
            st.success(
                f"✅ Done — {proc_result['saved']:,} file(s) staged · "
                f"{proc_result['errors']} error(s)"
            )

        if proc_result.get("errors", 0) > 0 and proc_result.get("error_details"):
            with st.expander(f"❌ {proc_result['errors']} error(s)", expanded=True):
                for err in proc_result["error_details"][:20]:
                    st.text(err)

        st.info("Switch to **✅ Review & Catalog** to assign UWIs and catalog.")

    # ── Excel export ──────────────────────────────────────────────────
    if file_list:
        st.divider()
        col_xl, _ = st.columns([1, 2])
        with col_xl:
            try:
                xl_bytes = _build_crawl_excel(file_list)
                st.download_button(
                    label=f"⬇ Export {len(file_list):,} files to Excel",
                    data=xl_bytes,
                    file_name="crawl_results.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="crawl_xl_btn",
                )
            except ImportError:
                st.warning("Install openpyxl: `pip install openpyxl`")
            except Exception as e:
                st.error(str(e))


def _render_scan(engine):
    st.subheader("Scan directory")
    st.markdown(
        "Scans a folder for well log files, extracts header well IDs, "
        "and fuzzy-matches against PPDM wells. Review matches in the "
        "**Review & Catalog** tab."
    )

    try:
        repos = list_repositories(engine)
    except Exception as e:
        st.error(str(e))
        return

    col1, col2 = st.columns([3, 1])
    with col1:
        folder = st.text_input(
            "Directory path", key="map_scan_folder",
            placeholder=r"e.g. C:\Data\DLIS_Delivery",
        )
    with col2:
        fmt = st.selectbox("Format", ["DLIS", "LIS", "LAS"], key="map_scan_fmt")

    # Repository selector
    if not repos.empty:
        repo_options = {
            f"{r['REPOSITORY_NAME']} ({r['BASE_PATH']})": r["REPOSITORY_ID"]
            for _, r in repos.iterrows()
        }
        repo_options = {"(none — assign later)": ""} | repo_options
        selected_repo = st.selectbox(
            "Repository", options=list(repo_options.keys()),
            key="map_scan_repo"
        )
        repository_id = repo_options[selected_repo]
    else:
        st.warning("No repositories registered. Register one in the Well Log Catalog.")
        repository_id = ""

    if not folder:
        return

    # Preview file count
    try:
        from pathlib import Path
        ext_map = {
            "DLIS": ["*.dlis", "*.DLIS"],
            "LIS":  ["*.lis",  "*.LIS"],
            "LAS":  ["*.las",  "*.LAS"],
        }
        files = []
        for pat in ext_map.get(fmt, []):
            files.extend(Path(folder).glob(pat))
        st.info(f"{len(files)} {fmt} file(s) found in directory.")
    except Exception:
        files = []

    max_workers = st.select_slider(
        "Parallel threads", options=[1, 2, 4, 6, 8, 10, 12, 16],
        value=__import__("os").cpu_count() and min(max(__import__("os").cpu_count()-2,2),12) or 4,
        key="map_scan_workers",
        help="More threads = faster scanning but higher memory use. "
             "4 is a good default for most machines."
    )

    if st.button(
        f"🔍 Scan {len(files)} file(s)",
        type="primary", key="map_scan_btn",
        disabled=len(files) == 0,
    ):
        with st.spinner("Loading PPDM wells…"):
            try:
                ppdm_uwis = fetch_ppdm_uwis(engine)
            except Exception as e:
                st.error(str(e))
                return

        progress = st.progress(0, text="Starting scan…")
        status   = st.empty()

        def _scan_progress(completed, total, filename):
            progress.progress(
                min(completed / total, 1.0),
                text=f"Scanning {filename} ({completed}/{total})…"
            )
            status.caption(filename)

        try:
            scan_df = scan_directory(
                folder, fmt, ppdm_uwis,
                repository_id=repository_id,
                max_workers=max_workers,
                progress_callback=_scan_progress,
            )
            progress.progress(1.0, text="Done.")
            status.empty()
        except Exception as e:
            st.error(str(e))
            return

        if scan_df.empty:
            st.info("No files found.")
            return

        # Save to staging
        try:
            scan_df = scan_df.where(scan_df.notna(), other=None)  # replace NaN → None
            saved = save_map(engine, scan_df)
            st.session_state["map_scan_preview"] = scan_df
        except Exception as e:
            st.error(str(e))
            return

        matched   = scan_df["UWI"].notna().sum()
        unmatched = scan_df["UWI"].isna().sum()
        st.success(
            f"✅ {saved} file(s) added to staging  |  "
            f"{matched} matched  |  {unmatched} need manual UWI"
        )

    # Preview with inline override
    if "map_scan_preview" in st.session_state:
        preview = st.session_state["map_scan_preview"]

        matched   = preview["UWI"].notna().sum()
        unmatched = preview["UWI"].isna().sum()

        st.divider()
        c1, c2, c3 = st.columns(3)
        c1.metric("Total files",  len(preview))
        c2.metric("Matched",      matched,   delta=None)
        c3.metric("Unmatched",    unmatched, delta=None,
                  delta_color="inverse" if unmatched > 0 else "off")

        # Split matched / unmatched for clarity
        tab_matched, tab_unmatched = st.tabs([
            f"✅ Matched ({matched})", f"❓ Unmatched ({unmatched})"
        ])

        with tab_matched:
            if matched == 0:
                st.info("No matched files yet.")
            else:
                matched_df = preview[preview["UWI"].notna()].copy()
                display_cols = [
                    "FILE_NAME", "UWI", "MATCH_WELL_NAME",
                    "MATCH_METHOD", "MATCH_SCORE", "HEADER_WELL_ID"
                ]
                st.dataframe(
                    matched_df[[c for c in display_cols if c in matched_df.columns]], hide_index=True,
                )

        with tab_unmatched:
            if unmatched == 0:
                st.success("All files matched!")
            else:
                unmatched_df = preview[preview["UWI"].isna()].copy()
                st.markdown(
                    "These files could not be matched automatically. "
                    "Assign a UWI below and click **Apply overrides** to update the staging table."
                )

                # Load PPDM UWIs for override dropdown
                try:
                    ppdm_uwis = fetch_ppdm_uwis(engine)
                    uwi_options = [""] + [r["UWI"] for r in ppdm_uwis]
                    uwi_label   = {r["UWI"]: f"{r['UWI']} — {r['WELL_NAME']}"
                                   for r in ppdm_uwis}
                except Exception:
                    uwi_options = [""]
                    uwi_label   = {}

                overrides = {}
                for _, row in unmatched_df.iterrows():
                    col_fn, col_hi, col_sel = st.columns([3, 2, 3])
                    with col_fn:
                        st.text(row["FILE_NAME"])
                    with col_hi:
                        st.caption(f"Header: {row.get('HEADER_WELL_ID') or '—'}")
                    with col_sel:
                        chosen = st.selectbox(
                            "UWI",
                            options=uwi_options,
                            format_func=lambda u: uwi_label.get(u, u) if u else "— select —",
                            key=f"map_ovr_{row['MAP_ID']}",
                            label_visibility="collapsed",
                        )
                        if chosen:
                            overrides[row["MAP_ID"]] = {
                                "uwi":       chosen,
                                "well_name": ppdm_uwis[[r["UWI"] for r in ppdm_uwis].index(chosen)]["WELL_NAME"]
                                             if chosen in [r["UWI"] for r in ppdm_uwis] else "",
                            }

                if overrides and st.button(
                    f"✔ Apply {len(overrides)} override(s)",
                    type="primary", key="map_apply_overrides_btn"
                ):
                    try:
                        for mid, vals in overrides.items():
                            update_uwi(engine, mid, vals["uwi"], vals["well_name"])
                        st.success(f"✅ {len(overrides)} UWI(s) assigned.")
                        # Refresh preview from DB
                        updated = load_map(engine)
                        st.session_state["map_scan_preview"] = updated[
                            updated["FILE_NAME"].isin(preview["FILE_NAME"])
                        ].reset_index(drop=True)
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))


# ─────────────────────────────────────────────────────────────────────────────
# MANIFEST TAB
# ─────────────────────────────────────────────────────────────────────────────

def _render_manifest(engine):
    st.subheader("Import manifest")
    st.markdown(
        "Upload a CSV or Excel file that maps filenames to UWIs. "
        "Manifest rows are marked CONFIRMED immediately — no fuzzy matching needed."
    )

    try:
        repos = list_repositories(engine)
    except Exception as e:
        st.error(str(e))
        return

    uploaded = st.file_uploader(
        "Manifest file", type=["csv", "xlsx", "xls"],
        key="map_manifest_upload"
    )

    if uploaded is None:
        st.info("Upload a CSV or Excel file with at least a filename column and a UWI column.")
        return

    # Preview the file to let user pick columns
    try:
        import tempfile, os
        suffix = Path(uploaded.name).suffix
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(uploaded.read())
            tmp_path = tmp.name

        if suffix.lower() in (".xlsx", ".xls"):
            preview_df = pd.read_excel(tmp_path, nrows=5)
        else:
            preview_df = pd.read_csv(tmp_path, nrows=5)

        st.markdown("**Preview (first 5 rows):**")
        st.dataframe(preview_df, hide_index=True)

        cols = list(preview_df.columns)
        col1, col2, col3 = st.columns(3)
        with col1:
            filename_col = st.selectbox("Filename column", cols, key="map_fn_col")
        with col2:
            uwi_col = st.selectbox("UWI column", cols, key="map_uwi_col")
        with col3:
            fmt = st.selectbox("Format", ["DLIS", "LIS", "LAS"], key="map_mfmt")

        folder = st.text_input(
            "Files base directory", key="map_manifest_folder",
            placeholder=r"e.g. C:\Data\DLIS_Delivery",
        )

        # Repository
        if not repos.empty:
            repo_options = {"(none — assign later)": ""} | {
                f"{r['REPOSITORY_NAME']} ({r['BASE_PATH']})": r["REPOSITORY_ID"]
                for _, r in repos.iterrows()
            }
            selected_repo = st.selectbox(
                "Repository", options=list(repo_options.keys()),
                key="map_manifest_repo"
            )
            repository_id = repo_options[selected_repo]
        else:
            repository_id = ""

        if st.button("📥 Import manifest", type="primary",
                      key="map_manifest_btn",
                      disabled=not folder):
            with st.spinner("Importing…"):
                try:
                    map_df = import_manifest(
                        tmp_path, filename_col, uwi_col,
                        folder, fmt, repository_id,
                    )
                    map_df = map_df.where(map_df.notna(), other=None)
                    saved = save_map(engine, map_df)
                    st.success(
                        f"✅ {saved} row(s) imported — "
                        f"all marked CONFIRMED. Review in the Review tab."
                    )
                except Exception as e:
                    st.error(str(e))

        os.unlink(tmp_path)

    except Exception as e:
        st.error(f"Could not read manifest: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# REVIEW & CATALOG TAB
# ─────────────────────────────────────────────────────────────────────────────

def _render_review(engine):
    st.subheader("Review & Catalog")
    from modules.wl_file_map import _now_str

    # Filters
    col1, col2, col3 = st.columns(3)
    with col1:
        fmt_filter = st.selectbox(
            "Format", ["All", "DLIS", "LIS", "LAS"], key="map_rev_fmt"
        )
    with col2:
        status_filter = st.selectbox(
            "Status", ["All", "PENDING", "CONFIRMED", "SKIPPED", "CATALOGUED"],
            key="map_rev_status"
        )
    with col3:
        since_filter = st.selectbox(
            "Added", ["All time", "Today", "Last hour"],
            key="map_rev_since"
        )

    fmt_arg    = "" if fmt_filter    == "All" else fmt_filter
    status_arg = "" if status_filter == "All" else status_filter

    # Build since datetime string
    since_arg = ""
    if since_filter == "Today":
        from datetime import datetime, timezone
        since_arg = datetime.now(timezone.utc).strftime("%Y-%m-%d") + " 00:00:00"
    elif since_filter == "Last hour":
        from datetime import datetime, timezone, timedelta
        since_arg = (datetime.now(timezone.utc) - timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")

    try:
        df = load_map(engine, fmt=fmt_arg, status=status_arg, since=since_arg)
    except Exception as e:
        st.error(str(e))
        return

    if df.empty:
        st.info("No staging rows found. Use the Crawl or Scan tab to add files.")
        return

    st.markdown(f"**{len(df)} file(s) in staging**")

    # Summary bar
    counts = df["STATUS"].value_counts().to_dict()
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Pending",    counts.get("PENDING",    0))
    c2.metric("Confirmed",  counts.get("CONFIRMED",  0))
    c3.metric("Skipped",    counts.get("SKIPPED",    0))
    c4.metric("Catalogued", counts.get("CATALOGUED", 0))

    st.divider()

    # Load PPDM UWIs for override dropdowns
    try:
        ppdm_uwis = fetch_ppdm_uwis(engine)
        uwi_options = [""] + [r["UWI"] for r in ppdm_uwis]
    except Exception:
        uwi_options = [""]

    # Editable grid
    st.markdown("**Review matches — override UWI or skip as needed:**")

    display_df = df[[
        "MAP_ID", "FILE_NAME", "FILE_FORMAT", "UWI",
        "MATCH_WELL_NAME", "MATCH_METHOD", "MATCH_SCORE",
        "HEADER_WELL_ID", "STATUS", "REMARK"
    ]].copy()

    edited = st.data_editor(
        display_df,
        hide_index=True,
        column_config={
            "MAP_ID":          st.column_config.Column(disabled=True, width="small"),
            "FILE_NAME":       st.column_config.Column(disabled=True),
            "FILE_FORMAT":     st.column_config.Column(disabled=True, width="small"),
            "MATCH_WELL_NAME": st.column_config.Column(disabled=True),
            "MATCH_METHOD":    st.column_config.Column(disabled=True, width="small"),
            "MATCH_SCORE":     st.column_config.NumberColumn(disabled=True, width="small"),
            "HEADER_WELL_ID":  st.column_config.Column(disabled=True),
            "UWI": st.column_config.SelectboxColumn(
                "UWI", options=uwi_options, width="medium"
            ),
            "STATUS": st.column_config.SelectboxColumn(
                "STATUS",
                options=["PENDING", "CONFIRMED", "SKIPPED"],
                width="small",
            ),
            "REMARK": st.column_config.TextColumn(width="medium"),
        },
        key="map_review_grid",
    )

    # Save edits button
    # ── Repository assignment ─────────────────────────────────────────
    # Required before cataloguing — rows without REPOSITORY_ID are skipped
    st.divider()
    st.markdown("**Assign repository**")

    try:
        repos = list_repositories(engine)
        if not repos.empty:
            repo_opts = {
                f"{r['REPOSITORY_NAME']} ({r['BASE_PATH']})": r["REPOSITORY_ID"]
                for _, r in repos.iterrows()
            }
            repo_id_to_name = {v: k for k, v in repo_opts.items()}

            # Show current repository assignment summary
            missing_repo = df["REPOSITORY_ID"].isna().sum()
            assigned = df["REPOSITORY_ID"].notna().sum()
            if missing_repo > 0:
                st.warning(
                    f"{missing_repo} row(s) have no repository — "
                    "they will be skipped when cataloguing."
                )
            if assigned > 0:
                current = df["REPOSITORY_ID"].dropna().mode()
                current_name = repo_id_to_name.get(
                    current.iloc[0] if not current.empty else "", "mixed"
                )
                st.caption(f"Current: {current_name}")

            col_repo, col_assign_all, col_assign_missing = st.columns([3, 1, 1])
            with col_repo:
                selected_repo_label = st.selectbox(
                    "Repository to assign", options=list(repo_opts.keys()),
                    key="map_rev_repo"
                )
            with col_assign_all:
                st.markdown("&nbsp;", unsafe_allow_html=True)
                if st.button("Assign to ALL", type="primary",
                              key="map_assign_repo_all_btn",
                              help="Overwrites any existing repository assignment"):
                    repo_id = repo_opts[selected_repo_label]
                    from sqlalchemy import text
                    with engine.begin() as con:
                        con.execute(text("""
                            UPDATE [las_catalog].[WL_FILE_UWI_MAP]
                            SET REPOSITORY_ID    = :repo,
                                ROW_CHANGED_BY   = 'DATA_WRANGLER',
                                ROW_CHANGED_DATE = :now
                            WHERE STATUS IN ('PENDING', 'CONFIRMED')
                        """), {"repo": repo_id, "now": _now_str()})
                    st.success("Repository updated for all active rows.")
                    st.rerun()
            with col_assign_missing:
                st.markdown("&nbsp;", unsafe_allow_html=True)
                if st.button("Assign missing",
                              key="map_assign_repo_missing_btn",
                              disabled=missing_repo == 0,
                              help="Only fills rows with no repository set"):
                    repo_id = repo_opts[selected_repo_label]
                    from sqlalchemy import text
                    with engine.begin() as con:
                        con.execute(text("""
                            UPDATE [las_catalog].[WL_FILE_UWI_MAP]
                            SET REPOSITORY_ID    = :repo,
                                ROW_CHANGED_BY   = 'DATA_WRANGLER',
                                ROW_CHANGED_DATE = :now
                            WHERE STATUS IN ('PENDING', 'CONFIRMED')
                              AND REPOSITORY_ID IS NULL
                        """), {"repo": repo_id, "now": _now_str()})
                    st.success(f"Repository assigned to {missing_repo} unassigned row(s).")
                    st.rerun()
        else:
            st.warning("No repositories registered. Register one in the Well Log Catalog first.")
    except Exception as e:
        st.error(str(e))

    st.divider()
    col_save, col_confirm, col_catalog, col_clear = st.columns(4)

    with col_save:
        if st.button("💾 Save edits", key="map_save_btn"):
            try:
                # Merge edits back — update UWI, STATUS, REMARK
                merged = df.copy()
                merged.set_index("MAP_ID", inplace=True)
                for _, erow in edited.iterrows():
                    mid = erow["MAP_ID"]
                    if mid in merged.index:
                        merged.at[mid, "UWI"]    = erow["UWI"]
                        merged.at[mid, "STATUS"] = erow["STATUS"]
                        merged.at[mid, "REMARK"] = erow["REMARK"]
                merged.reset_index(inplace=True)
                merged = merged.where(merged.notna(), other=None)
                saved = save_map(engine, merged)
                st.success(f"✅ {saved} row(s) saved.")
                st.rerun()
            except Exception as e:
                st.error(str(e))

    with col_confirm:
        pending_ids = df[df["STATUS"] == "PENDING"]["MAP_ID"].tolist()
        if st.button(
            f"✔ Confirm all matched ({len([i for i in pending_ids if df.loc[df['MAP_ID']==i, 'UWI'].iloc[0]])})",
            key="map_confirm_btn",
            disabled=len(pending_ids) == 0,
        ):
            try:
                # Confirm pending rows that have a UWI
                ids_with_uwi = df[
                    (df["STATUS"] == "PENDING") & df["UWI"].notna()
                ]["MAP_ID"].tolist()
                n = confirm_rows(engine, ids_with_uwi)
                st.success(f"✅ {n} row(s) confirmed.")
                st.rerun()
            except Exception as e:
                st.error(str(e))

    with col_catalog:
        confirmed_count = counts.get("CONFIRMED", 0)
        if st.button(
            f"🚀 Catalog {confirmed_count} confirmed",
            type="primary",
            key="map_catalog_btn",
            disabled=confirmed_count == 0,
        ):
            prog  = st.progress(0, text="Starting…")
            stat  = st.empty()
            def _cb_confirmed(done, total, fname):
                pct = done / total if total > 0 else 0
                prog.progress(min(pct, 1.0),
                    text=f"Cataloguing {fname} ({done}/{total})…")
                stat.caption(fname)
            try:
                result = catalog_confirmed(engine,
                                           progress_callback=_cb_confirmed)
                prog.empty()
                stat.empty()
                if result["errors"] == 0:
                    st.success(
                        f"✅ {result['catalogued']} catalogued"
                    )
                else:
                    st.warning(
                        f"{result['catalogued']} catalogued · "
                        f"{result['errors']} error(s)"
                    )
                if result["details"]:
                    with st.expander("Details", expanded=result["errors"] > 0):
                        st.dataframe(
                            pd.DataFrame(result["details"]), hide_index=True,
                        )
                st.rerun()
            except Exception as e:
                st.error(str(e))

    with col_clear:
        catalogued_count = counts.get("CATALOGUED", 0)
        if st.button(
            f"🧹 Clear {catalogued_count} catalogued",
            key="map_clear_btn",
            disabled=catalogued_count == 0,
        ):
            try:
                n = clear_catalogued(engine)
                st.success(f"Cleared {n} row(s).")
                st.rerun()
            except Exception as e:
                st.error(str(e))

    # Delete selected rows from staging
    st.divider()
    col_del, col_del_all = st.columns([1, 1])
    with col_del:
        if st.button(
            f"🗑 Delete all {len(df)} visible rows from staging",
            key="map_del_visible_btn",
            type="primary" if len(df) > 0 else "secondary",
            disabled=len(df) == 0,
            help="Removes all rows currently shown (respects Format/Status/Date filters)"
        ):
            try:
                from sqlalchemy import text
                map_ids = df["MAP_ID"].tolist()
                with engine.begin() as con:
                    for mid in map_ids:
                        con.execute(text(
                            "DELETE FROM [las_catalog].[WL_FILE_UWI_MAP] "
                            "WHERE MAP_ID = :id"
                        ), {"id": mid})
                st.success(f"Deleted {len(map_ids)} row(s) from staging.")
                st.rerun()
            except Exception as e:
                st.error(str(e))
    with col_del_all:
        if st.button(
            "🗑 Clear ALL staging rows",
            key="map_del_all_btn",
            help="Removes every row from WL_FILE_UWI_MAP regardless of filters"
        ):
            try:
                from sqlalchemy import text
                with engine.begin() as con:
                    result = con.execute(text(
                        "DELETE FROM [las_catalog].[WL_FILE_UWI_MAP]"
                    ))
                st.success(f"Cleared all staging rows.")
                st.rerun()
            except Exception as e:
                st.error(str(e))

    # ── Rename section ────────────────────────────────────────────────
    confirmed_count = counts.get("CONFIRMED", 0)
    if confirmed_count > 0:
        st.divider()
        with st.expander("✏️ Rename files with UWI prefix", expanded=False):
            st.markdown(
                "Prepends the assigned UWI to the filename before cataloguing, "
                "so the fuzzy matcher can identify the file in future scans. "
                "e.g. Chevron_A12a.DLIS becomes 17-031-10035-0000_Chevron_A12a.DLIS. "
                "Files are renamed in place. Original files are replaced."
            )

            # Preview
            try:
                preview_df = preview_rename(engine)
            except Exception as e:
                st.error(str(e))
                preview_df = pd.DataFrame()

            if not preview_df.empty:
                to_rename = preview_df[~preview_df["ALREADY_RENAMED"]]
                already   = preview_df[preview_df["ALREADY_RENAMED"]]

                if not to_rename.empty:
                    st.markdown(f"**{len(to_rename)} file(s) will be renamed:**")
                    st.dataframe(
                        to_rename[["FILE_NAME", "NEW_FILE_NAME", "UWI"]], hide_index=True,
                    )
                if not already.empty:
                    st.caption(
                        f"{len(already)} file(s) already have UWI prefix — will be skipped."
                    )

                col_dry, col_rename = st.columns([1, 1])
                with col_dry:
                    if st.button("👁 Dry run", key="map_rename_dry_btn",
                                  disabled=to_rename.empty):
                        try:
                            r = rename_files(engine, dry_run=True)
                            st.info(
                                f"Would rename {r['renamed']} file(s), "
                                f"skip {r['skipped']} already prefixed."
                            )
                        except Exception as e:
                            st.error(str(e))

                with col_rename:
                    if st.button(
                        f"✏️ Rename {len(to_rename)} file(s)",
                        type="primary",
                        key="map_rename_btn",
                        disabled=to_rename.empty,
                    ):
                        try:
                            r = rename_files(engine, dry_run=False)
                            if r["errors"] == 0:
                                st.success(
                                    f"✅ {r['renamed']} renamed, "
                                    f"{r['skipped']} skipped."
                                )
                            else:
                                st.warning(
                                    f"{r['renamed']} renamed · "
                                    f"{r['skipped']} skipped · "
                                    f"{r['errors']} error(s)"
                                )
                            if r["details"]:
                                with st.expander("Details", expanded=r["errors"] > 0):
                                    st.dataframe(
                                        pd.DataFrame(r["details"]),
                                        hide_index=True,
                                    )
                            st.rerun()
                        except Exception as e:
                            st.error(str(e))
            else:
                st.info("No confirmed rows with UWI to rename.")


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────────────────────
# SEISMIC TAB
# ─────────────────────────────────────────────────────────────────────────────

def _render_seismic(engine):
    st.subheader("Seismic File Catalog")
    # Ensure all columns exist (adds any missing from newer schema versions)
    try:
        from modules.segy_catalog import ensure_seis_catalog_columns
        added = ensure_seis_catalog_columns(engine)
        if added:
            st.info(f"Schema updated — added columns: {', '.join(added)}")
    except Exception as _e:
        st.warning(f"Schema check failed: {_e}")
    st.caption("Browse and search SEG-Y (2D and 3D) and P1/90 positioning files.")

    # Summary metrics
    try:
        segy_stats = get_segy_summary(engine)
        p190_stats = get_p190_summary(engine)

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("SEG-Y files",    f"{segy_stats['file_count']:,}")
        c2.metric("2D lines",       f"{segy_stats['count_2d']:,}")
        c3.metric("3D surveys",     f"{segy_stats['count_3d']:,}")
        c4.metric("P190 files",     f"{p190_stats['file_count']:,}")
        c5.metric("P190 shot pts",  f"{p190_stats['total_shots']:,}")
    except Exception:
        pass

    st.divider()

    tab_search, tab_surveys = st.tabs([
        "🔍 Search", "🗂 Surveys"
    ])

    with tab_search:
        _render_seis_search(engine)

    with tab_surveys:
        _render_seis_surveys(engine)



def _render_seis_ppdm_mapping(engine):
    """
    Review catalogued SEG-Y files not yet linked to PPDM SEIS_SET,
    then optionally create SEIS_SET + SEIS_LINE records and write
    SEIS_SET_ID / SEIS_LINE_ID back to SEIS_FILE_CATALOG.
    """
    st.subheader("PPDM Mapping")
    st.caption(
        "Link catalogued SEG-Y files to PPDM **dbo.SEIS_SET** and **dbo.SEIS_LINE**. "
        "Files are grouped by survey name. You can create new PPDM records or link "
        "to existing ones."
    )

    from sqlalchemy import text
    from datetime import datetime, timezone

    # ── Check PPDM tables exist ───────────────────────────────────────────────
    try:
        with engine.connect() as con:
            for tbl in ("SEIS_SET", "SEIS_LINE"):
                exists = con.execute(text(
                    "SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES "
                    "WHERE TABLE_SCHEMA='dbo' AND TABLE_NAME=:t"
                ), {"t": tbl}).scalar()
                if not exists:
                    st.error(
                        f"**dbo.{tbl}** does not exist in your PPDM database. "
                        "Run the PPDM 3.9 DDL script first."
                    )
                    return
    except Exception as e:
        st.error(f"Cannot check PPDM tables: {e}")
        return

    # ── Load unlinked files grouped by survey ─────────────────────────────────
    try:
        with engine.connect() as con:
            rows = con.execute(text("""
                SELECT
                    SURVEY_NAME,
                    DIMENSIONALITY,
                    COUNT(*)            AS file_count,
                    SUM(FILE_SIZE_KB)/1024.0 AS size_mb,
                    MIN(LINE_NAME)      AS sample_line,
                    MIN(SEIS_FILE_ID)   AS sample_id
                FROM [las_catalog].[SEIS_FILE_CATALOG]
                WHERE FILE_FORMAT = 'SEGY'
                  AND SEIS_SET_ID IS NULL
                GROUP BY SURVEY_NAME, DIMENSIONALITY
                ORDER BY SURVEY_NAME
            """)).fetchall()
        if not rows:
            st.success("✅ All catalogued SEG-Y files are already linked to PPDM SEIS_SET records.")
            return
        unlinked = pd.DataFrame(rows, columns=[
            "SURVEY_NAME", "DIMENSIONALITY", "file_count", "size_mb",
            "sample_line", "sample_id"
        ])
    except Exception as e:
        st.error(f"Could not load unlinked files: {e}")
        return

    st.info(
        f"**{len(unlinked)} survey group(s)** with "
        f"**{unlinked['file_count'].sum():,} file(s)** not yet linked to PPDM."
    )

    # ── Load existing SEIS_SET records for fuzzy matching ─────────────────────
    existing_sets = []
    try:
        with engine.connect() as con:
            ex_rows = con.execute(text(
                "SELECT SEIS_SET_ID, SEIS_SET_NAME, SEIS_TYPE FROM dbo.SEIS_SET"
            )).fetchall()
        existing_sets = [{"id": r[0], "name": r[1] or "", "type": r[2] or ""}
                         for r in ex_rows]
    except Exception:
        pass

    # ── Simple fuzzy match helper ─────────────────────────────────────────────
    def _fuzzy_match(survey_name: str) -> dict | None:
        if not existing_sets or not survey_name:
            return None
        try:
            from difflib import SequenceMatcher
            best, best_ratio = None, 0.0
            for rec in existing_sets:
                ratio = SequenceMatcher(
                    None, survey_name.upper(), rec["name"].upper()
                ).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best = rec
            return {"match": best, "score": best_ratio} if best_ratio > 0.6 else None
        except Exception:
            return None

    # ── Review table ──────────────────────────────────────────────────────────
    st.divider()
    st.markdown("**Survey groups awaiting PPDM linkage**")

    display = unlinked.copy()
    display["size_mb"] = display["size_mb"].apply(
        lambda x: f"{float(x or 0):.1f} MB" if x else "—"
    )
    display["Suggested SEIS_SET_ID"] = ""
    display["Confidence"] = ""

    for i, row in unlinked.iterrows():
        match = _fuzzy_match(str(row["SURVEY_NAME"] or ""))
        if match:
            display.at[i, "Suggested SEIS_SET_ID"] = match["match"]["id"]
            display.at[i, "Confidence"] = f"{match['score']*100:.0f}%"
        else:
            display.at[i, "Suggested SEIS_SET_ID"] = "(create new)"
            display.at[i, "Confidence"] = "—"

    st.dataframe(
        display[["SURVEY_NAME","DIMENSIONALITY","file_count","size_mb",
                 "Suggested SEIS_SET_ID","Confidence"]].rename(columns={
            "SURVEY_NAME": "Survey Name",
            "DIMENSIONALITY": "Type",
            "file_count": "Files",
            "size_mb": "Size",
        }),
        hide_index=True, use_container_width=True
    )

    # ── Action buttons ────────────────────────────────────────────────────────
    st.divider()
    col_all, col_sel, col_source = st.columns([1, 1, 2])

    with col_source:
        source = st.text_input("Source", value="DATA_WRANGLER",
                               key="seis_map_source")

    with col_all:
        seed_all = st.button(
            f"🌱 Create PPDM records for all {len(unlinked)} survey(s)",
            type="primary", key="seis_map_seed_all", use_container_width=True
        )

    # ── Execute seeding ───────────────────────────────────────────────────────
    if seed_all:
        now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")
        seeded, linked, errors = 0, 0, []
        prog = st.progress(0, text="Creating PPDM records…")

        for i, row in unlinked.iterrows():
            prog.progress((i + 1) / len(unlinked),
                          text=f"Processing {row['SURVEY_NAME']}…")
            survey_name = str(row["SURVEY_NAME"] or "UNKNOWN")
            dim         = str(row["DIMENSIONALITY"] or "2D")
            seis_type   = "3D" if dim == "3D" else "2D"
            subtype     = "SEIS_3D" if dim == "3D" else "SEIS_LINE"

            # Generate deterministic ID
            import hashlib
            seis_set_id = hashlib.sha1(
                f"SEGY|{survey_name}".encode()
            ).hexdigest()[:40].upper()

            try:
                with engine.begin() as con:
                    # Check / insert SEIS_SET
                    exists = con.execute(text(
                        "SELECT SEIS_SET_ID FROM dbo.SEIS_SET "
                        "WHERE SEIS_SET_ID = :id"
                    ), {"id": seis_set_id}).scalar()

                    if not exists:
                        con.execute(text("""
                            INSERT INTO dbo.SEIS_SET
                                (SEIS_SET_ID, SEIS_SET_SUBTYPE, SEIS_SET_NAME,
                                 SEIS_TYPE, ACTIVE_IND, SOURCE,
                                 ROW_CREATED_BY, ROW_CREATED_DATE,
                                 ROW_CHANGED_BY, ROW_CHANGED_DATE,
                                 ROW_EFFECTIVE_DATE, ROW_EXPIRY_DATE, PPDM_GUID)
                            VALUES
                                (:id, :subtype, :name,
                                 :stype, 'Y', :src,
                                 'DATA_WRANGLER', :now,
                                 'DATA_WRANGLER', :now,
                                 CAST('1900-01-01' AS DATETIME2),
                                 CAST('2099-12-31' AS DATETIME2),
                                 NEWID())
                        """), {
                            "id":      seis_set_id,
                            "subtype": subtype,
                            "name":    survey_name[:255],
                            "stype":   seis_type,
                            "src":     source,
                            "now":     now,
                        })
                        seeded += 1

                    # Update SEIS_FILE_CATALOG — link all files in this survey
                    con.execute(text("""
                        UPDATE [las_catalog].[SEIS_FILE_CATALOG]
                        SET SEIS_SET_ID = :sid
                        WHERE FILE_FORMAT = 'SEGY'
                          AND SURVEY_NAME  = :survey
                          AND SEIS_SET_ID IS NULL
                    """), {"sid": seis_set_id, "survey": survey_name})
                    linked += 1

            except Exception as e:
                errors.append(f"{survey_name}: {e}")

        prog.empty()

        if seeded:
            st.success(f"✅ Created {seeded} new SEIS_SET record(s).")
        if linked:
            st.success(f"🔗 Linked {linked} survey group(s) in SEIS_FILE_CATALOG.")
        if errors:
            with st.expander(f"❌ {len(errors)} error(s)"):
                for e in errors:
                    st.text(e)

        st.rerun()


def _render_segy_catalog(engine, repo_opts):
    st.markdown("**Catalog SEG-Y files**")
    sub_single, sub_dir = st.tabs(["Single file", "Directory"])
    with sub_single:
        _render_segy_single(engine, repo_opts)
    with sub_dir:
        _render_segy_directory(engine, repo_opts)


def _render_segy_single(engine, repo_opts):
            file_path = st.text_input(
                "Full file path", key="segy_single_path",
                placeholder=r"e.g. C:\Seismic\line01.segy"
            )
            col_repo, col_ppdm = st.columns([3, 1])
            with col_repo:
                repo_label = st.selectbox("Repository", options=list(repo_opts.keys()),
                                           key="segy_single_repo")
            with col_ppdm:
                seed_ppdm = st.checkbox("Seed PPDM", key="segy_single_ppdm",
                                         help="Insert dbo.SEIS_SET and dbo.SEIS_LINE")

            if not file_path:
                return

            valid, msg = _validate_file_path(file_path, [".segy", ".sgy", ".SEGY", ".SGY"])
            if not valid:
                st.error(msg)
                return
            st.caption(msg)

            if st.button("🔍 Preview header", key="segy_preview_btn"):
                with st.spinner("Reading SEG-Y header…"):
                    try:
                        hdr = parse_segy_header(file_path)
                        st.session_state["segy_header"] = hdr
                        st.session_state["segy_header_path"] = file_path
                    except Exception as e:
                        st.error(str(e))

            if "segy_header" not in st.session_state:
                return
            if st.session_state.get("segy_header_path") != file_path:
                st.session_state.pop("segy_header", None)
                return

            hdr = st.session_state["segy_header"]
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Dimensionality", hdr["dimensionality"])
            c2.metric("Traces",         f"{hdr['trace_count']:,}" if hdr["trace_count"] else "—")
            c3.metric("Sample interval",f"{hdr['sample_interval_us']} µs" if hdr["sample_interval_us"] else "—")
            c4.metric("SEG-Y revision", hdr["segy_revision"])

            col_info, col_bbox = st.columns(2)
            with col_info:
                st.markdown("**File metadata**")
                info = {
                    "Survey":   hdr["survey_name"] or "—",
                    "Line":     hdr["line_name"]   or "—",
                    "Client":   hdr["client_name"] or "—",
                    "Vessel":   hdr["vessel_name"] or "—",
                    "Format":   hdr["data_format"] or "—",
                    "Date":     hdr["acq_date_start"] or "—",
                    "CRS":      hdr["coord_system"][:50] + "…" if len(hdr["coord_system"]) > 50
                                else hdr["coord_system"] or "—",
                }
                for k, v in info.items():
                    st.text(f"{k}: {v}")

            with col_bbox:
                st.markdown("**Bounding box**")
                if hdr["min_lat"] is not None:
                    st.text(f"Lat: {hdr['min_lat']:.5f} – {hdr['max_lat']:.5f}")
                    st.text(f"Lon: {hdr['min_lon']:.5f} – {hdr['max_lon']:.5f}")
                if hdr["min_x"] is not None:
                    st.text(f"X: {hdr['min_x']:,.1f} – {hdr['max_x']:,.1f}")
                    st.text(f"Y: {hdr['min_y']:,.1f} – {hdr['max_y']:,.1f}")
                if hdr["min_inline"] is not None:
                    st.text(f"Inline: {hdr['min_inline']} – {hdr['max_inline']}")
                    st.text(f"Xline:  {hdr['min_crossline']} – {hdr['max_crossline']}")

            with st.expander("EBCDIC text header"):
                st.code("\n".join(hdr["text_header"]), language=None)

            _render_segy_section_viewer(file_path, hdr)

            st.divider()
            st.markdown("**Survey map**")
            _has_coords = (
                hdr.get("min_lat") is not None or
                hdr.get("min_x")   is not None
            )
            if _has_coords:
                try:
                    _render_segy_map(hdr)
                except Exception as _map_err:
                    st.warning(f"Map could not be rendered: {_map_err}")

                # Coordinate summary table
                _coord_rows = []
                if hdr.get("min_lat") is not None:
                    _coord_rows += [
                        {"Field": "Min Latitude",  "Value": f"{hdr['min_lat']:.6f} °N"},
                        {"Field": "Max Latitude",  "Value": f"{hdr['max_lat']:.6f} °N"},
                        {"Field": "Min Longitude", "Value": f"{hdr['min_lon']:.6f} °E"},
                        {"Field": "Max Longitude", "Value": f"{hdr['max_lon']:.6f} °E"},
                    ]
                if hdr.get("min_x") is not None:
                    _coord_rows += [
                        {"Field": "Min Easting",   "Value": f"{hdr['min_x']:,.2f}"},
                        {"Field": "Max Easting",   "Value": f"{hdr['max_x']:,.2f}"},
                        {"Field": "Min Northing",  "Value": f"{hdr['min_y']:,.2f}"},
                        {"Field": "Max Northing",  "Value": f"{hdr['max_y']:,.2f}"},
                    ]
                if hdr.get("min_inline") is not None:
                    _coord_rows += [
                        {"Field": "Min Inline",    "Value": str(hdr['min_inline'])},
                        {"Field": "Max Inline",    "Value": str(hdr['max_inline'])},
                        {"Field": "Min Crossline", "Value": str(hdr['min_crossline'])},
                        {"Field": "Max Crossline", "Value": str(hdr['max_crossline'])},
                    ]
                if _coord_rows:
                    with st.expander("📐 Coordinate details", expanded=False):
                        st.dataframe(
                            pd.DataFrame(_coord_rows),
                            hide_index=True,
                            use_container_width=True
                        )
            else:
                scalar = hdr.get("coord_scalar_raw")
                samples = hdr.get("coord_samples", 0)
                diag = []
                if scalar == 0:
                    diag.append("coordinate scalar is 0 (no scaling applied)")
                elif scalar is not None:
                    diag.append(f"coordinate scalar = {scalar}")
                if samples == 0:
                    diag.append("no traces could be sampled")
                else:
                    diag.append(f"{samples} trace(s) sampled")
                st.info(
                    "No valid coordinates found in trace headers. "
                    + (("Diagnostics: " + "; ".join(diag) + ".") if diag else "")
                    + "\n\nPossible causes: zero-filled XY fields, "
                    "coordinates stored in non-standard trace header bytes, "
                    "or the file uses a proprietary byte location for coordinates."
                )

            st.divider()
            if st.button("📥 Catalog SEG-Y file", type="primary", key="segy_catalog_btn"):
                with st.spinner("Cataloguing…"):
                    r = catalog_segy_file(
                        engine, file_path,
                        repository_id=repo_opts[repo_label],
                        seed_ppdm=seed_ppdm,
                    )
                if r["ok"]:
                    st.success(f"✅ {r['action'].capitalize()} | ID: {r['seis_file_id']}" +
                               (" | PPDM seeded" if r["seeded_ppdm"] else ""))
                    if r.get("error"):
                        st.warning(r["error"])
                else:
                    st.error(r["error"])


def _render_segy_directory(engine, repo_opts):
            folder = st.text_input("Directory path", key="segy_dir_path",
                                    placeholder=r"e.g. C:\Seismic\2D_Lines")
            col_repo2, col_ppdm2, col_wk = st.columns([3, 1, 1])
            with col_repo2:
                repo_label2 = st.selectbox("Repository", options=list(repo_opts.keys()),
                                            key="segy_dir_repo")
            with col_ppdm2:
                seed_ppdm2 = st.checkbox("Seed PPDM", key="segy_dir_ppdm")
            with col_wk:
                workers = st.select_slider("Threads",
                    options=[1, 2, 4, 6, 8, 10, 12, 16],
                    value=__import__("os").cpu_count() and
                          min(max(__import__("os").cpu_count()-2, 2), 12) or 4,
                    key="segy_dir_workers")

            if folder:
                try:
                    p = Path(folder)
                    if not p.exists():
                        st.error(f"Directory not found: `{folder}`")
                        files = []
                    elif not p.is_dir():
                        st.error(f"Not a directory: `{folder}`")
                        files = []
                    else:
                        _segy_exts = {".segy", ".sgy", ".seg"}
                        all_files = list(p.rglob("*"))
                        all_files = [f for f in all_files if f.is_file()]
                        files = [f for f in all_files
                                 if f.suffix.lower() in _segy_exts]
                        n_dirs = len(set(f.parent for f in files))
                        st.info(
                            f"{len(files)} SEG-Y file(s) found"
                            + (f" across {n_dirs} sub-folder(s)" if n_dirs > 1 else "")
                            + f" (searched recursively)"
                        )
                        if len(all_files) == 0:
                            st.caption("Directory appears empty.")
                        elif len(files) == 0:
                            exts = sorted(set(f.suffix for f in all_files))
                            st.caption(f"Extensions found: {', '.join(exts) or 'none'}")
                            names = [f.name for f in all_files[:5]]
                            st.caption(f"Files: {', '.join(names)}"
                                       + (" …" if len(all_files) > 5 else ""))
                except Exception as e:
                    st.error(str(e))
                    files = []

                segy_dir_survey = ""  # auto-detect from each file's header

                if st.button(f"🚀 Catalog {len(files)} SEG-Y file(s)", type="primary",
                              key="segy_dir_btn", disabled=len(files) == 0):
                    prog = st.progress(0, text="Starting…")
                    def _cb(done, total, name):
                        prog.progress(min(done/total, 1.0), text=f"{name} ({done}/{total})…")
                    try:
                        # Store survey name override in session state for
                        # catalog_segy_file to pick up via the pre-parse hook
                        if segy_dir_survey.strip():
                            st.session_state["segy_survey_name_override"] = segy_dir_survey.strip()
                        else:
                            st.session_state.pop("segy_survey_name_override", None)
                        results = catalog_segy_directory(
                            engine, folder,
                            repository_id=repo_opts[repo_label2],
                            seed_ppdm=seed_ppdm2,
                            max_workers=workers,
                            progress_callback=_cb,
                        )
                        prog.empty()
                        ins = sum(1 for r in results if r.get("action") == "inserted")
                        upd = sum(1 for r in results if r.get("action") == "updated")
                        err = sum(1 for r in results if not r.get("ok"))
                        st.success(f"✅ {ins} new · {upd} updated · {err} error(s)")
                        if err:
                            with st.expander(f"❌ {err} error(s)", expanded=True):
                                st.dataframe(pd.DataFrame([
                                    {"File": r["file_name"], "Error": r["error"]}
                                    for r in results if not r.get("ok")
                                ]), hide_index=True, use_container_width=True)
                    except Exception as e:
                        st.error(str(e))



def _render_p190_catalog(engine, repo_opts):
    st.markdown("**Catalog P1/90 positioning files**")
    sub_single, sub_dir = st.tabs(["Single file", "Directory"])
    with sub_single:
        _render_p190_single(engine, repo_opts)
    with sub_dir:
        _render_p190_directory(engine, repo_opts)


def _render_p190_single(engine, repo_opts):
            file_path = st.text_input(
                "Full file path", key="p190_single_path",
                placeholder=r"e.g. C:\Nav\line01.p190"
            )
            col_repo, col_ppdm = st.columns([3, 1])
            with col_repo:
                repo_label = st.selectbox("Repository", options=list(repo_opts.keys()),
                                           key="p190_single_repo")
            with col_ppdm:
                seed_ppdm = st.checkbox("Seed PPDM", key="p190_single_ppdm",
                                         help="Insert dbo.SEIS_SET and dbo.SEIS_LINE")

            if file_path:
                valid, msg = _validate_file_path(file_path, [".p190", ".P190", ".p90", ".P90"])
                if not valid:
                    st.error(msg)
                else:
                    st.caption(msg)
                    if st.button("🔍 Preview header", key="p190_preview_btn"):
                        with st.spinner("Reading P190 file…"):
                            try:
                                hdr = parse_p190_header(file_path)
                                st.session_state["p190_header"] = hdr
                                st.session_state["p190_header_path"] = file_path
                            except Exception as e:
                                st.error(str(e))

            _p190_ready = (
                file_path and
                "p190_header" in st.session_state and
                st.session_state.get("p190_header_path") == file_path
            )
            if file_path and st.session_state.get("p190_header_path") != file_path:
                st.session_state.pop("p190_header", None)

            if not _p190_ready:
                return

            hdr = st.session_state["p190_header"]
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total records", f"{hdr['record_count']:,}")
            c2.metric("Shot records",  f"{hdr['shot_count']:,}")
            c3.metric("First SP",      hdr["first_shot_point"] or "—")
            c4.metric("Last SP",       hdr["last_shot_point"] or "—")

            col_info, col_bbox = st.columns(2)
            with col_info:
                st.markdown("**File metadata**")
                info = {
                    "Survey":  hdr["survey_name"] or "—",
                    "Line":    hdr["line_name"]   or "—",
                    "Vessel":  hdr["vessel_name"] or "—",
                    "Client":  hdr["client_name"] or "—",
                    "Nav sys": hdr["nav_system"]  or "—",
                    "Date start": hdr["acq_date_start"] or "—",
                    "Date end":   hdr["acq_date_end"]   or "—",
                }
                for k, v in info.items():
                    st.text(f"{k}: {v}")
                st.markdown("**Record type counts**")
                counts_df = pd.DataFrame([
                    {"Type": k, "Count": v}
                    for k, v in sorted(hdr["record_type_counts"].items())
                ])
                st.dataframe(counts_df, hide_index=True, use_container_width=True)

            with col_bbox:
                st.markdown("**Bounding box**")
                if hdr["min_lat"] is not None:
                    st.text(f"Lat: {hdr['min_lat']:.5f} – {hdr['max_lat']:.5f}")
                    st.text(f"Lon: {hdr['min_lon']:.5f} – {hdr['max_lon']:.5f}")
                if hdr["min_x"] is not None:
                    st.text(f"X: {hdr['min_x']:,.1f} – {hdr['max_x']:,.1f}")
                    st.text(f"Y: {hdr['min_y']:,.1f} – {hdr['max_y']:,.1f}")
                if not hdr["min_lat"] and not hdr["min_x"]:
                    st.info("No coordinates extracted from shot records.")

            # Show raw H records from the file
            if file_path and st.checkbox("Show raw header records (H lines)", key="p190_show_raw"):
                try:
                    import os as _os
                    if _os.path.exists(file_path):
                        _h_lines = []
                        with open(file_path, "r", encoding="latin-1", errors="replace") as _f:
                            for _ln in _f:
                                if _ln.startswith("H"):
                                    _h_lines.append(_ln.rstrip())
                                elif _h_lines:
                                    break  # stop at first non-H record
                        if _h_lines:
                            with st.expander(f"📄 P190 Header Records ({len(_h_lines)} lines)", expanded=True):
                                st.code("\n".join(_h_lines), language=None)
                        else:
                            st.info("No H records found.")
                    else:
                        st.warning("File not accessible from this machine.")
                except Exception as _e:
                    st.warning(f"Could not read file: {_e}")

            st.divider()
            st.markdown("**Track map**")
            try:
                _render_p190_map(hdr)
            except Exception as _map_err:
                st.warning(f"Map could not be rendered: {_map_err}")

            st.divider()
            p190_survey_override = st.text_input(
                "Survey name override",
                value=hdr.get("survey_name", "") or "",
                key="p190_survey_override",
                help="Groups this line with others from the same survey on the map."
            )

            if st.button("📥 Catalog P190 file", type="primary", key="p190_catalog_btn"):
                with st.spinner("Cataloguing…"):
                    if p190_survey_override.strip():
                        hdr["survey_name"] = p190_survey_override.strip()
                    r = catalog_p190_file(
                        engine, file_path,
                        repository_id=repo_opts[repo_label],
                        seed_ppdm=seed_ppdm,
                    )
                if r["ok"]:
                    st.success(f"✅ {r['action'].capitalize()} | ID: {r['seis_file_id']}" +
                               (" | PPDM seeded" if r["seeded_ppdm"] else ""))
                    if r.get("error"):
                        st.warning(r["error"])
                else:
                    st.error(r["error"])


def _render_p190_directory(engine, repo_opts):
            folder = st.text_input("Directory path", key="p190_dir_path",
                                    placeholder=r"e.g. C:\Nav\P190_Files")
            col_repo2, col_ppdm2, col_wk = st.columns([3, 1, 1])
            with col_repo2:
                repo_label2 = st.selectbox("Repository", options=list(repo_opts.keys()),
                                            key="p190_dir_repo")
            with col_ppdm2:
                seed_ppdm2 = st.checkbox("Seed PPDM", key="p190_dir_ppdm")
            with col_wk:
                workers = st.select_slider("Threads",
                    options=[1, 2, 4, 6, 8, 10, 12, 16],
                    value=__import__("os").cpu_count() and
                          min(max(__import__("os").cpu_count()-2, 2), 12) or 4,
                    key="p190_dir_workers")

            if folder:
                try:
                    _p190_exts = {".p190", ".p90", ".p1"}
                    files = [f for f in Path(folder).rglob("*")
                             if f.is_file() and f.suffix.lower() in _p190_exts]
                    st.info(f"{len(files)} P190 file(s) found.")
                except Exception:
                    files = []

                if st.button(f"🚀 Catalog {len(files)} P190 file(s)", type="primary",
                              key="p190_dir_btn", disabled=len(files) == 0):
                    prog = st.progress(0, text="Starting…")
                    def _cb(done, total, name):
                        prog.progress(min(done/total, 1.0), text=f"{name} ({done}/{total})…")
                    try:
                        results = catalog_p190_directory(
                            engine, folder,
                            repository_id=repo_opts[repo_label2],
                            seed_ppdm=seed_ppdm2,
                            max_workers=workers,
                            progress_callback=_cb,
                        )
                        prog.empty()
                        ins = sum(1 for r in results if r.get("action") == "inserted")
                        upd = sum(1 for r in results if r.get("action") == "updated")
                        err = sum(1 for r in results if not r.get("ok"))
                        st.success(f"✅ {ins} new · {upd} updated · {err} error(s)")
                        if err:
                            with st.expander(f"❌ {err} error(s)", expanded=True):
                                st.dataframe(pd.DataFrame([
                                    {"File": r["file_name"], "Error": r["error"]}
                                    for r in results if not r.get("ok")
                                ]), hide_index=True, use_container_width=True)
                    except Exception as e:
                        st.error(str(e))




# ─────────────────────────────────────────────────────────────────────────────
# SEISMIC SECTION VIEWER
# ─────────────────────────────────────────────────────────────────────────────

def _read_segy_section(file_path: str, max_traces: int = 1000) -> dict | None:
    """
    Read SEG-Y trace data and return a dict with the section array and metadata.
    Tries big-endian IEEE float first, falls back to IBM float conversion.
    Subsamples traces if file has more than max_traces.
    """
    import struct as _struct
    import numpy as _np

    def _ibm_to_ieee(data: bytes) -> "_np.ndarray":
        ibm = _np.frombuffer(data, dtype=">u4")
        sign     = (ibm >> 31) & 1
        exp      = (ibm >> 24) & 0x7F
        mantissa = (ibm & 0x00FFFFFF).astype(_np.float64)
        result   = ((1 - 2*sign) * mantissa / (1 << 24)
                    * _np.float_power(16.0, exp.astype(_np.float64) - 64))
        return result.astype(_np.float32)

    try:
        with open(file_path, "rb") as f:
            # Binary header
            f.seek(3200)
            bin_hdr    = f.read(400)
            n_samples  = _struct.unpack(">h", bin_hdr[20:22])[0]
            si_us      = _struct.unpack(">h", bin_hdr[16:18])[0]
            data_fmt   = _struct.unpack(">h", bin_hdr[24:26])[0]

            if n_samples <= 0 or si_us <= 0:
                return None

            bps = {1: 4, 2: 4, 3: 2, 5: 4}.get(data_fmt, 4)
            trace_bytes = 240 + n_samples * bps

            # Count traces from file size
            file_size  = __import__("os").path.getsize(file_path)
            n_traces   = max(1, (file_size - 3600) // trace_bytes)

            # Subsample if needed
            step = max(1, n_traces // max_traces)
            indices = list(range(0, n_traces, step))

            traces = []
            cdps   = []
            for idx in indices:
                offset = 3600 + idx * trace_bytes
                f.seek(offset)
                hdr = f.read(240)
                if len(hdr) < 240:
                    break
                cdp = _struct.unpack(">i", hdr[20:24])[0]
                cdps.append(cdp if cdp != 0 else idx + 1)
                raw = f.read(n_samples * bps)
                if len(raw) < n_samples * bps:
                    break

                if data_fmt == 1:
                    # Try IEEE first — if values look reasonable use it
                    arr_ieee = _np.frombuffer(raw, dtype=">f4").copy()
                    if _np.abs(arr_ieee).max() < 1e6:
                        traces.append(arr_ieee)
                    else:
                        traces.append(_ibm_to_ieee(raw))
                elif data_fmt == 2:
                    traces.append(_np.frombuffer(raw, dtype=">i4").astype(_np.float32).copy())
                elif data_fmt == 3:
                    traces.append(_np.frombuffer(raw, dtype=">i2").astype(_np.float32).copy())
                else:
                    traces.append(_np.frombuffer(raw, dtype=">f4").copy())

        if not traces:
            return None

        section = _np.array(traces).T  # (n_samples, n_traces_read)

        # Find first live sample (skip leading zeros)
        live_mask  = section.any(axis=1)
        live_start = int(_np.argmax(live_mask)) if live_mask.any() else 0
        live_end   = int(len(live_mask) - _np.argmax(live_mask[::-1])) if live_mask.any() else n_samples

        times = _np.arange(n_samples) * si_us / 1000.0  # ms

        return {
            "section":    section,
            "section_live": section[live_start:live_end, :],
            "times":      times,
            "times_live": times[live_start:live_end],
            "cdps":       cdps,
            "n_samples":  n_samples,
            "n_traces":   n_traces,
            "n_read":     len(traces),
            "si_ms":      si_us / 1000.0,
            "live_start": live_start,
            "data_fmt":   data_fmt,
            "step":       step,
        }
    except Exception as e:
        return {"error": str(e)}


def _render_segy_section_viewer(file_path: str, hdr: dict):
    """
    Render an interactive seismic section viewer for a SEG-Y file.
    Called from the single file preview tab.
    """
    st.divider()
    st.markdown("**📈 Seismic section viewer**")

    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.colors as mcolors
        import numpy as np
    except ImportError:
        st.warning("matplotlib is required. Run: `pip install matplotlib`")
        return

    # ── Controls ──────────────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        display_mode = st.selectbox(
            "Display", ["Variable Density", "Wiggle", "VD + Wiggle"],
            key="seis_view_mode"
        )
    with c2:
        colourmap = st.selectbox(
            "Colour map",
            ["RdBu_r", "Greys", "seismic", "bwr", "RdGy_r"],
            key="seis_view_cmap"
        )
    with c3:
        clip_pct = st.slider(
            "Clip %", 90, 100, 98, 1,
            key="seis_view_clip",
            help="Percentile clip for amplitude scaling"
        )
    with c4:
        max_traces = st.select_slider(
            "Resolution",
            options=[100, 250, 500, 1000],
            value=500,
            key="seis_view_res",
            help="Number of traces to read (higher = slower but more detail)"
        )

    # Determine live data range from cached section if available
    _cached = st.session_state.get("seis_section") if st.session_state.get("seis_section_path") == file_path else None
    _max_t  = round(int(hdr.get("sample_count") or 3000) * int(hdr.get("sample_interval_us") or 4000) / 1000.0)
    _si_ms  = int(hdr.get("sample_interval_us") or 4000) / 1000.0
    if _cached and "live_start" in _cached:
        _default_t_min = round(_cached["live_start"] * _si_ms / 1) * 1.0
    else:
        _default_t_min = 0.0

    # Time range
    t1, t2 = st.columns(2)
    with t1:
        t_min = st.number_input(
            "Start time (ms)", value=_default_t_min,
            min_value=0.0, max_value=float(_max_t),
            step=100.0, key="seis_t_min"
        )
    with t2:
        _default_t_max = min(_default_t_min + 4000.0, float(_max_t)) if _default_t_min > 0 else float(_max_t)
        t_max = st.number_input(
            "End time (ms)", value=_default_t_max,
            min_value=0.0, max_value=float(_max_t),
            step=100.0, key="seis_t_max"
        )

    # CDP / trace range
    _cached_cdps = _cached.get("cdps", []) if _cached else []
    _cdp_min = int(min(_cached_cdps)) if _cached_cdps else 1
    _cdp_max = int(max(_cached_cdps)) if _cached_cdps else 99999
    c1, c2 = st.columns(2)
    with c1:
        cdp_min = st.number_input(
            "CDP from", value=_cdp_min,
            step=50, key="seis_cdp_min",
            help="Filter to a CDP range — leave at min/max to show full line"
        )
    with c2:
        cdp_max = st.number_input(
            "CDP to", value=_cdp_max,
            step=50, key="seis_cdp_max"
        )

    # For 3D files hand off to the 3D viewer
    if hdr.get("dimensionality") == "3D" or (
        hdr.get("min_inline") is not None and hdr.get("min_inline", 0) != 0
    ):
        _render_segy_3d_viewer(file_path, hdr)
        return

    if st.button("📈 Render section", type="primary", key="seis_render_btn"):
        with st.spinner(f"Reading {max_traces} traces…"):
            data = _read_segy_section(file_path, max_traces=max_traces)
            st.session_state["seis_section"] = data
            st.session_state["seis_section_path"] = file_path

    # Use cached section if available for same file
    if (st.session_state.get("seis_section_path") != file_path):
        st.session_state.pop("seis_section", None)

    data = st.session_state.get("seis_section")
    if not data:
        st.caption("Click **Render section** to load the trace data.")
        return

    if "error" in data:
        st.error(f"Could not read trace data: {data['error']}")
        return

    section = data["section"]
    times   = data["times"]
    cdps    = data["cdps"]

    # Time window filter
    t_mask   = (times >= t_min) & (times <= t_max)
    # Apply time filter
    section_w = section[t_mask, :]
    times_w   = times[t_mask]

    # Apply CDP range filter
    cdp_arr = np.array(data["cdps"])
    cdp_mask = (cdp_arr >= cdp_min) & (cdp_arr <= cdp_max)
    if cdp_mask.any():
        section_w = section_w[:, cdp_mask]
        cdp_arr   = cdp_arr[cdp_mask]
    else:
        st.warning(f"No traces in CDP range {cdp_min}–{cdp_max}.")
        return

    if section_w.size == 0:
        st.warning("No data in selected time range.")
        return

    vmax = float(np.percentile(np.abs(section_w), clip_pct))
    if vmax == 0:
        # Find where live data actually starts
        live_mask  = section[t_mask, :] if True else section
        any_live   = np.any(section != 0, axis=1)
        live_start_sample = int(np.argmax(any_live)) if any_live.any() else 0
        live_start_ms = live_start_sample * int(hdr.get("sample_interval_us") or 4000) / 1000.0
        st.warning(
            f"All amplitudes are zero in the selected time window "
            f"({t_min:.0f}–{t_max:.0f} ms).\n\n"
            f"Live data starts at approximately **{live_start_ms:.0f} ms**. "
            f"Adjust the Start time above."
        )
        return

    # ── Render ────────────────────────────────────────────────────────────
    fig_h = max(5, min(10, len(times_w) / 200))
    fig, ax = plt.subplots(figsize=(14, fig_h))
    ax.set_facecolor("#f0f0f0")

    extent  = [cdp_arr[0], cdp_arr[-1], times_w[-1], times_w[0]]

    if "Variable Density" in display_mode or "VD" in display_mode:
        ax.imshow(
            section_w, aspect="auto",
            cmap=colourmap, vmin=-vmax, vmax=vmax,
            extent=extent, interpolation="bilinear"
        )

    if "Wiggle" in display_mode:
        # Plot every Nth trace as a wiggle
        n_wiggles = min(100, section_w.shape[1])
        wig_step  = max(1, section_w.shape[1] // n_wiggles)
        wig_scale = (cdp_arr[-1] - cdp_arr[0]) / max(section_w.shape[1], 1) * wig_step * 0.6
        for wi in range(0, section_w.shape[1], wig_step):
            cdp_x = float(cdp_arr[wi]) if wi < len(cdp_arr) else float(cdp_arr[-1])
            trace = section_w[:, wi] / (vmax + 1e-10) * wig_scale
            ax.plot(cdp_x + trace, times_w, "k-", linewidth=0.3, alpha=0.6)
            # Fill positive
            ax.fill_betweenx(times_w, cdp_x, cdp_x + trace,
                             where=(trace > 0), color="black", alpha=0.4)

    ax.set_xlabel("CDP", fontsize=10)
    ax.set_ylabel("Time (ms)", fontsize=10)
    survey = hdr.get("survey_name", "") or ""
    line   = hdr.get("line_name", "") or ""
    title  = f"{survey}  {line}".strip() or __import__("pathlib").Path(file_path).stem
    ax.set_title(
        f"{title}  |  CDP {cdp_arr[0]}–{cdp_arr[-1]}  |  "
        f"{times_w[0]:.0f}–{times_w[-1]:.0f} ms",
        fontsize=10, pad=8
    )
    ax.grid(True, linestyle="--", alpha=0.2, linewidth=0.5)

    # Colourbar for VD mode
    if "Variable Density" in display_mode or "VD" in display_mode:
        sm = plt.cm.ScalarMappable(
            cmap=colourmap,
            norm=mcolors.Normalize(vmin=-vmax, vmax=vmax)
        )
        sm.set_array([])
        plt.colorbar(sm, ax=ax, shrink=0.6, pad=0.01, label="Amplitude")

    plt.tight_layout()
    st.pyplot(fig, use_container_width=True)
    plt.close(fig)

    st.caption(
        f"File: `{__import__('pathlib').Path(file_path).name}`  |  "
        f"{data['n_traces']:,} total traces  |  "
        f"Step: every {data['step']} trace(s)  |  "
        f"Sample interval: {data['si_ms']:.1f} ms  |  "
        f"Data format: {data['data_fmt']}"
    )


# ─────────────────────────────────────────────────────────────────────────────
# 3D SEG-Y VIEWER
# ─────────────────────────────────────────────────────────────────────────────

def _build_segy_3d_index(file_path: str, progress_cb=None) -> dict | None:
    """
    Build a trace index for a 3D SEG-Y file.
    Reads only the 240-byte trace headers — skips all data samples.
    Returns dict with inline/crossline arrays and file offsets.
    """
    import struct as _struct
    import numpy as _np
    import os as _os

    try:
        fsize = _os.path.getsize(file_path)
        with open(file_path, "rb") as f:
            f.seek(3200)
            bin_hdr   = f.read(400)
            n_samples = _struct.unpack(">h", bin_hdr[20:22])[0]
            si_us     = _struct.unpack(">h", bin_hdr[16:18])[0]
            data_fmt  = _struct.unpack(">h", bin_hdr[24:26])[0]

        bps         = {1: 4, 2: 4, 3: 2, 5: 4}.get(data_fmt, 4)
        trace_bytes = 240 + n_samples * bps
        n_traces    = max(1, (fsize - 3600) // trace_bytes)

        inlines  = _np.zeros(n_traces, dtype=_np.int32)
        xlines   = _np.zeros(n_traces, dtype=_np.int32)
        offsets  = _np.zeros(n_traces, dtype=_np.int64)

        with open(file_path, "rb") as f:
            for i in range(n_traces):
                off = 3600 + i * trace_bytes
                f.seek(off)
                hdr = f.read(240)
                if len(hdr) < 240:
                    n_traces = i
                    break
                inlines[i] = _struct.unpack(">i", hdr[188:192])[0]
                xlines[i]  = _struct.unpack(">i", hdr[192:196])[0]
                offsets[i] = off
                if progress_cb and i % 5000 == 0:
                    progress_cb(i, n_traces)

        inlines  = inlines[:n_traces]
        xlines   = xlines[:n_traces]
        offsets  = offsets[:n_traces]

        return {
            "inlines":    inlines,
            "xlines":     xlines,
            "offsets":    offsets,
            "n_traces":   n_traces,
            "n_samples":  n_samples,
            "si_us":      si_us,
            "data_fmt":   data_fmt,
            "bps":        bps,
            "trace_bytes": trace_bytes,
            "il_min": int(inlines.min()), "il_max": int(inlines.max()),
            "xl_min": int(xlines.min()),  "xl_max": int(xlines.max()),
        }
    except Exception as e:
        return {"error": str(e)}


def _read_3d_inline(file_path: str, index: dict, il: int) -> "_np.ndarray | None":
    """Read all traces for a given inline number. Returns (n_samples, n_xlines)."""
    import numpy as _np
    mask    = index["inlines"] == il
    offsets = index["offsets"][mask]
    xlines  = index["xlines"][mask]
    n_samp  = index["n_samples"]
    bps     = index["bps"]
    fmt     = index["data_fmt"]

    if len(offsets) == 0:
        return None, None

    # Sort by crossline
    sort_idx = _np.argsort(xlines)
    offsets  = offsets[sort_idx]
    xlines   = xlines[sort_idx]

    traces = []
    with open(file_path, "rb") as f:
        for off in offsets:
            f.seek(off + 240)
            raw = f.read(n_samp * bps)
            if fmt in (1, 5):
                arr = _np.frombuffer(raw, dtype=">f4").copy()
                if _np.abs(arr).max() > 1e6:
                    arr = _ibm_to_ieee_3d(raw)
            elif fmt == 2:
                arr = _np.frombuffer(raw, dtype=">i4").astype(_np.float32).copy()
            elif fmt == 3:
                arr = _np.frombuffer(raw, dtype=">i2").astype(_np.float32).copy()
            else:
                arr = _np.frombuffer(raw, dtype=">f4").copy()
            traces.append(arr)

    return _np.array(traces).T, xlines  # (n_samples, n_xlines), xline_nums


def _read_3d_crossline(file_path: str, index: dict, xl: int) -> "_np.ndarray | None":
    """Read all traces for a given crossline. Returns (n_samples, n_inlines)."""
    import numpy as _np
    mask    = index["xlines"] == xl
    offsets = index["offsets"][mask]
    ils     = index["inlines"][mask]

    sort_idx = _np.argsort(ils)
    offsets  = offsets[sort_idx]
    ils      = ils[sort_idx]

    n_samp = index["n_samples"]
    bps    = index["bps"]
    fmt    = index["data_fmt"]

    traces = []
    with open(file_path, "rb") as f:
        for off in offsets:
            f.seek(off + 240)
            raw = f.read(n_samp * bps)
            arr = _np.frombuffer(raw, dtype=">f4").copy()
            traces.append(arr)

    return _np.array(traces).T, ils


def _read_3d_timeslice(file_path: str, index: dict, sample_idx: int) -> "_np.ndarray | None":
    """Read one time sample from all traces. Returns 2D grid (il, xl)."""
    import numpy as _np
    bps   = index["bps"]
    ils   = index["inlines"]
    xls   = index["xlines"]
    offs  = index["offsets"]

    il_vals = _np.unique(ils)
    xl_vals = _np.unique(xls)
    il_map  = {v: i for i, v in enumerate(il_vals)}
    xl_map  = {v: i for i, v in enumerate(xl_vals)}
    grid    = _np.zeros((len(il_vals), len(xl_vals)), dtype=_np.float32)

    sample_offset = sample_idx * bps

    with open(file_path, "rb") as f:
        for i in range(len(offs)):
            f.seek(offs[i] + 240 + sample_offset)
            raw = f.read(bps)
            if len(raw) == bps:
                val = _np.frombuffer(raw, dtype=">f4")[0]
                grid[il_map[ils[i]], xl_map[xls[i]]] = val

    return grid, il_vals, xl_vals


def _ibm_to_ieee_3d(data: bytes) -> "_np.ndarray":
    import numpy as _np
    ibm      = _np.frombuffer(data, dtype=">u4")
    sign     = (ibm >> 31) & 1
    exp      = (ibm >> 24) & 0x7F
    mantissa = (ibm & 0x00FFFFFF).astype(_np.float64)
    result   = ((1 - 2*sign) * mantissa / (1 << 24)
                * _np.float_power(16.0, exp.astype(_np.float64) - 64))
    return result.astype(_np.float32)


def _render_segy_3d_viewer(file_path: str, hdr: dict):
    """Interactive 3D SEG-Y viewer — inline, crossline and time slice."""
    import numpy as np

    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.colors as mcolors
    except ImportError:
        st.warning("matplotlib is required: `pip install matplotlib`")
        return

    st.divider()
    st.markdown("**🧊 3D Volume viewer**")
    st.caption(
        "Builds a trace index then lets you navigate inlines, crosslines and time slices. "
        "Index building reads only trace headers — no amplitude data loaded until you render."
    )

    idx_key  = f"segy3d_index_{file_path}"
    idx_path = f"segy3d_path_{file_path}"

    # ── Build index ────────────────────────────────────────────────────────
    if idx_key not in st.session_state:
        if st.button("🗂 Build trace index", type="primary", key="segy3d_index_btn"):
            prog = st.progress(0, text="Reading trace headers…")
            def _cb(done, total):
                prog.progress(min(done/total, 1.0),
                              text=f"Indexing trace {done:,} / {total:,}…")
            with st.spinner("Building index — reads headers only, skips data…"):
                idx = _build_segy_3d_index(file_path, progress_cb=_cb)
            prog.empty()
            if "error" in idx:
                st.error(f"Index failed: {idx['error']}")
                return
            st.session_state[idx_key] = idx
            st.success(
                f"✅ Index built — {idx['n_traces']:,} traces  |  "
                f"IL {idx['il_min']}–{idx['il_max']}  |  "
                f"XL {idx['xl_min']}–{idx['xl_max']}"
            )
            st.rerun()
        return

    idx = st.session_state[idx_key]

    st.success(
        f"Index ready — {idx['n_traces']:,} traces  |  "
        f"IL {idx['il_min']}–{idx['il_max']}  |  "
        f"XL {idx['xl_min']}–{idx['xl_max']}  |  "
        f"{idx['n_samples']} samples @ {idx['si_us']/1000:.1f} ms"
    )

    # ── Slice controls ─────────────────────────────────────────────────────
    c1, c2, c3 = st.columns(3)
    with c1:
        colourmap = st.selectbox("Colour map",
            ["RdBu_r","Greys","seismic","bwr"], key="segy3d_cmap")
    with c2:
        clip_pct = st.slider("Clip %", 90, 100, 98, key="segy3d_clip")

    tab_il, tab_xl, tab_ts = st.tabs([
        "📐 Inline", "📏 Crossline", "🗺️ Time slice"
    ])

    def _plot_section(section, x_vals, x_label, times, title, cmap, clip):
        vmax = float(np.percentile(np.abs(section), clip))
        if vmax == 0: vmax = 1.0
        fig, ax = plt.subplots(figsize=(13, 6))
        ax.imshow(section, aspect="auto", cmap=cmap,
                  vmin=-vmax, vmax=vmax,
                  extent=[x_vals[0], x_vals[-1], times[-1], times[0]],
                  interpolation="bilinear")
        ax.set_xlabel(x_label, fontsize=10)
        ax.set_ylabel("Time (ms)", fontsize=10)
        ax.set_title(title, fontsize=10)
        ax.grid(True, linestyle="--", alpha=0.2)
        sm = plt.cm.ScalarMappable(cmap=cmap,
             norm=mcolors.Normalize(vmin=-vmax, vmax=vmax))
        sm.set_array([])
        plt.colorbar(sm, ax=ax, shrink=0.6, pad=0.01, label="Amplitude")
        plt.tight_layout()
        st.pyplot(fig, use_container_width=True)
        plt.close(fig)

    times = np.arange(idx["n_samples"]) * idx["si_us"] / 1000.0
    survey = hdr.get("survey_name","") or ""

    with tab_il:
        il_num = st.slider("Inline", idx["il_min"], idx["il_max"],
                           (idx["il_min"] + idx["il_max"]) // 2,
                           key="segy3d_il_num")
        if st.button("Render inline", type="primary", key="segy3d_il_btn"):
            with st.spinner(f"Reading inline {il_num}…"):
                section, xl_vals = _read_3d_inline(file_path, idx, il_num)
            if section is None:
                st.warning(f"No traces found for inline {il_num}")
            else:
                _plot_section(section, xl_vals, "Crossline",
                              times, f"{survey} — Inline {il_num}",
                              colourmap, clip_pct)

    with tab_xl:
        xl_num = st.slider("Crossline", idx["xl_min"], idx["xl_max"],
                           (idx["xl_min"] + idx["xl_max"]) // 2,
                           key="segy3d_xl_num")
        if st.button("Render crossline", type="primary", key="segy3d_xl_btn"):
            with st.spinner(f"Reading crossline {xl_num}…"):
                section, il_vals = _read_3d_crossline(file_path, idx, xl_num)
            if section is None:
                st.warning(f"No traces found for crossline {xl_num}")
            else:
                _plot_section(section, il_vals, "Inline",
                              times, f"{survey} — Crossline {xl_num}",
                              colourmap, clip_pct)

    with tab_ts:
        max_t_ms = float(times[-1])
        t_ms = st.slider("Time (ms)", 0.0, max_t_ms,
                         max_t_ms * 0.3, step=float(idx["si_us"]/1000),
                         key="segy3d_t_ms")
        sample_idx = int(t_ms / (idx["si_us"] / 1000.0))
        if st.button("Render time slice", type="primary", key="segy3d_ts_btn"):
            with st.spinner(f"Reading time slice at {t_ms:.0f} ms…"):
                grid, il_vals, xl_vals = _read_3d_timeslice(
                    file_path, idx, sample_idx)
            vmax = float(np.percentile(np.abs(grid), clip_pct))
            if vmax == 0: vmax = 1.0
            fig, ax = plt.subplots(figsize=(10, 8))
            im = ax.imshow(grid, aspect="auto", cmap=colourmap,
                           vmin=-vmax, vmax=vmax,
                           extent=[xl_vals[0], xl_vals[-1],
                                   il_vals[-1], il_vals[0]],
                           interpolation="bilinear")
            ax.set_xlabel("Crossline", fontsize=10)
            ax.set_ylabel("Inline", fontsize=10)
            ax.set_title(f"{survey} — Time slice {t_ms:.0f} ms",
                         fontsize=10)
            plt.colorbar(im, ax=ax, shrink=0.6, label="Amplitude")
            plt.tight_layout()
            st.pyplot(fig, use_container_width=True)
            plt.close(fig)

def _render_segy_map(hdr: dict, title: str = ""):
    """
    Render a bounding box map and survey geometry for a SEG-Y file.
    Works for both 2D (line endpoint map) and 3D (survey rectangle).
    Uses matplotlib only — no external mapping libs required.
    """
    import math
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
    except ImportError:
        st.warning(
            "**matplotlib** is not installed. Run: "
            "`pip install matplotlib` then restart Data Wrangler."
        )
        return

    has_geo  = hdr.get("min_lat") is not None
    has_proj = hdr.get("min_x")   is not None
    dim      = hdr.get("dimensionality", "2D")

    if not has_geo and not has_proj:
        st.info("No coordinate data available in this file header.")
        return

    n_panels = 2 if (has_geo and has_proj) else 1
    fig, axes = plt.subplots(1, n_panels, figsize=(12 if n_panels == 2 else 6, 5))
    # Normalise to always be a flat list of Axes objects
    if n_panels == 1:
        ax_list = [axes]
    else:
        ax_list = list(axes)

    def _draw_bbox(ax, x0, x1, y0, y1, xlabel, ylabel, color, dim, hdr):
        w = x1 - x0
        h = y1 - y0
        pad_x = max(w * 0.15, abs(x0) * 0.001 + 0.001)
        pad_y = max(h * 0.15, abs(y0) * 0.001 + 0.001)

        if dim == "3D":
            rect = mpatches.FancyBboxPatch(
                (x0, y0), w, h,
                boxstyle="square,pad=0",
                linewidth=2, edgecolor=color,
                facecolor=color, alpha=0.15,
            )
            ax.add_patch(rect)
            ax.set_xlim(x0 - pad_x, x1 + pad_x)
            ax.set_ylim(y0 - pad_y, y1 + pad_y)

            # Label corners
            for cx, cy, lbl in [
                (x0, y0, f"IL{hdr.get('min_inline','?')}\nXL{hdr.get('min_crossline','?')}"),
                (x1, y0, f"IL{hdr.get('max_inline','?')}\nXL{hdr.get('min_crossline','?')}"),
                (x0, y1, f"IL{hdr.get('min_inline','?')}\nXL{hdr.get('max_crossline','?')}"),
                (x1, y1, f"IL{hdr.get('max_inline','?')}\nXL{hdr.get('max_crossline','?')}"),
            ]:
                ax.plot(cx, cy, "o", color=color, markersize=6)
                ax.annotate(lbl, (cx, cy), textcoords="offset points",
                            xytext=(6, 4), fontsize=7, color=color)

        else:  # 2D — draw as a line between two endpoints
            ax.plot([x0, x1], [y0, y1], "-o", color=color,
                    linewidth=2.5, markersize=7)
            ax.annotate("Start", (x0, y0), textcoords="offset points",
                        xytext=(6, 4), fontsize=8, color=color, fontweight="bold")
            ax.annotate("End",   (x1, y1), textcoords="offset points",
                        xytext=(6, 4), fontsize=8, color=color, fontweight="bold")
            ax.set_xlim(x0 - pad_x, x1 + pad_x)
            ax.set_ylim(y0 - pad_y, y1 + pad_y)

        ax.set_xlabel(xlabel, fontsize=9)
        ax.set_ylabel(ylabel, fontsize=9)
        ax.grid(True, linestyle="--", alpha=0.4, linewidth=0.5)
        ax.set_facecolor("#f8f9fb")
        ax.tick_params(labelsize=8)

    colours = ["#1F4E79", "#2CA02C"]
    idx = 0

    if has_geo and ax_list:
        _draw_bbox(
            ax_list[idx],
            hdr["min_lon"], hdr["max_lon"],
            hdr["min_lat"], hdr["max_lat"],
            "Longitude (°E)", "Latitude (°N)",
            colours[idx], dim, hdr
        )
        ax_list[idx].set_title("Geographic (lat/lon)", fontsize=10, pad=6)
        idx += 1

    if has_proj and idx < len(ax_list):
        _draw_bbox(
            ax_list[idx],
            hdr["min_x"], hdr["max_x"],
            hdr["min_y"], hdr["max_y"],
            "Easting (m)", "Northing (m)",
            colours[idx], dim, hdr
        )
        ax_list[idx].set_title("Projected coordinates", fontsize=10, pad=6)

    sup = title or (hdr.get("survey_name") or hdr.get("line_name") or "SEG-Y Survey")
    sup += f"  [{dim}]"
    if dim == "3D" and hdr.get("min_inline") is not None:
        sup += f"  Inlines {hdr['min_inline']}–{hdr['max_inline']}  Xlines {hdr['min_crossline']}–{hdr['max_crossline']}"

    fig.suptitle(sup, fontsize=11, y=1.01, fontweight="bold", color="#1F4E79")
    plt.tight_layout()
    st.pyplot(fig, use_container_width=True)
    plt.close(fig)


def _render_p190_map(hdr: dict, title: str = ""):
    """Render a track map for P190 positioning data."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
    except ImportError:
        st.warning(
            "**matplotlib** is not installed. Run: "
            "`pip install matplotlib` then restart Data Wrangler."
        )
        return

    has_geo  = hdr.get("min_lat") is not None
    has_proj = hdr.get("min_x")   is not None

    if not has_geo and not has_proj:
        st.info("No coordinate data found in this P190 file.")
        return

    n_panels = 2 if (has_geo and has_proj) else 1
    fig, axes = plt.subplots(1, n_panels, figsize=(12 if n_panels == 2 else 6, 5))
    if n_panels == 1:
        ax_list = [axes]
    else:
        ax_list = list(axes)

    def _draw_line(ax, x0, x1, y0, y1, xlabel, ylabel, color, hdr):
        w = x1 - x0
        h = y1 - y0
        pad_x = max(w * 0.15, 0.001)
        pad_y = max(h * 0.15, 0.001)

        ax.plot([x0, x1], [y0, y1], "-o", color=color,
                linewidth=2.5, markersize=7)

        first_sp = hdr.get("first_shot_point")
        last_sp  = hdr.get("last_shot_point")
        ax.annotate(f"SP {first_sp}" if first_sp else "Start",
                    (x0, y0), textcoords="offset points",
                    xytext=(6, 4), fontsize=8, color=color, fontweight="bold")
        ax.annotate(f"SP {last_sp}" if last_sp else "End",
                    (x1, y1), textcoords="offset points",
                    xytext=(6, 4), fontsize=8, color=color, fontweight="bold")
        ax.set_xlim(x0 - pad_x, x1 + pad_x)
        ax.set_ylim(y0 - pad_y, y1 + pad_y)
        ax.set_xlabel(xlabel, fontsize=9)
        ax.set_ylabel(ylabel, fontsize=9)
        ax.grid(True, linestyle="--", alpha=0.4, linewidth=0.5)
        ax.set_facecolor("#f8f9fb")
        ax.tick_params(labelsize=8)

    colours = ["#D62728", "#FF7F0E"]
    idx = 0

    if has_geo and ax_list:
        _draw_line(ax_list[idx],
                   hdr["min_lon"], hdr["max_lon"],
                   hdr["min_lat"], hdr["max_lat"],
                   "Longitude (°E)", "Latitude (°N)",
                   colours[idx], hdr)
        ax_list[idx].set_title("Geographic (lat/lon)", fontsize=10, pad=6)
        idx += 1

    if has_proj and idx < len(ax_list):
        _draw_line(ax_list[idx],
                   hdr["min_x"], hdr["max_x"],
                   hdr["min_y"], hdr["max_y"],
                   "Easting (m)", "Northing (m)",
                   colours[idx], hdr)
        ax_list[idx].set_title("Projected coordinates", fontsize=10, pad=6)

    sup = title or (hdr.get("survey_name") or hdr.get("line_name") or "P190 Line")
    shots = hdr.get("shot_count", 0)
    if shots:
        sup += f"  [{shots:,} shot points]"
    fig.suptitle(sup, fontsize=11, y=1.01, fontweight="bold", color="#D62728")
    plt.tight_layout()
    st.pyplot(fig, use_container_width=True)
    plt.close(fig)


def _safe_num(v):
    """Convert a value to float or None — for building map_hdr from search results."""
    try:
        f = float(v)
        return None if (f != f) else f  # NaN check
    except (TypeError, ValueError):
        return None


# ─────────────────────────────────────────────────────────────────────────────
# SURVEYS TAB
# ─────────────────────────────────────────────────────────────────────────────

def _render_seis_surveys(engine):
    """
    Manage survey groupings for catalogued seismic files.
    Allows renaming survey names, reassigning lines to surveys,
    and creating new survey groups from ungrouped files.
    """
    from sqlalchemy import text

    st.subheader("Manage Surveys")
    st.caption(
        "Group seismic lines into surveys. Survey names are stored in SURVEY_NAME "
        "on each file record — no separate survey table is required."
    )

    # Load all catalogued seismic files
    try:
        with engine.connect() as con:
            rows = con.execute(text(
                "SELECT SEIS_FILE_ID, FILE_FORMAT, SURVEY_NAME, LINE_NAME, "
                "DIMENSIONALITY, FILE_NAME, FILE_SIZE_KB "
                "FROM [las_catalog].[SEIS_FILE_CATALOG] "
                "ORDER BY FILE_FORMAT, SURVEY_NAME, LINE_NAME"
            )).fetchall()
    except Exception as e:
        st.error(str(e))
        return

    if not rows:
        st.info("No seismic files catalogued yet.")
        return

    df = pd.DataFrame(rows, columns=[
        "SEIS_FILE_ID", "FORMAT", "SURVEY", "LINE", "DIM", "FILE_NAME", "SIZE_KB"
    ])
    df["SURVEY"] = df["SURVEY"].fillna("").str.strip()
    df["_display_survey"] = df["SURVEY"].where(df["SURVEY"] != "", "— ungrouped —")

    # ── Summary ──────────────────────────────────────────────────────────
    surveys    = sorted(df[df["SURVEY"] != ""]["SURVEY"].unique().tolist())
    ungrouped  = df[df["SURVEY"] == ""]
    mc1, mc2, mc3 = st.columns(3)
    mc1.metric("Total files",   len(df))
    mc2.metric("Surveys",       len(surveys))
    mc3.metric("Ungrouped",     len(ungrouped))

    st.divider()

    sub_rename, sub_assign, sub_new = st.tabs([
        "✏️ Rename survey", "🔀 Reassign lines", "➕ Create survey"
    ])

    # ── Rename survey ─────────────────────────────────────────────────────
    with sub_rename:
        st.markdown("**Rename an existing survey**")
        if not surveys:
            st.info("No named surveys yet.")
        else:
            sel_survey = st.selectbox("Select survey", surveys, key="srv_rename_sel")
            n_lines = len(df[df["SURVEY"] == sel_survey])
            st.caption(f"{n_lines} line(s) in this survey")

            new_name = st.text_input("New survey name", value=sel_survey,
                                      key="srv_rename_new")
            if st.button("Rename", type="primary", key="srv_rename_btn"):
                if not new_name.strip():
                    st.error("Name cannot be empty.")
                elif new_name.strip() == sel_survey:
                    st.warning("Name is unchanged.")
                else:
                    try:
                        with engine.begin() as con:
                            con.execute(text("""
                                UPDATE [las_catalog].[SEIS_FILE_CATALOG]
                                SET SURVEY_NAME = :new,
                                    ROW_CHANGED_BY = 'DATA_WRANGLER',
                                    ROW_CHANGED_DATE = GETUTCDATE()
                                WHERE SURVEY_NAME = :old
                            """), {"new": new_name.strip(), "old": sel_survey})
                        st.success(f"✅ Renamed **{sel_survey}** → **{new_name.strip()}**")
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))

    # ── Reassign lines ────────────────────────────────────────────────────
    with sub_assign:
        st.markdown("**Move lines to a different survey**")

        # Source: pick a survey (or ungrouped)
        src_options = ["— ungrouped —"] + surveys
        src_survey  = st.selectbox("From survey", src_options, key="srv_src_sel")

        if src_survey == "— ungrouped —":
            src_df = df[df["SURVEY"] == ""]
        else:
            src_df = df[df["SURVEY"] == src_survey]

        if src_df.empty:
            st.info("No files in this group.")
        else:
            # Multi-select lines
            line_opts = (src_df["LINE"].fillna(src_df["FILE_NAME"])).tolist()
            sel_lines = st.multiselect(
                "Select lines to move",
                options=line_opts,
                default=line_opts,
                key="srv_lines_sel"
            )

            # Target survey — existing or new
            target_options = [s for s in surveys if s != src_survey] + ["[ new survey… ]"]
            tgt_choice = st.selectbox("To survey", target_options, key="srv_tgt_sel")                          if target_options else None

            if tgt_choice == "[ new survey… ]" or not target_options:
                tgt_name = st.text_input("New survey name", key="srv_tgt_new")
            else:
                tgt_name = tgt_choice

            if st.button("🔀 Reassign", type="primary", key="srv_assign_btn"):
                if not tgt_name or not tgt_name.strip():
                    st.error("Target survey name is required.")
                elif not sel_lines:
                    st.error("Select at least one line.")
                else:
                    # Get file IDs for selected lines
                    sel_ids = src_df[
                        src_df["LINE"].fillna(src_df["FILE_NAME"]).isin(sel_lines)
                    ]["SEIS_FILE_ID"].tolist()
                    try:
                        for fid in sel_ids:
                            with engine.begin() as con:
                                con.execute(text("""
                                    UPDATE [las_catalog].[SEIS_FILE_CATALOG]
                                    SET SURVEY_NAME = :sv,
                                        ROW_CHANGED_BY = 'DATA_WRANGLER',
                                        ROW_CHANGED_DATE = GETUTCDATE()
                                    WHERE SEIS_FILE_ID = :id
                                """), {"sv": tgt_name.strip(), "id": fid})
                        st.success(
                            f"✅ Moved {len(sel_ids)} line(s) → "
                            f"**{tgt_name.strip()}**"
                        )
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))

    # ── Create survey from ungrouped ──────────────────────────────────────
    with sub_new:
        st.markdown("**Create a new survey from ungrouped files**")
        if ungrouped.empty:
            st.success("All files are assigned to a survey. ✅")
        else:
            st.info(f"{len(ungrouped)} ungrouped file(s):")
            st.dataframe(
                ungrouped[["FORMAT","LINE","FILE_NAME","DIM","SIZE_KB"]],
                hide_index=True, use_container_width=True
            )

            new_survey_name = st.text_input(
                "Survey name for selected files",
                key="srv_new_name",
                placeholder="e.g. North Sea 2D 1989"
            )
            sel_new = st.multiselect(
                "Select files to assign",
                options=ungrouped["FILE_NAME"].tolist(),
                default=ungrouped["FILE_NAME"].tolist(),
                key="srv_new_sel"
            )

            if st.button("➕ Create survey", type="primary", key="srv_new_btn"):
                if not new_survey_name.strip():
                    st.error("Survey name is required.")
                elif not sel_new:
                    st.error("Select at least one file.")
                else:
                    ids = ungrouped[
                        ungrouped["FILE_NAME"].isin(sel_new)
                    ]["SEIS_FILE_ID"].tolist()
                    try:
                        for fid in ids:
                            with engine.begin() as con:
                                con.execute(text("""
                                    UPDATE [las_catalog].[SEIS_FILE_CATALOG]
                                    SET SURVEY_NAME = :sv,
                                        ROW_CHANGED_BY = 'DATA_WRANGLER',
                                        ROW_CHANGED_DATE = GETUTCDATE()
                                    WHERE SEIS_FILE_ID = :id
                                """), {"sv": new_survey_name.strip(), "id": fid})
                        st.success(
                            f"✅ Created survey **{new_survey_name.strip()}** "
                            f"with {len(ids)} file(s)."
                        )
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))

    # ── Current survey listing ────────────────────────────────────────────
    st.divider()
    st.markdown("**Current survey assignments**")
    summary = df.groupby("_display_survey").agg(
        Lines=("SEIS_FILE_ID", "count"),
        Formats=("FORMAT", lambda x: ", ".join(sorted(x.unique()))),
        Size_MB=("SIZE_KB", lambda x: round(x.sum() / 1024, 1))
    ).reset_index().rename(columns={"_display_survey": "Survey"})
    st.dataframe(summary, hide_index=True, use_container_width=True)


def _render_seis_search(engine):
    from sqlalchemy import text

    st.markdown("**Search seismic file catalog**")

    col1, col2, col3 = st.columns(3)
    with col1:
        fmt_filter = st.selectbox("Format", ["All", "SEGY", "P190"],
                                   key="seis_search_fmt")
    with col2:
        survey_filter = st.text_input("Survey name (partial)", key="seis_search_survey")
    with col3:
        line_filter = st.text_input("Line name (partial)", key="seis_search_line")

    if st.button("🔍 Search", type="primary", key="seis_search_btn"):
        where  = ["1=1"]
        params = {}
        if fmt_filter != "All":
            where.append("FILE_FORMAT = :fmt"); params["fmt"] = fmt_filter
        if survey_filter:
            where.append("SURVEY_NAME LIKE :sv"); params["sv"] = f"%{survey_filter}%"
        if line_filter:
            where.append("LINE_NAME LIKE :ln"); params["ln"] = f"%{line_filter}%"
        try:
            with engine.connect() as con:
                rows = con.execute(text(
                    f"SELECT f.SEIS_FILE_ID, f.FILE_FORMAT, f.SURVEY_NAME, f.LINE_NAME, "
                    f"f.DIMENSIONALITY, f.VESSEL_NAME, f.CLIENT_NAME, "
                    f"f.FIRST_SHOT_POINT, f.LAST_SHOT_POINT, f.SHOT_COUNT, "
                    f"f.TRACE_COUNT, f.SAMPLE_COUNT, f.SAMPLE_INTERVAL_US, "
                    f"f.NAV_SYSTEM, f.COORD_SYSTEM, "
                    f"CONVERT(NVARCHAR(30), f.ACQ_DATE_START) AS ACQ_DATE_START, "
                    f"CONVERT(NVARCHAR(30), f.ACQ_DATE_END)   AS ACQ_DATE_END, "
                    f"f.MIN_LAT, f.MAX_LAT, f.MIN_LON, f.MAX_LON, "
                    f"f.MIN_X, f.MAX_X, f.MIN_Y, f.MAX_Y, "
                    f"f.MIN_INLINE, f.MAX_INLINE, f.MIN_CROSSLINE, f.MAX_CROSSLINE, "
                    f"f.FILE_SIZE_KB, f.FILE_NAME, f.SEIS_SET_ID, f.SEIS_LINE_ID, "
                    f"CASE WHEN r.BASE_PATH IS NULL THEN f.FILE_NAME "
                    f"     WHEN RIGHT(r.BASE_PATH,1)=\'\\\' THEN r.BASE_PATH + f.FILE_NAME "
                    f"     ELSE r.BASE_PATH + \'\\\' + f.FILE_NAME END AS FULL_PATH "
                    f"FROM [las_catalog].[SEIS_FILE_CATALOG] f "
                    f"LEFT JOIN [las_catalog].[WL_REPOSITORY] r "
                    f"  ON r.REPOSITORY_ID = f.REPOSITORY_ID "
                    f"WHERE {' AND '.join(where)} "
                    f"ORDER BY f.FILE_FORMAT, f.SURVEY_NAME, f.LINE_NAME"
                ), params).fetchall()

            if not rows:
                st.info("No results found.")
                st.session_state.pop("seis_search_df", None)
                st.stop()

            df = pd.DataFrame(rows, columns=[
                "SEIS_FILE_ID", "FORMAT", "SURVEY", "LINE", "DIM",
                "VESSEL", "CLIENT", "FIRST_SP", "LAST_SP", "SHOT_COUNT",
                "TRACE_COUNT", "SAMPLE_COUNT", "SI_US", "NAV_SYSTEM", "COORD_SYSTEM",
                "ACQ_DATE_START", "ACQ_DATE_END",
                "MIN_LAT", "MAX_LAT", "MIN_LON", "MAX_LON",
                "MIN_X", "MAX_X", "MIN_Y", "MAX_Y",
                "MIN_INLINE", "MAX_INLINE", "MIN_CROSSLINE", "MAX_CROSSLINE",
                "SIZE_KB", "FILE_NAME", "SEIS_SET_ID", "SEIS_LINE_ID", "FULL_PATH"
            ])
            df["PPDM_SEEDED"] = df["SEIS_SET_ID"].notna().map({True: "✅", False: "—"})
            # Store results in session state so they persist across reruns
            st.session_state["seis_search_df"] = df

        except Exception as e:
            st.error(str(e))
            return

    # ── Results — rendered outside button block so they survive reruns ────
    if "seis_search_df" not in st.session_state:
        return

    df = st.session_state["seis_search_df"]

    st.success(f"{len(df)} file(s) found.")

    # Summary metrics
    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("P190 files",  len(df[df["FORMAT"] == "P190"]))
    mc2.metric("SEG-Y files", len(df[df["FORMAT"] == "SEGY"]))
    mc3.metric("PPDM seeded",  df["SEIS_SET_ID"].notna().sum())
    mc4.metric("Not seeded",   df["SEIS_SET_ID"].isna().sum())

    # Build a clean display frame — use FILE_NAME as line if LINE is empty
    disp = df.copy()
    disp["LINE"] = disp["LINE"].fillna("").str.strip()
    disp["LINE"] = disp.apply(
        lambda r: r["LINE"] if r["LINE"] else r["FILE_NAME"], axis=1
    )
    # Show X/Y when LAT/LON absent
    has_geo  = df["MIN_LAT"].notna().any()
    has_proj = df["MIN_X"].notna().any()
    coord_cols = []
    if has_geo:  coord_cols += ["MIN_LAT", "MAX_LAT", "MIN_LON", "MAX_LON"]
    if has_proj: coord_cols += ["MIN_X", "MAX_X", "MIN_Y", "MAX_Y"]

    display_cols = [c for c in [
        "FORMAT", "SURVEY", "LINE", "DIM",
        "FIRST_SP", "LAST_SP", "SHOT_COUNT", "TRACE_COUNT",
        "ACQ_DATE_START",
    ] + coord_cols + ["SIZE_KB", "PPDM_SEEDED", "FILE_NAME"]
    if c in disp.columns]
    st.dataframe(disp[display_cols], use_container_width=True, hide_index=True)

    # ── Copy / Export ─────────────────────────────────────────────────────
    st.divider()
    st.markdown("#### 📂 Copy files to folder")
    c_dest, c_ow = st.columns([4, 1])
    seis_dest = c_dest.text_input(
        "Destination folder", key="seis_export_dest",
        placeholder=r"e.g. C:\Downloads\Seismic"
    )
    seis_ow = c_ow.checkbox("Overwrite", value=False, key="seis_export_ow")

    c_btn_copy, c_btn_csv = st.columns([1, 1])
    if c_btn_copy.button(
        f"📋 Copy {len(df)} file(s)", type="primary",
        key="seis_copy_btn", use_container_width=True,
        disabled=not seis_dest.strip()
    ):
        import shutil, os
        os.makedirs(seis_dest.strip(), exist_ok=True)
        copied = skipped = missing = errors_c = 0
        details = []
        for _, row in df.iterrows():
            src = str(row.get("FULL_PATH") or "").strip()
            if not src or src in ("", "None", "nan"):
                missing += 1
                details.append({"File": row["FILE_NAME"], "Status": "❌ No path stored"})
                continue
            from pathlib import Path as _P
            src_p = _P(src)
            if not src_p.exists():
                missing += 1
                details.append({"File": src_p.name, "Status": "❌ Not found"})
                continue
            dst_p = _P(seis_dest.strip()) / src_p.name
            if dst_p.exists() and not seis_ow:
                skipped += 1
                details.append({"File": src_p.name, "Status": "⏭ Skipped (exists)"})
                continue
            try:
                shutil.copy2(str(src_p), str(dst_p))
                copied += 1
                details.append({"File": src_p.name, "Status": "✅ Copied"})
            except Exception as ce:
                errors_c += 1
                details.append({"File": src_p.name, "Status": f"❌ {ce}"})
        if errors_c == 0 and missing == 0:
            st.success(f"✅ {copied} copied · {skipped} skipped")
        else:
            st.warning(f"{copied} copied · {skipped} skipped · {missing} not found · {errors_c} error(s)")
        with st.expander("Copy details", expanded=errors_c > 0):
            st.dataframe(pd.DataFrame(details), hide_index=True, use_container_width=True)

    with c_btn_csv:
        st.download_button(
            label="⬇ Export Summary CSV",
            data=df.to_csv(index=False),
            file_name="seis_catalog_results.csv",
            mime="text/csv",
            key="seis_csv_btn",
        )

    # ── Survey / line viewer ──────────────────────────────────────────────
    st.divider()
    st.markdown("**View survey map**")

    # Build survey list — group by SURVEY name, fall back to FILE_NAME
    df["_survey_key"] = df["SURVEY"].fillna("").str.strip()
    df["_survey_key"] = df["_survey_key"].where(df["_survey_key"] != "", df["FILE_NAME"])

    surveys = sorted(df["_survey_key"].unique().tolist())

    sel_survey = st.selectbox(
        "Select survey",
        options=["— select —"] + surveys,
        key="seis_survey_sel"
    )

    if sel_survey and sel_survey != "— select —":
        survey_df = df[df["_survey_key"] == sel_survey]
        fmt = str(survey_df["FORMAT"].iloc[0]).upper()
        dim = str(survey_df["DIM"].iloc[0]) if "DIM" in survey_df.columns else "2D"
        n_lines = len(survey_df)

        st.caption(
            f"{fmt} · {dim} · {n_lines} line(s) · "
            f"Survey: **{sel_survey}**"
        )

        # ── Multi-line map ────────────────────────────────────────────────
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            import matplotlib.patches as mpatches

            # Colour palette — one per line
            _COLOURS = [
                "#1F4E79","#D62728","#2CA02C","#FF7F0E","#9467BD",
                "#8C564B","#E377C2","#7F7F7F","#BCBD22","#17BECF",
            ]

            # Check what coordinate systems are available
            has_geo  = survey_df["MIN_LAT"].notna().any()
            has_proj = survey_df["MIN_X"].notna().any()

            if not has_geo and not has_proj:
                st.info("No coordinate data available for this survey.")
            else:
                n_panels = 2 if (has_geo and has_proj) else 1
                fig, axes = plt.subplots(
                    1, n_panels, figsize=(13 if n_panels == 2 else 6.5, 5)
                )
                ax_list = [axes] if n_panels == 1 else list(axes)

                legend_handles = []

                for li, (_, row) in enumerate(survey_df.iterrows()):
                    colour   = _COLOURS[li % len(_COLOURS)]
                    line_lbl = str(row.get("LINE", "") or row.get("FILE_NAME", f"Line {li+1}"))
                    row_dim  = str(row.get("DIM", "2D"))

                    def _draw_line_or_box(ax, x0, x1, y0, y1, colour, row_dim, line_lbl, li):
                        if x0 is None or x1 is None or y0 is None or y1 is None:
                            return
                        if row_dim == "3D":
                            w = x1 - x0; h = y1 - y0
                            rect = mpatches.FancyBboxPatch(
                                (x0, y0), w, h,
                                boxstyle="square,pad=0",
                                linewidth=2, edgecolor=colour,
                                facecolor=colour, alpha=0.12,
                            )
                            ax.add_patch(rect)
                            # Corner dots
                            for cx, cy in [(x0,y0),(x1,y0),(x0,y1),(x1,y1)]:
                                ax.plot(cx, cy, "o", color=colour, markersize=5)
                        else:
                            ax.plot([x0, x1], [y0, y1], "-o",
                                    color=colour, linewidth=2, markersize=6,
                                    label=line_lbl)
                            # Label endpoints for first and last line only
                            if li == 0:
                                ax.annotate(
                                    line_lbl, (x0, y0),
                                    textcoords="offset points", xytext=(5, 4),
                                    fontsize=7, color=colour
                                )
                            ax.annotate(
                                line_lbl, (x1, y1),
                                textcoords="offset points", xytext=(5, 4),
                                fontsize=7, color=colour
                            )

                    _ax_idx = 0
                    if has_geo:
                        _draw_line_or_box(
                            ax_list[_ax_idx],
                            _safe_num(row.get("MIN_LON")), _safe_num(row.get("MAX_LON")),
                            _safe_num(row.get("MIN_LAT")), _safe_num(row.get("MAX_LAT")),
                            colour, row_dim, line_lbl, li
                        )
                        _ax_idx += 1
                    if has_proj:
                        _draw_line_or_box(
                            ax_list[_ax_idx],
                            _safe_num(row.get("MIN_X")), _safe_num(row.get("MAX_X")),
                            _safe_num(row.get("MIN_Y")), _safe_num(row.get("MAX_Y")),
                            colour, row_dim, line_lbl, li
                        )
                    legend_handles.append(
                        mpatches.Patch(color=colour, label=line_lbl)
                    )

                # Format axes
                _panel_labels = []
                if has_geo:  _panel_labels.append(("Longitude (°E)", "Latitude (°N)", "Geographic (lat/lon)"))
                if has_proj: _panel_labels.append(("Easting (m)", "Northing (m)", "Projected coordinates"))

                for ax, (xl, yl, title) in zip(ax_list, _panel_labels):
                    ax.autoscale()
                    ax.margins(0.12)
                    ax.set_xlabel(xl, fontsize=9)
                    ax.set_ylabel(yl, fontsize=9)
                    ax.set_title(title, fontsize=10, pad=6)
                    ax.grid(True, linestyle="--", alpha=0.4, linewidth=0.5)
                    ax.set_facecolor("#f8f9fb")
                    ax.tick_params(labelsize=8)

                # Legend — only if multiple lines
                if n_lines > 1 and n_lines <= 20:
                    fig.legend(
                        handles=legend_handles,
                        loc="lower center",
                        ncol=min(n_lines, 5),
                        fontsize=8,
                        framealpha=0.8,
                        bbox_to_anchor=(0.5, -0.05)
                    )

                fig.suptitle(
                    f"{sel_survey}  [{fmt} · {dim} · {n_lines} line(s)]",
                    fontsize=11, fontweight="bold", color="#1F4E79", y=1.01
                )
                plt.tight_layout()
                st.pyplot(fig, use_container_width=True)
                plt.close(fig)

        except Exception as _map_err:
            st.warning(f"Map could not be rendered: {_map_err}")

        # ── Per-line header viewer ────────────────────────────────────────
        st.divider()
        st.markdown("**View header for line**")

        line_opts = survey_df["LINE"].fillna(survey_df["FILE_NAME"]).tolist()
        sel_line  = st.selectbox(
            "Select line",
            options=["— select —"] + line_opts,
            key="seis_line_hdr_sel"
        )

        if sel_line and sel_line != "— select —":
            line_row = survey_df[
                (survey_df["LINE"].fillna(survey_df["FILE_NAME"])) == sel_line
            ].iloc[0]
            file_id  = str(line_row.get("SEIS_FILE_ID", ""))
            sel_path = str(line_row.get("FULL_PATH", ""))

            _hdr_lines  = []
            _hdr_source = ""

            if file_id:
                try:
                    with engine.connect() as _con:
                        _rows = _con.execute(text(
                            "SELECT LINE_NO, HEADER_TEXT "
                            "FROM [las_catalog].[SEIS_FILE_HEADER] "
                            "WHERE SEIS_FILE_ID = :id ORDER BY LINE_NO"
                        ), {"id": file_id}).fetchall()
                    _hdr_lines  = [r[1] for r in _rows if r[1]]
                    _hdr_source = "catalog"
                except Exception as _e:
                    st.warning(f"Could not load from catalog: {_e}")

            if not _hdr_lines:
                try:
                    import os as _os
                    if _os.path.exists(sel_path):
                        if fmt == "SEGY":
                            _ph = parse_segy_header(sel_path)
                            _hdr_lines  = _ph.get("text_header", [])
                            _hdr_source = "file"
                        elif fmt == "P190":
                            _ph = parse_p190_header(sel_path)
                            _hdr_lines  = [
                                f"Record type '{k}': {v} records"
                                for k, v in sorted(_ph["record_type_counts"].items())
                            ]
                            _hdr_source = "file"
                    else:
                        st.info("File not accessible from this machine.")
                except Exception as _e:
                    st.warning(f"Could not parse file: {_e}")

            if _hdr_lines:
                if fmt == "SEGY":
                    _all = "\n".join(_hdr_lines)
                    _pct = sum(1 for c in _all if 32 <= ord(c) <= 126) / max(len(_all), 1)
                    _icon = "✅" if _pct > 0.8 else ("⚠️" if _pct > 0.4 else "❌")
                    with st.expander(
                        f"{_icon} EBCDIC Text Header — {sel_line} "
                        f"({_pct*100:.0f}% readable · {_hdr_source})",
                        expanded=_pct > 0.5
                    ):
                        st.code(_all, language=None)
                elif fmt == "P190":
                    with st.expander(
                        f"📄 P190 Record Summary — {sel_line} · {_hdr_source}",
                        expanded=True
                    ):
                        for _ln in _hdr_lines:
                            st.text(_ln)

                # ── Section viewer for SEG-Y lines ────────────────────────
                if fmt == "SEGY":
                    import os as _os
                    sel_full_path = str(line_row.get("FULL_PATH", ""))
                    _hdr_dict = {
                        "survey_name":        str(line_row.get("SURVEY", "")),
                        "line_name":          sel_line,
                        "sample_count":       line_row.get("SAMPLE_COUNT") or line_row.get("TRACE_COUNT"),
                        "sample_interval_us": line_row.get("SI_US") or line_row.get("SAMPLE_INTERVAL_US"),
                        "dimensionality":     str(line_row.get("DIM", "2D")),
                        "min_inline":         line_row.get("MIN_INLINE"),
                    }
                    if sel_full_path and _os.path.exists(sel_full_path):
                        _render_segy_section_viewer(sel_full_path, _hdr_dict)
                    else:
                        st.divider()
                        st.markdown("**📈 Seismic section viewer**")
                        st.warning(
                            f"File not found at catalogued path:\n\n"
                            f"`{sel_full_path}`\n\n"
                            "Enter the current file path below:"
                        )
                        _override_path = st.text_input(
                            "File path", key=f"seis_path_override_{sel_line}",
                            placeholder=r"e.g. C:\Seismic\splee-2d-1982-t-pstm-line-201.sgy"
                        )
                        if _override_path and _os.path.exists(_override_path):
                            _render_segy_section_viewer(_override_path, _hdr_dict)
                        elif _override_path:
                            st.error(f"File not found: `{_override_path}`")

    # ── Actions ───────────────────────────────────────────────────────────
    st.divider()
    col_csv, col_recatalog, col_seed = st.columns([1, 1, 2])
    with col_csv:
        st.download_button(
            "⬇ Export CSV",
            data=df.to_csv(index=False),
            file_name="seis_catalog_results.csv",
            mime="text/csv",
            key="seis_dl_btn"
        )
    with col_recatalog:
        # Count files missing both geo and projected coordinates
        def _is_missing_coords(row):
            def _null(v):
                return v is None or (hasattr(v, '__class__') and
                       v.__class__.__name__ in ('NoneType','NAType')) or                        str(v).strip() in ('', 'None', 'nan', 'NaN', '<NA>')
            return _null(row.get("MIN_X")) and _null(row.get("MIN_LAT"))

        missing_coords = df[df.apply(_is_missing_coords, axis=1)]
        n_missing = len(missing_coords)
        btn_label = (
            f"🔄 Re-catalog {n_missing} file(s) missing coords"
            if n_missing > 0 else "🔄 Re-catalog all"
        )
        recatalog_target = missing_coords if n_missing > 0 else df
        if st.button(
            btn_label,
            key="seis_recatalog_btn",
            help="Re-parse files using the updated parser to populate coordinates, "
                 "survey names and other metadata."
        ):
                from modules.segy_catalog  import catalog_segy_file
                from modules.p190_catalog  import catalog_p190_file
                prog = st.progress(0, text="Re-cataloguing…")
                errors = []
                done  = 0
                for _, row in recatalog_target.iterrows():
                    fp  = str(row.get("FULL_PATH",""))
                    fmt = str(row.get("FORMAT","")).upper()
                    prog.progress(
                        (done+1)/len(recatalog_target),
                        text=f"Re-cataloguing {row.get('FILE_NAME','')}…"
                    )
                    try:
                        import os as _os
                        if _os.path.exists(fp):
                            # Catalog will do a full UPDATE if record exists
                            fn = catalog_p190_file if fmt == "P190" else catalog_segy_file
                            r  = fn(engine, fp)
                            if not r.get("ok"):
                                errors.append(f"{row['FILE_NAME']}: {r.get('error','')}")
                        else:
                            errors.append(f"{row['FILE_NAME']}: file not accessible")
                    except Exception as _e:
                        errors.append(f"{row['FILE_NAME']}: {_e}")
                    done += 1
                prog.empty()
                if errors:
                    with st.expander(f"❌ {len(errors)} error(s)"):
                        for e in errors: st.text(e)
                else:
                    st.success(f"✅ Re-catalogued {len(recatalog_target)} file(s).")
                st.session_state.pop("seis_search_df", None)
                st.rerun()
    with col_seed:
        unseeded = df[df["SEIS_SET_ID"].isna()]
        if not unseeded.empty:
            if st.button(
                f"🌱 Seed PPDM for {len(unseeded)} unseeded file(s)",
                key="seis_seed_btn",
                help="Insert dbo.SEIS_SET and dbo.SEIS_LINE for files not yet linked to PPDM"
            ):
                from modules.segy_catalog import catalog_segy_file
                from modules.p190_catalog import catalog_p190_file
                seeded_count = 0
                errors = []
                prog = st.progress(0, text="Seeding PPDM…")
                for _i, (_, urow) in enumerate(unseeded.iterrows()):
                    prog.progress((_i + 1) / len(unseeded),
                                  text=f"Seeding {urow['FILE_NAME']}…")
                    try:
                        fn  = str(urow["FULL_PATH"])
                        fmt = str(urow["FORMAT"]).upper()
                        r = (catalog_p190_file if fmt == "P190" else catalog_segy_file)(
                            engine, fn, seed_ppdm=True
                        )
                        if r["seeded_ppdm"]:
                            seeded_count += 1
                        elif r.get("error"):
                            errors.append(f"{urow['FILE_NAME']}: {r['error']}")
                    except Exception as _e:
                        errors.append(f"{urow['FILE_NAME']}: {_e}")
                prog.empty()
                if seeded_count:
                    st.success(f"✅ {seeded_count} file(s) seeded to PPDM.")
                if errors:
                    with st.expander(f"❌ {len(errors)} error(s)"):
                        for _e in errors:
                            st.text(_e)
                # Refresh results
                st.session_state.pop("seis_search_df", None)
                st.rerun()


def run_landing(engine=None):
    """
    Petroleum File Catalog landing page.
    Shown when user clicks "Std Format Browser" — unified entry point
    showing both well log and seismic catalog search.
    """
    import streamlit as st

    st.title("🛢️ Petroleum File Catalog")
    st.caption(
        "Search and browse cataloged petroleum files — "
        "LAS, DLIS, LIS, SEG-Y and P190."
    )

    if engine is None:
        engine = _get_engine()
    if engine is None:
        st.warning("No database connection. Connect via the main pipeline first.")
        return

    tab_wl, tab_seis = st.tabs([
        "🛢️ Well Logs  (LAS · DLIS · LIS)",
        "🌊 Seismic  (SEG-Y · P190)",
    ])

    with tab_wl:
        try:
            _render_search(engine)
        except Exception as e:
            st.error(f"Well log search error: {e}")

    with tab_seis:
        try:
            _render_seis_search(engine)
        except Exception as e:
            st.error(f"Seismic search error: {e}")

